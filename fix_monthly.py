import openpyxl, re, sys
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
sys.stdout.reconfigure(encoding='utf-8')

FILLS = {
    'match':    (None, None),
    'mismatch': (PatternFill(fill_type='solid', fgColor='FFD7D7'), Font(bold=True, color='CC0000')),
    'adj_only': (PatternFill(fill_type='solid', fgColor='D6E4F7'), None),
    'fallback': (PatternFill(fill_type='solid', fgColor='FFF3CD'), None),
    'none':     (PatternFill(fill_type='solid', fgColor='E8E8E8'), Font(color='888888')),
}

def parse생산구분(val, qty):
    if val is None: return 0, qty
    s = str(val).strip()
    if '수입' in s or 'FLC' in s: return 0, qty
    if s in ('생', '셍'): return qty, 0
    m = re.match(r'^생(\d+)$', s)
    if m:
        n = int(m.group(1))
        return n, max(0, qty - n)
    return 0, qty

def resolve_key(d, 주문no, 품목코드):
    key = f"{주문no}||{품목코드}"
    if key in d: return d.get(key)
    if re.match(r'^\d{15,}$', 주문no) and 주문no.endswith('0'):
        for d2 in range(1, 10):
            alt = f"{주문no[:-1]}{d2}||{품목코드}"
            if alt in d: return d.get(alt)
    return None

def proportional(total_val, qty, total_qty, is_last, already):
    if is_last:
        return max(0, total_val - already)
    return round(total_val * qty / total_qty) if total_qty > 0 else 0

# ── 조정 파일 ──
wb_adj = openpyxl.load_workbook(
    r"C:\Users\AJWorld\todo-app\대외비_OJC 주문건 조정 관련_SCM팀_박정원_260605_02.xlsx",
    data_only=True)
ws_adj = wb_adj['미출하현황 1']
headers = [str(c.value or '').replace('\n','').strip() for c in ws_adj[1]]
idx = {
    '주문NO':   next(i for i,h in enumerate(headers) if '주문NO' in h),
    '품목코드': headers.index('품목코드'),
    '수입출고': next(i for i,h in enumerate(headers) if '수입출고' in h),
    '생산출고': headers.index('생산출고'),
}
adj_map = defaultdict(lambda: [0, 0])
for row in ws_adj.iter_rows(min_row=2, values_only=True):
    주문no = str(row[idx['주문NO']] or '').strip()
    품목코드 = str(row[idx['품목코드']] or '').strip()
    if not 주문no or not 품목코드: continue
    adj_map[f"{주문no}||{품목코드}"][0] += int(row[idx['생산출고']] or 0)
    adj_map[f"{주문no}||{품목코드}"][1] += int(row[idx['수입출고']] or 0)

# ── 출하현황 ──
wb_ship = openpyxl.load_workbook(r"C:\Users\AJWorld\todo-app\P4DTS85219LDA6A.xlsx", data_only=True)
ws_ship = wb_ship.active
sh = [str(c.value or '').strip() for c in list(ws_ship.rows)[1]]
si = {'주문NO': sh.index('주문NO(4)'), '품목코드': sh.index('품목코드'),
      '수량': sh.index('수량'), '생산구분': sh.index('생산구분')}
ship_map = defaultdict(lambda: [0, 0])
for row in ws_ship.iter_rows(min_row=3, values_only=True):
    주문no = str(row[si['주문NO']] or '').strip()
    품목코드 = str(row[si['품목코드']] or '').strip()
    수량 = int(row[si['수량']] or 0)
    if not 주문no or not 품목코드 or 수량 == 0: continue
    m, i = parse생산구분(row[si['생산구분']], 수량)
    ship_map[f"{주문no}||{품목코드}"][0] += m
    ship_map[f"{주문no}||{품목코드}"][1] += i

# ── 수익률 분석 파일 ──
wb_out = openpyxl.load_workbook(
    r"C:\Users\AJWorld\todo-app\비밀_2026년 5월 OJC 판매 수익률 분석_영업관리팀_김한나_260605_작성중.xlsx")
ws_out = wb_out.active
out_h = [str(c.value or '').strip() for c in ws_out[1]]
col = {k: out_h.index(v)+1 for k,v in {
    '주문NO':'주문NO(4)', '품목코드':'품목코드', '수량':'수량',
    '맥산수량':'맥산수량', '수입수량':'수입수량'}.items()}

# 사전 스캔: 키별 총 수량
total_qty_map = defaultdict(int)
for r in range(2, ws_out.max_row+1):
    주문no = str(ws_out.cell(r, col['주문NO']).value or '').strip()
    품목코드 = str(ws_out.cell(r, col['품목코드']).value or '').strip()
    if not 주문no or not 품목코드: continue
    total_qty_map[f"{주문no}||{품목코드}"] += int(ws_out.cell(r, col['수량']).value or 0)

assigned_m = defaultdict(int)
assigned_i = defaultdict(int)
consumed   = defaultdict(int)

비고_col = ws_out.max_column + 1
ws_out.cell(1, 비고_col).value = '검토사항'
ws_out.cell(1, 비고_col).font = Font(bold=True)

counts = defaultdict(int)

for r in range(2, ws_out.max_row+1):
    주문no = str(ws_out.cell(r, col['주문NO']).value or '').strip()
    품목코드 = str(ws_out.cell(r, col['품목코드']).value or '').strip()
    if not 주문no or not 품목코드: continue

    key      = f"{주문no}||{품목코드}"
    a        = resolve_key(adj_map, 주문no, 품목코드)
    s        = resolve_key(ship_map, 주문no, 품목코드)
    수량     = int(ws_out.cell(r, col['수량']).value or 0)
    total_q  = total_qty_map[key] or 수량
    is_last  = consumed[key] + 수량 >= total_q

    if a and s:
        맥산 = proportional(a[0], 수량, total_q, is_last, assigned_m[key])
        수입 = proportional(a[1], 수량, total_q, is_last, assigned_i[key])
        status = 'match' if a[0]==s[0] and a[1]==s[1] else 'mismatch'
        note = '' if status == 'match' else f'조정(맥산{a[0]}/수입{a[1]}) vs 출하(맥산{s[0]}/수입{s[1]})'
    elif a:
        맥산 = proportional(a[0], 수량, total_q, is_last, assigned_m[key])
        수입 = proportional(a[1], 수량, total_q, is_last, assigned_i[key])
        status = 'adj_only'; note = '조정 파일만 (출하현황 없음)'
    elif s:
        맥산 = proportional(s[0], 수량, total_q, is_last, assigned_m[key])
        수입 = proportional(s[1], 수량, total_q, is_last, assigned_i[key])
        status = 'fallback'; note = '출하현황 기준 (조정 파일 없음)'
    else:
        맥산 = 0; 수입 = 0; status = 'none'; note = '미매칭'

    # 행별 상한: 맥산+수입 <= 수량
    맥산 = max(0, min(맥산, 수량))
    수입 = max(0, min(수입, 수량 - 맥산))

    consumed[key] += 수량
    assigned_m[key] += 맥산
    assigned_i[key] += 수입
    counts[status] += 1

    ws_out.cell(r, col['맥산수량']).value = 맥산
    ws_out.cell(r, col['수입수량']).value = 수입
    ws_out.cell(r, 비고_col).value = note

    fill, font = FILLS[status]
    for ci in [col['맥산수량'], col['수입수량'], 비고_col]:
        if fill: ws_out.cell(r, ci).fill = fill
        if font: ws_out.cell(r, ci).font = font

out_path = r"C:\Users\AJWorld\todo-app\OJC_판매수익률분석_완성_260608.xlsx"
wb_out.save(out_path)
print(f"완료: {out_path}")
for s, c in counts.items():
    print(f"  {s}: {c}건")

# 14-K-021 검증
print("\n=== 14-K-021 검증 ===")
wb2 = openpyxl.load_workbook(out_path)
ws2 = wb2.active
h2 = [str(c.value or '').strip() for c in ws2[1]]
cc = {k: h2.index(v)+1 for k,v in {'품목코드':'품목코드','수량':'수량','맥산수량':'맥산수량','수입수량':'수입수량'}.items()}
ts, tm, ti = 0, 0, 0
for r in range(2, ws2.max_row+1):
    if str(ws2.cell(r, cc['품목코드']).value or '').strip() != '14-K-021': continue
    s = int(ws2.cell(r, cc['수량']).value or 0)
    m = int(ws2.cell(r, cc['맥산수량']).value or 0)
    i = int(ws2.cell(r, cc['수입수량']).value or 0)
    ts += s; tm += m; ti += i
    print(f"  수량:{s} 맥산:{m} 수입:{i}")
print(f"  합계 → 수량:{ts} 맥산:{tm} 수입:{ti} (맥산+수입={tm+ti})")
