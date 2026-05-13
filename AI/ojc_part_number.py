"""
OJC 자동 품번 생성 로직
품번 체계: A14-[OJC종류]-[코어종류][심선수][재질]-[커넥터A][커넥터B]-[길이코드][마킹]

나중에 VS Code 내부 서버로 구현 예정.
확인 필요 항목은 TODO 주석으로 표시.
"""

import re
from typing import Optional


# ─── 코드 테이블 ──────────────────────────────────────────────────────────────

# 위치 4: OJC 종류
OJC_TYPE = {
    "A": "Simplex 0.9mm (SOJC)",
    "B": "Simplex 3.0mm",
    "C": "Duplex (DOJC)",
    "D": "MOJC / OJC-C2  0.9mm",
    "E": "MOJC / OJC-C2  2.0mm",
    "F": "Pigtail 0.9mm",
    "G": "Pigtail 2.0mm",
    "H": "Pigtail 3.0mm",
    "I": "Pigtail 성단용",
    "J": "리본케이블",
    "K": "Drop cable 3.0mm",
    "L": "Drop cable 5.0mm",
    "M": "Optical Cable 3.0mm (인입광 / Optical Cable Parts)",
    "N": "FLAT 3mm×2mm",
}

# 위치 5: 코어 종류
CORE_TYPE = {
    "A": "SMF 652D",
    "B": "SMF 657.A1",
    "C": "SMF 657.A2",
    "D": "SMF 657.B3",
    "E": "OM1",
    "F": "OM2",
    "G": "OM3",
    "H": "OM4",
    "I": "OM5",
}

# 위치 6: 심선수  ← 실제 품번체계 기준 (3심·5심·9심·11심 코드 없음)
CORE_COUNT = {
    "1": 1,
    "2": 2,
    "3": 4,
    "4": 6,
    "5": 7,
    "6": 8,
    "7": 10,
    "8": 12,
    "9": 13,
    "A": 14,
    "B": 15,
    "C": 16,
    "D": 17,
    "E": 18,
    "F": 20,
    "G": 22,
    "H": 24,
    "I": 36,
    "J": 48,
}
CORE_COUNT_INV = {v: k for k, v in CORE_COUNT.items()}  # 심수 → 코드

# 위치 7: 재질
MATERIAL = {
    "0": "PVC YELLOW",
    "1": "PVC BLUE",
    "2": "PVC AQUA",
    "3": "PVC ORANGE",
    "4": "PVC BLACK",
    "5": "PVC WHITE",
    "6": "PVC 2색",
    "7": "PVC 4색",
    "8": "PVC 6색",
    "9": "PVC 8색",
    "A": "PVC 12색",
    "B": "LSZH BLACK",
    "C": "LSZH YELLOW",
    "D": "PU YELLOW",
    "E": "PU BLUE",
    "F": "PU AQUA",
    "G": "PU ORANGE",
    "H": "PU BLACK",
    "I": "PU WHITE",
    "J": "PU 2색",
    "K": "PU 4색",
    "L": "PU 6색",
    "M": "PU 8색",
    "N": "PU 12색",
    "O": "PE YELLOW",
    "P": "PE BLUE",
    "Q": "PE AQUA",
    "R": "PE ORANGE",
    "S": "PE BLACK",
    "T": "PE WHITE",
    "U": "PE 2색",
    "V": "PE 4색",
    "W": "PE 6색",
    "X": "PE 8색",
    "Y": "PE 12색",
}

# 위치 8, 9: 커넥터 종류 및 연마면
CONNECTOR = {
    "0": "SC/PC",
    "1": "SC/APC",
    "2": "SC/PC CLIP",
    "3": "SC/APC CLIP",
    "4": "LC/PC",
    "5": "LC/APC",
    "6": "LC/PC CLIP",
    "7": "LC/APC CLIP",
    "8": "FC/PC",
    "9": "FC/APC",
    "A": "ST/PC",
    "B": "ST/APC",
    "C": "MU/PC",
    "D": "MU/APC",
    "E": "MTRJ/PC",
    "F": "MPO A",
    "G": "MPO B",
    "H": "MPO C",
    "I": "MTP A",
    "J": "MTP B",
    "K": "MTP C",
}
CONNECTOR_INV = {v: k for k, v in CONNECTOR.items()}  # 이름 → 코드

# 커넥터 이름 파싱용 별칭
CONNECTOR_ALIAS = {
    "SC/PC":       "0",
    "SC/APC":      "1",
    "SC/PC CLIP":  "2",
    "SC/PC,CLIP":  "2",
    "SC/APC CLIP": "3",
    "SC/APC,CLIP": "3",
    "LC/PC":       "4",
    "LC/APC":      "5",
    "LC/PC CLIP":  "6",
    "LC/PC,CLIP":  "6",
    "LC DUPLEX KIT(CLIP)": "6",
    "LC/APC CLIP": "7",
    "LC/APC,CLIP": "7",
    "FC/PC":       "8",
    "FC/APC":      "9",
    "ST/PC":       "A",
    "ST/APC":      "B",
}

# 위치 13: 마킹
MARKING = {
    "N": "NON",
    "K": "KT",
    "L": "LG",
    "O": "고려오트론",
    "P": "포앤티",
    "Z": "수출",
    "X": "케이블 사급",
}


# ─── 길이 코드 변환 ───────────────────────────────────────────────────────────

_SPECIAL_LEN = {100: "100", 150: "150", 200: "200", 250: "250", 300: "300"}

def length_to_code(m: float) -> str:
    """길이(M) → 3자리 코드. 표준 길이 외 입력 시 ValueError."""
    if m in _SPECIAL_LEN:
        return _SPECIAL_LEN[m]
    if m < 1:
        d = round(m * 10)
        return f"00{chr(ord('A') + d - 1)}"
    if m < 10:
        i, d = int(m), round((m - int(m)) * 10)
        return f"00{i}" if d == 0 else f"0{i}{chr(ord('A') + d - 1)}"
    if m < 100:
        i, d = int(m), round((m - int(m)) * 10)
        return f"{i:02d}J" if d == 0 else f"{i:02d}{chr(ord('A') + d - 1)}"
    raise ValueError(f"지원하지 않는 길이: {m}M")

def parse_length(spec: str) -> Optional[float]:
    """규격 문자열에서 길이(M) 추출. 예: '40M' → 40.0, '300MM' → 0.3"""
    spec = spec.upper().strip()
    # MM 단위 (300MM → 0.3M)
    m = re.search(r"(\d+(?:\.\d+)?)\s*MM(?!\s*M)", spec)
    if m:
        return round(float(m.group(1)) / 1000, 3)
    # M 단위
    m = re.search(r"(\d+(?:\.\d+)?)\s*M(?!M)", spec)
    if m:
        return float(m.group(1))
    return None


# ─── 자동 판별 규칙 ───────────────────────────────────────────────────────────

def detect_marking(product_name: str) -> str:
    """품목명 prefix → 마킹 코드 자동 판별"""
    n = product_name.upper().strip()
    if n.startswith("OJC-A1-") or n.startswith("OJC-C2-"):
        return "K"
    if n.startswith("SOJC-") or n.startswith("DOJC-") or n.startswith("MOJC-"):
        return "L"
    # DROP, PIGTAIL, Optical Cable Parts → 비계약 품목 → NON
    return "N"


def detect_material(core_code: str, diameter_mm: float) -> Optional[str]:
    """
    코어 종류 코드 + 직경 → 재질 코드 자동 판별

    확정 규칙:
        B3 (D) + 2.0mm  → 0 (PVC YELLOW)
        B3 (D) + 3.0mm  → 0 (PVC YELLOW)
        A1 (B) + 2.0mm  → 0 (PVC YELLOW)
        A1 (B) + 3.0mm  → B (LSZH BLACK)  ← DROP cable
        A2 (C) + 3.0mm  → B (LSZH BLACK)  ← DROP cable
        (A2 2.0mm, B3 3.0mm DP → 존재하지 않음)

    TODO: MOJC (D/E) 재질 규칙 확인 필요
    TODO: OJC-C2 재질 규칙 확인 필요
    TODO: Pigtail 재질 규칙 확인 필요
    """
    if core_code == "D":                          # SMF 657.B3
        return "0"                                # PVC YELLOW (2.0/3.0 공통)
    if core_code == "B" and diameter_mm == 2.0:  # SMF 657.A1 2.0mm
        return "0"                                # PVC YELLOW
    if core_code in ("B", "C") and diameter_mm == 3.0:  # A1/A2 3.0mm DROP
        return "B"                                # LSZH BLACK
    return None  # 확인 필요


def detect_ojc_type(product_name: str, diameter_mm: Optional[float] = None) -> Optional[str]:
    """
    품목명 → OJC 종류 코드 자동 판별

    TODO: MOJC 직경 규칙 확인 필요 (SM이 0.9mm인지 2.0mm인지)
    TODO: OJC-A1 단심/다심 구분 규칙 확인 필요
    """
    n = product_name.upper().strip()
    if n.startswith("SOJC-"):
        return "A"   # Simplex 0.9mm
    if n.startswith("DOJC-"):
        return "C"   # Duplex
    if n.startswith("MOJC-") or n.startswith("OJC-C2-"):
        # TODO: 0.9mm=D, 2.0mm=E 규칙 확인 필요
        if diameter_mm == 0.9:
            return "D"
        if diameter_mm == 2.0:
            return "E"
        return None  # 직경 정보 필요
    if n.startswith("DROP-CABLE"):
        return "K"   # Drop cable 3.0mm (기본)
    if n.startswith("PIGTAIL-"):
        return "F"   # Pigtail 0.9mm (기본, 직경에 따라 G/H)
    if n.startswith("OPTICAL CABLE PARTS"):
        return "M"   # Optical Cable 3.0mm 인입광
    return None


def detect_core_type(spec: str) -> Optional[str]:
    """규격 문자열에서 코어 종류 코드 추출"""
    s = spec.upper().strip()
    if "657B3" in s or "657.B3" in s or "B3" in s:
        return "D"
    if "657A2" in s or "657.A2" in s or "A2" in s:
        return "C"
    if "657A1" in s or "657.A1" in s or "A1" in s:
        return "B"
    if "652D" in s or "652" in s:
        return "A"
    if "OM5" in s: return "I"
    if "OM4" in s: return "H"
    if "OM3" in s: return "G"
    if "OM2" in s: return "F"
    if "OM1" in s: return "E"
    # SM만 표기된 경우 → TODO: 제품별 기본 코어 확인 필요
    return None


def detect_core_count(product_name: str) -> Optional[str]:
    """품목명에서 심선수 코드 추출. 예: MOJC-SM-6C → 4 (6심)"""
    m = re.search(r"[-_](\d+)C(?:$|[-_\s])", product_name.upper())
    if not m:
        # DOJC/SOJC 등 명시 없는 경우 기본값
        n = product_name.upper()
        if n.startswith("DOJC-"):
            return "2"   # 항상 2심
        if n.startswith("SOJC-"):
            return "1"   # 항상 1심
        return None
    count = int(m.group(1))
    return CORE_COUNT_INV.get(count)  # 없으면 None


def detect_connectors(product_name: str):
    """
    품목명에서 커넥터A, 커넥터B 코드 추출
    반환: (conn_a, conn_b) 코드 튜플

    지원 패턴:
        DOJC-SM-LC/PC-LC/PC-5M         → ('4', '4')
        DOJC-SM-SC/PC-FC/APC           → ('0', '9')
        OJC-A1-SC/LC-SM-3-APC/PC-SP   → pos3=커넥터, pos6=페롤 → 결합
        DROP-CABLE(LC/PC)              → ('4', '4')
        DROP-CABLE(LC/PC-SC/APC)       → ('4', '1')
        PIGTAIL-LC/APC-SM-12C          → ('5', None)
    """
    n = product_name.upper().strip()

    # DROP-CABLE(XX) or DROP-CABLE(XX-YY)
    m = re.search(r"DROP-CABLE\(([^)]+)\)", n)
    if m:
        inner = m.group(1)
        parts = inner.split("-")
        ca = CONNECTOR_ALIAS.get(parts[0].strip())
        cb = CONNECTOR_ALIAS.get(parts[1].strip()) if len(parts) > 1 else ca
        return ca, cb

    # OJC-A1 / OJC-C2: pos3=커넥터, pos6=페롤 → 조합
    if n.startswith("OJC-A1-") or n.startswith("OJC-C2-"):
        parts = n.split("-")
        # parts: ['OJC', 'A1', 커넥터/커넥터, SM, 길이, 페롤/페롤, ...]
        if len(parts) >= 6:
            conn_raw = parts[2]   # SC/LC
            ferrule_raw = parts[5]  # APC/PC
            connectors = conn_raw.split("/")
            ferrules = ferrule_raw.split("/")
            if len(connectors) == 2 and len(ferrules) == 2:
                ta = f"{connectors[0]}/{ferrules[0]}"
                tb = f"{connectors[1]}/{ferrules[1]}"
                return CONNECTOR_ALIAS.get(ta), CONNECTOR_ALIAS.get(tb)

    # PIGTAIL: 타입1만
    m = re.search(r"PIGTAIL-([A-Z]+/[A-Z]+)", n)
    if m:
        return CONNECTOR_ALIAS.get(m.group(1)), None

    # LG DOJC/SOJC/MOJC: XXXX-SM-커넥터A-커넥터B 또는 커넥터A-커넥터B-길이
    # 예: DOJC-SM-LC/PC-LC/PC-5M
    conn_pattern = re.findall(r"((?:SC|LC|FC|ST|MU)/(?:PC|APC)(?:\s*CLIP)?)", n)
    if conn_pattern:
        ca = CONNECTOR_ALIAS.get(conn_pattern[0])
        cb = CONNECTOR_ALIAS.get(conn_pattern[1]) if len(conn_pattern) > 1 else None
        return ca, cb

    return None, None


# ─── 메인 파싱 함수 ───────────────────────────────────────────────────────────

def parse_product(product_name: str, spec: str,
                  diameter_mm: Optional[float] = None,
                  overrides: dict = None) -> dict:
    """
    품목명 + 규격 → 품번 구성 요소 딕셔너리 반환

    반환값:
        {
          'ojc_type':    str | None,   # 위치 4
          'core_type':   str | None,   # 위치 5
          'core_count':  str | None,   # 위치 6
          'material':    str | None,   # 위치 7
          'connector_a': str | None,   # 위치 8
          'connector_b': str | None,   # 위치 9
          'length_code': str | None,   # 위치 10~12
          'marking':     str | None,   # 위치 13
          'missing':     list[str],    # 확인 필요 항목
        }
    """
    result = {}
    missing = []

    # OJC 종류
    result['ojc_type'] = detect_ojc_type(product_name, diameter_mm)
    if not result['ojc_type']:
        missing.append('ojc_type (직경 정보 필요)')

    # 코어 종류
    result['core_type'] = detect_core_type(spec)
    if not result['core_type']:
        missing.append('core_type (SMF 종류 확인 필요)')

    # 심선수
    result['core_count'] = detect_core_count(product_name)
    if not result['core_count']:
        missing.append('core_count (심선수 확인 필요)')

    # 재질 (코어 + 직경으로 자동)
    diam = diameter_mm
    if diam is None:
        # 규격에서 직경 추출 시도
        m = re.search(r"(\d+(?:\.\d+)?)\s*MM(?:$|[^M])", spec.upper())
        if m:
            diam = float(m.group(1))
    if result['core_type'] and diam:
        result['material'] = detect_material(result['core_type'], diam)
        if not result['material']:
            missing.append('material (재질 규칙 미정의 조합)')
    else:
        result['material'] = None
        missing.append('material (직경 정보 필요)')

    # 커넥터
    result['connector_a'], result['connector_b'] = detect_connectors(product_name)
    if not result['connector_a']:
        missing.append('connector_a')
    if result['connector_b'] is None and not product_name.upper().startswith("PIGTAIL"):
        missing.append('connector_b')

    # 길이
    length_m = parse_length(spec)
    if length_m:
        try:
            result['length_code'] = length_to_code(length_m)
        except ValueError as e:
            result['length_code'] = None
            missing.append(f'length_code ({e})')
    else:
        result['length_code'] = None
        missing.append('length_code (길이 정보 없음)')

    # 마킹
    result['marking'] = detect_marking(product_name)

    # overrides 적용 (수동 입력으로 덮어쓰기)
    if overrides:
        for k, v in overrides.items():
            if v:
                result[k] = v
                if k in missing:
                    missing = [x for x in missing if not x.startswith(k)]

    result['missing'] = missing
    return result


def generate_part_number(product_name: str, spec: str,
                         diameter_mm: Optional[float] = None,
                         overrides: dict = None) -> dict:
    """
    품목명 + 규격 → 품번 생성

    반환값:
        {
          'part_number': str | None,  # 생성된 품번 (모든 항목 확인 시)
          'fields':      dict,         # 각 필드 값
          'missing':     list[str],    # 확인 필요 항목
          'error':       str | None,   # 오류 메시지
        }
    """
    fields = parse_product(product_name, spec, diameter_mm, overrides)
    missing = fields.pop('missing')

    required = ['ojc_type', 'core_type', 'core_count', 'material',
                'connector_a', 'length_code', 'marking']
    incomplete = [f for f in required if not fields.get(f)]
    # PIGTAIL은 connector_b 없어도 됨
    if not product_name.upper().startswith("PIGTAIL"):
        if not fields.get('connector_b'):
            incomplete.append('connector_b')

    if incomplete or missing:
        return {
            'part_number': None,
            'fields': fields,
            'missing': missing or incomplete,
            'error': f"확인 필요: {', '.join(missing or incomplete)}"
        }

    # 조합
    t = fields
    cb = t.get('connector_b', '0')
    pn = (f"A14-{t['ojc_type']}-"
          f"{t['core_type']}{t['core_count']}{t['material']}-"
          f"{t['connector_a']}{cb}-"
          f"{t['length_code']}{t['marking']}")
    return {
        'part_number': pn,
        'fields': fields,
        'missing': [],
        'error': None
    }


# ─── 중복 체크 & 등록 ────────────────────────────────────────────────────────

def check_duplicate(part_number: str, list_path: str) -> bool:
    """품번리스트 Excel에서 중복 여부 확인"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(list_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == part_number.strip():
                return True
        return False
    except Exception as e:
        raise RuntimeError(f"품번리스트 읽기 실패: {e}")


def register_part(part_number: str, product_name: str, spec: str,
                  note: str = "", list_path: str = None) -> bool:
    """품번리스트 Excel에 신규 품번 등록"""
    import openpyxl
    from datetime import date

    if list_path is None:
        import os
        list_path = os.path.join(os.path.dirname(__file__),
                                 "..", "OJC_품번생성기.xlsx")
    try:
        wb = openpyxl.load_workbook(list_path)
        ws = wb["품번리스트"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, part_number)
        ws.cell(next_row, 2, product_name)
        ws.cell(next_row, 3, spec)
        ws.cell(next_row, 4, note)
        ws.cell(next_row, 5, date.today().isoformat())
        wb.save(list_path)
        return True
    except Exception as e:
        raise RuntimeError(f"등록 실패: {e}")


# ─── 테스트 ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":

    TEST_CASES = [
        # (품목명, 규격, 직경, 예상품번)
        ("DOJC-SM-LC/PC-LC/PC",     "17M, B3, 2.0mm",  2.0, "A14-C-D24-44-17JL"),
        ("DOJC-SM-SC/PC-FC/APC",    "60M, B3, 2.0mm",  2.0, "A14-C-D24-09-60JL"),
        ("DROP-CABLE(SC/PC)",        "75M",             3.0, None),   # core 확인 필요
        ("MOJC-SM-6C-SC/PC-SC/PC",  "40M",             None, None),  # TODO: 직경 확인
        ("OJC-C2-SC/LC-SM-10-PC/PC-12C", "10M",        None, None),  # TODO
        ("Optical Cable Parts,LC/PC-SC/APC-Single mode-5m-2Core",
         "657A2, LSZH, 5m", 3.0, None),
    ]

    print("=" * 65)
    print(f"{'품목명':<40} {'결과':<25}")
    print("=" * 65)
    for name, spec, diam, expected in TEST_CASES:
        r = generate_part_number(name, spec, diam)
        pn = r['part_number'] or f"⚠ {r['error']}"
        ok = ""
        if expected and r['part_number']:
            ok = "✅" if r['part_number'] == expected else f"❌ (예상:{expected})"
        print(f"{name[:38]:<40} {pn:<25} {ok}")
    print("=" * 65)
