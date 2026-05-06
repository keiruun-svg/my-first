"""
OJC 판매량 필터링 스크립트
-------------------------------
용도  : 전체 판매량 Excel 파일에서 OJC 완제품 항목만 추출하여 정리된 Excel 파일 생성
작성자: AJWorld 품질기술팀 학습자료 기반
버전  : 1.0 (2026-04-28)

사용법:
    python ojc_filter.py <입력파일.xlsx> [출력파일.xlsx]

    입력파일: 판매량 원본 Excel (컬럼: 년,월,일,품목코드,품목명,규격명,수량,거래처,창고)
    출력파일: 생략 시 'OJC_판매량_정리.xlsx'로 저장

출력 탭:
    - OJC 판매량    : OJC 완제품 (음수 수량 제거)
    - Distribution 케이블 : Distribution 케이블 별도 관리
"""

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 분류 기준 ──────────────────────────────────────────────
# OJC 완제품: 품목명이 아래 prefix로 시작하는 항목
OJC_PREFIXES = {
    'KT OJC'            : ('OJC-A1-', 'OJC-C2-'),
    'LG OJC'            : ('SOJC-', 'DOJC-', 'MOJC-'),
    'DROP'              : ('DROP-CABLE',),
    'PIGTAIL'           : ('PIGTAIL-',),
    'Optical Cable Parts': ('Optical Cable Parts',),
    'DX-MM'             : ('DX-MM',),
}

# Distribution 케이블: 별도 탭으로 관리
DIST_PREFIXES = ('Distribution-CABLE', 'DISTRIBUTION CABLE')

# OJC 탭 종류별 배경색 (Hex, 6자리)
COLOR_MAP = {
    'KT OJC'            : 'DEEAF1',
    'LG OJC'            : 'E2EFDA',
    'DROP'              : 'FFF2CC',
    'PIGTAIL'           : 'FCE4D6',
    'Optical Cable Parts': 'F4E6FF',
    'DX-MM'             : 'EDEDED',
}


# ── 분류 함수 ──────────────────────────────────────────────
def classify_ojc(name: str) -> str | None:
    if not isinstance(name, str):
        return None
    for label, prefixes in OJC_PREFIXES.items():
        if name.startswith(prefixes):
            return label
    return None


def classify_dist(name: str) -> bool:
    if not isinstance(name, str):
        return False
    return name.startswith(DIST_PREFIXES)


# ── 스타일 적용 ────────────────────────────────────────────
def apply_style(ws, color_map: dict = None, ojc_col_idx: int = None):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    for cell in ws[1]:
        cell.fill = PatternFill('solid', start_color='1F4E79')
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 20

    # 바디
    body_font = Font(name='Arial', size=9)
    for row in ws.iter_rows(min_row=2):
        if color_map and ojc_col_idx is not None:
            fill_color = color_map.get(row[ojc_col_idx].value, 'FFFFFF')
        else:
            fill_color = 'EAF2FF'
        fill = PatternFill('solid', start_color=fill_color)
        for cell in row:
            cell.fill = fill
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'


# ── 검증 ──────────────────────────────────────────────────
def validate(ojc_df: pd.DataFrame) -> list[str]:
    warnings = []

    # 음수 수량 잔존
    neg = ojc_df[ojc_df['수량'] <= 0]
    if len(neg):
        warnings.append(f'[경고] 음수/0 수량 {len(neg)}건 잔존')

    # 결측값
    for col in ['년', '월', '일', '품목명', '수량', '거래처']:
        if col in ojc_df.columns and ojc_df[col].isna().any():
            warnings.append(f'[경고] {col} 컬럼에 결측값 있음')

    # 품명 규칙 이상
    for name in ojc_df['품목명'].unique():
        parts = name.split('-')
        if name.startswith(('OJC-A1-', 'OJC-C2-')) and len(parts) < 7:
            warnings.append(f'[품명 이상] KT: {name}')
        elif name.startswith(('SOJC-', 'DOJC-', 'MOJC-')) and len(parts) < 4:
            warnings.append(f'[품명 이상] LG: {name}')

    return warnings


# ── 메인 ──────────────────────────────────────────────────
def run(input_path: str, output_path: str = 'OJC_판매량_정리.xlsx'):
    print(f'[1/5] 파일 읽는 중: {input_path}')
    df = pd.read_excel(input_path)
    print(f'      전체 {len(df):,}건')

    print('[2/5] OJC 항목 분류 중...')
    df['OJC종류'] = df['품목명'].apply(classify_ojc)
    ojc_df = df[df['OJC종류'].notna()].copy()
    dist_df = df[df['품목명'].apply(classify_dist)].copy()

    print('[3/5] 취소 수량(음수) 제거 중...')
    before = len(ojc_df)
    ojc_df = ojc_df[ojc_df['수량'] > 0].copy()
    print(f'      제거: {before - len(ojc_df)}건 → 잔여 {len(ojc_df):,}건')

    print('[4/5] 검증 중...')
    warnings = validate(ojc_df)
    if warnings:
        for w in warnings:
            print(f'      {w}')
    else:
        print('      이상 없음')

    print('[5/5] 파일 저장 중...')
    ojc_cols = ['년', '월', '일', 'OJC종류', '품목코드', '품목명', '규격명', '수량', '거래처', '창고']
    dist_cols = ['년', '월', '일', '품목코드', '품목명', '규격명', '수량', '거래처', '창고']
    ojc_out = ojc_df[[c for c in ojc_cols if c in ojc_df.columns]]
    dist_out = dist_df[[c for c in dist_cols if c in dist_df.columns]]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        ojc_out.to_excel(writer, index=False, sheet_name='OJC 판매량')
        dist_out.to_excel(writer, index=False, sheet_name='Distribution 케이블')

    wb = load_workbook(output_path)

    ws1 = wb['OJC 판매량']
    apply_style(ws1, color_map=COLOR_MAP, ojc_col_idx=3)
    for i, w in enumerate([6,5,5,18,14,55,20,8,20,18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb['Distribution 케이블']
    apply_style(ws2)
    for i, w in enumerate([6,5,5,14,40,20,8,20,18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)

    # 요약 출력
    print(f'\n✅ 완료: {output_path}')
    print('\n[OJC 판매량]')
    print(ojc_out.groupby('OJC종류')['수량'].agg(건수='count', 수량합계='sum').to_string())
    print(f'\n[Distribution 케이블] {len(dist_out)}건, 수량합계: {dist_out["수량"].sum():,}')
    if warnings:
        print(f'\n⚠️  검증 경고 {len(warnings)}건 — 상사 확인 필요')

    return ojc_out, dist_out


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('사용법: python ojc_filter.py <입력파일.xlsx> [출력파일.xlsx]')
        sys.exit(1)
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'OJC_판매량_정리.xlsx'
    run(input_file, output_file)
