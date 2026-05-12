#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AJW 생산자재 발주계획 웹앱
실행: streamlit run web_app.py --server.address=0.0.0.0 --server.port=8501
"""
import streamlit as st
import json, tempfile, os, pandas as pd
from pathlib import Path
from datetime import datetime

SETTINGS_FILE   = Path(__file__).parent / "settings.json"
METADATA_FILE   = Path(__file__).parent / "metadata.json"
INVENTORY_FILE  = Path(__file__).parent / "inventory.json"
SALES_FILE      = Path(__file__).parent / "sales_analysis.json"

DEFAULT_SETTINGS = {
    "lead_time_default": 60,
    "colors": {
        "main_header": "1F3864",
        "year_23": "2F5597",
        "year_24": "2E75B6",
        "year_25": "155480",
    }
}

def load_settings() -> dict:
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return DEFAULT_SETTINGS.copy()

def save_settings(s: dict):
    SETTINGS_FILE.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding='utf-8')

def load_metadata() -> dict:
    if METADATA_FILE.exists():
        try:
            return json.loads(METADATA_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {"cable": {}, "housing": {}}

def save_metadata(m: dict):
    METADATA_FILE.write_text(json.dumps(m, ensure_ascii=False, indent=2), encoding='utf-8')

def load_inventory() -> dict:
    if INVENTORY_FILE.exists():
        try:
            return json.loads(INVENTORY_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {"cable": {}, "housing": {}}

def save_inventory(inv: dict):
    INVENTORY_FILE.write_text(json.dumps(inv, ensure_ascii=False, indent=2), encoding='utf-8')

def load_sales_analysis() -> dict:
    if SALES_FILE.exists():
        try:
            return json.loads(SALES_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {}

def save_sales_analysis(d: dict):
    SALES_FILE.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding='utf-8')

FERRULE_TYPES = ['LC/PC', 'LC/APC', 'SC/PC', 'SC/APC', 'FC/PC', 'FC/APC']

def meta_to_step1(metadata: dict):
    """metadata.json → step1_core용 cable_meta / housing_meta_in 딕셔너리"""
    cable_meta = {}
    for k, v in metadata.get('cable', {}).items():
        parts = k.split('|', 1)
        if len(parts) == 2:
            cable_meta[(parts[0], parts[1])] = v
    housing_meta_in = {}
    for k, v in metadata.get('housing', {}).items():
        parts = k.split('|', 1)
        if len(parts) == 2:
            housing_meta_in[(parts[0], parts[1])] = v if isinstance(v, list) else [v]
    return cable_meta, housing_meta_in

def get_ferrule_meta(metadata: dict) -> dict:
    """metadata.json ferrule 섹션 → {커넥터타입: {품번, 품명, ...}} (6종 기본값 보장)"""
    base = {t: {"품번": "", "품명": f"FERRULE (W/ FLANGE,{t} TYPE)", "구매처": "", "리드타임": None}
            for t in FERRULE_TYPES}
    base.update(metadata.get('ferrule', {}))
    return base

def save_upload(uploaded_file, tmp_dir: str, filename: str) -> str:
    path = os.path.join(tmp_dir, filename)
    with open(path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return path

def ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def parse_production_data(file_bytes: bytes) -> dict:
    """구매관리(맥산) Excel → {품목코드: {'23': qty, '24': qty, '25': qty}}
    컬럼: 구매처명, 입고일자, 발주서일자, 품목코드, 품목명, 창고명, 규격명, 수량, ...
    입고일자 형식: '25/01/09-2' 또는 '2025/01/08 -2'
    """
    import io as _io, re as _re
    df = pd.read_excel(_io.BytesIO(file_bytes), dtype=str)

    col_map = {}
    for col in df.columns:
        s = str(col).strip()
        if s == '품목코드':                  col_map[col] = '품목코드'
        elif s == '수량':                    col_map[col] = '수량'
        elif '입고' in s and '일자' in s:    col_map[col] = '입고일자'
    df.rename(columns=col_map, inplace=True)

    def _yr(d):
        if not d or str(d).strip() in ('', 'nan', 'None'): return None
        s = _re.sub(r'\s*-\d+\s*$', '', str(d).strip())
        if '/' in s:
            p = s.split('/')[0].strip()
            if len(p) == 4: return p[2:]
            if len(p) == 2: return p
        return None

    result = {}
    for _, row in df.iterrows():
        code = str(row.get('품목코드', '') or '').strip()
        if not code or code in ('nan', 'None'): continue
        try:
            qty_raw = row.get('수량', 0)
            qty = int(float(str(qty_raw))) if pd.notna(qty_raw) and str(qty_raw).strip() not in ('', 'nan') else 0
        except Exception:
            qty = 0
        if qty <= 0: continue
        yr = _yr(row.get('입고일자', ''))
        if yr not in ('23', '24', '25'): continue
        result.setdefault(code, {'23': 0, '24': 0, '25': 0})[yr] += qty
    return result

def _build_ojc_excel(ojc_df, dist_df) -> bytes:
    """OJC 필터링 결과 DataFrame → 스타일 적용 Excel bytes"""
    import io as _io, openpyxl as _xl
    COLOR_MAP = {
        'KT OJC': 'DEEAF1', 'LG OJC': 'E2EFDA', 'DROP': 'FFF2CC',
        'PIGTAIL': 'FCE4D6', 'Optical Cable Parts': 'F4E6FF', 'DX-MM': 'EDEDED',
    }
    ojc_cols  = ['년','월','일','OJC종류','품목코드','품목명','규격명','수량','거래처','창고']
    dist_cols = ['년','월','일','품목코드','품목명','규격명','수량','거래처','창고']
    ojc_out  = ojc_df[[c for c in ojc_cols  if c in ojc_df.columns]]
    dist_out = dist_df[[c for c in dist_cols if c in dist_df.columns]]

    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        ojc_out.to_excel(writer,  index=False, sheet_name='OJC 판매량')
        dist_out.to_excel(writer, index=False, sheet_name='Distribution 케이블')

    buf.seek(0)
    wb = _xl.load_workbook(buf)
    for ws, ojc_col_idx, col_widths in [
        (wb['OJC 판매량'],          3, [6,5,5,18,14,55,20,8,20,18]),
        (wb['Distribution 케이블'], None, [6,5,5,14,40,20,8,20,18]),
    ]:
        from openpyxl.styles import Font as F, PatternFill as PF, Alignment as AL, Border as BD, Side as SD
        thin = SD(style='thin', color='CCCCCC')
        bdr  = BD(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill      = PF('solid', start_color='1F4E79')
            cell.font      = F(name='Arial', bold=True, color='FFFFFF', size=10)
            cell.alignment = AL(horizontal='center', vertical='center')
            cell.border    = bdr
        ws.row_dimensions[1].height = 20
        for row in ws.iter_rows(min_row=2):
            if ojc_col_idx is not None:
                fc = COLOR_MAP.get(row[ojc_col_idx].value, 'FFFFFF')
            else:
                fc = 'EAF2FF'
            fill = PF('solid', start_color=fc)
            for cell in row:
                cell.fill = fill; cell.font = F(name='Arial', size=9)
                cell.border = bdr; cell.alignment = AL(vertical='center')
        ws.freeze_panes = 'A2'
        from openpyxl.utils import get_column_letter as gcl
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[gcl(i)].width = w

    out = _io.BytesIO(); wb.save(out); return out.getvalue()


def build_sales_analysis(sales_bytes: bytes, purchase_bytes: bytes,
                          progress_cb=None) -> tuple:
    """전체 판매량.xlsx + 구매관리(맥산).xlsx
    → (analysis_dict, summary_dict, errors, ojc_xl_bytes)
    analysis_dict = {품목코드: {'품목명':..., '23':{'sales':N,'production':N,'ratio':R}, ...}}
    """
    import sys as _sys, importlib as _il, io as _io
    _sys.path.insert(0, str(Path(__file__).parent))
    import ojc_filter as _ojf
    _il.reload(_ojf)

    def _prog(pct, msg=''):
        if progress_cb: progress_cb(pct, msg)

    errors = []

    _prog(5, "판매량 파일 읽는 중...")
    try:
        df = pd.read_excel(_io.BytesIO(sales_bytes), dtype=str)
        df['OJC종류'] = df['품목명'].apply(
            lambda x: _ojf.classify_ojc(str(x)) if pd.notna(x) else None)
        ojc_df  = df[df['OJC종류'].notna()].copy()
        dist_df = df[df['품목명'].apply(
            lambda x: _ojf.classify_dist(str(x)) if pd.notna(x) else False)].copy()

        def _qty(v):
            try: return int(float(str(v))) if pd.notna(v) and str(v).strip() not in ('','nan') else 0
            except: return 0

        def _yr_from_row(row):
            yr_raw = row.get('년', '')
            if pd.notna(yr_raw) and str(yr_raw).strip() not in ('', 'nan'):
                yr_s = str(int(float(str(yr_raw))))
                return yr_s[2:] if len(yr_s) == 4 else yr_s[-2:]
            return None

        ojc_df['_qty'] = ojc_df['수량'].apply(_qty)
        ojc_df = ojc_df[ojc_df['_qty'] > 0].copy()
        ojc_df['_yr'] = ojc_df.apply(_yr_from_row, axis=1)
    except Exception as e:
        errors.append(f"판매량 파일 오류: {e}")
        return {}, {}, errors, b''

    _prog(30, f"OJC 분류 완료 ({len(ojc_df):,}건) — 집계 중...")
    sales_by = {}
    for _, row in ojc_df.iterrows():
        code = str(row.get('품목코드', '') or '').strip()
        yr = row.get('_yr')
        qty = int(row.get('_qty', 0))
        if not code or code in ('nan','None') or yr not in ('23','24','25') or qty <= 0: continue
        if code not in sales_by:
            sales_by[code] = {'품목명': str(row.get('품목명','') or ''), '23':0,'24':0,'25':0}
        sales_by[code][yr] += qty

    _prog(55, "구매관리(맥산) 파일 읽는 중...")
    if not purchase_bytes:
        prod_by = {}
    else:
        try:
            prod_by = parse_production_data(purchase_bytes)
        except Exception as e:
            errors.append(f"구매관리 파일 오류: {e}")
            prod_by = {}

    _prog(75, "생산비중 계산 중...")
    all_codes = set(sales_by) | set(prod_by)
    analysis = {}
    for code in all_codes:
        s = sales_by.get(code, {'품목명':'','23':0,'24':0,'25':0})
        p = prod_by.get(code, {'23':0,'24':0,'25':0})
        yr_data = {}
        for yr in ('23','24','25'):
            sq = s.get(yr, 0); pq = p.get(yr, 0)
            ratio = round(pq / sq, 4) if sq > 0 else (1.0 if pq > 0 else 0.0)
            yr_data[yr] = {'sales': sq, 'production': pq, 'ratio': ratio}
        analysis[code] = {'품목명': s.get('품목명',''), **yr_data}

    _prog(90, "Excel 파일 생성 중...")
    try:
        ojc_xl = _build_ojc_excel(ojc_df, dist_df)
    except Exception as e:
        errors.append(f"Excel 생성 오류: {e}")
        ojc_xl = b''

    summary = {
        'total': len(all_codes),
        'both':  sum(1 for c in all_codes if c in sales_by and c in prod_by),
        'sales_only': sum(1 for c in all_codes if c in sales_by and c not in prod_by),
        'prod_only':  sum(1 for c in all_codes if c not in sales_by and c in prod_by),
        'ojc_rows': len(ojc_df),
    }
    _prog(100, "완료")
    return analysis, summary, errors, ojc_xl

# ── Excel 양식 빌더 ──────────────────────────────────────────
def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color='BBBBBB')
    thick = Side(style='medium', color='888888')
    return (
        Font, PatternFill, Alignment,
        Border(left=thin, right=thin, top=thin,    bottom=thin),
        Border(left=thin, right=thin, top=thick,   bottom=thin),   # header bottom thick
    )

def _xl_title(ws, ncols, text):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    c = ws['A1']
    c.value = text
    c.font = Font(italic=True, name='Arial', size=9, color='555555')
    c.fill = PatternFill('solid', start_color='EFEFEF')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 16

def _xl_header(ws, row, headers, bg):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color='FFFFFF')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = bdr
    ws.row_dimensions[row].height = 24

def _xl_example(ws, row, values):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill('solid', start_color='FFF2CC')
    for ci, v in enumerate(values, 1):
        c = ws.cell(row, ci, v)
        c.font      = Font(italic=True, color='7F6000', name='Arial', size=9)
        c.fill      = fill
        c.border    = bdr
        c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 15

def _xl_data(ws, row, values, alt=False):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill('solid', start_color='F8F8F8') if alt else None
    for ci, v in enumerate(values, 1):
        c = ws.cell(row, ci, v)
        c.font      = Font(name='Arial', size=9)
        c.border    = bdr
        c.alignment = Alignment(vertical='center')
        if fill: c.fill = fill
    ws.row_dimensions[row].height = 15

def _xl_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────
def build_meta_excel(metadata: dict) -> bytes:
    """품번 관리 Excel 양식 생성 (예시 행 포함)"""
    import io, openpyxl
    wb = openpyxl.Workbook()

    # ── 케이블 시트 ──────────────────────────────────────────
    ws = wb.active; ws.title = '케이블'
    hdrs = ['파이', '케이블종류', '품번', '품명', '구매처', '리드타임(일)']
    _xl_title(ws, len(hdrs),
              '※ 파이 값: 2.0mm / 3.0mm / 0.9mm  |  노란 예시 행은 삭제 후 실제 데이터를 입력하세요.')
    _xl_header(ws, 2, hdrs, '2E75B6')
    _xl_example(ws, 3, ['2.0mm', '※예시 A1-SP',   'P14-RM-4188', 'OPTICAL CABLE 2.0mm G657A1 SP-SM (YELLOW,황)',          'HUAMAI',   60])
    _xl_example(ws, 4, ['3.0mm', '※예시 DROP-SP',  'P14-RM-4191', 'DROP OPTICAL CABLE 3.0mm SP-SM (BLACK)',                'HUAMAI',   60])
    ri = 5
    for idx, (k, v) in enumerate(sorted(metadata.get('cable', {}).items())):
        pai, ct = k.split('|',1) if '|' in k else (k,'')
        _xl_data(ws, ri, [pai, ct, v.get('품번',''), v.get('품명',''), v.get('구매처',''), v.get('리드타임') or ''], idx%2==1)
        ri += 1
    _xl_col_widths(ws, [9, 14, 18, 48, 14, 12])
    ws.freeze_panes = 'A3'

    # ── 하우징 시트 ─────────────────────────────────────────
    ws = wb.create_sheet('하우징')
    hdrs = ['파이', '하우징타입', '품번', '품명', '구매처', '리드타임(일)']
    _xl_title(ws, len(hdrs),
              '※ 같은 하우징 타입에 부품이 여러 개(하우징·페롤·스프링·더스트캡)면 파이·하우징타입을 동일하게 쓰고 행을 추가하세요.  |  노란 예시 행 삭제 후 입력')
    _xl_header(ws, 2, hdrs, '375623')
    _xl_example(ws, 3, ['2.0mm', '※예시 LC/PC 청색', 'P14-LS-4228',  'OJC HOUSING KIT (OJC LC/PC 2.0 BLUE)',   'FIBERCAN',   60])
    _xl_example(ws, 4, ['2.0mm', '※예시 LC/PC 청색', 'N94-1-04075',  'SPRING',                                  '상영스프링',  30])
    ri = 5
    for idx, (k, v) in enumerate(sorted(metadata.get('housing', {}).items())):
        pai, htype = k.split('|',1) if '|' in k else (k,'')
        for item in (v if isinstance(v, list) else [v]):
            _xl_data(ws, ri, [pai, htype, item.get('품번',''), item.get('품명',''), item.get('구매처',''), item.get('리드타임') or ''], idx%2==1)
            ri += 1
    _xl_col_widths(ws, [9, 16, 18, 48, 14, 12])
    ws.freeze_panes = 'A3'

    # ── 페롤 시트 ────────────────────────────────────────────
    ws = wb.create_sheet('페롤')
    hdrs = ['커넥터타입', '품번', '품명', '구매처', '리드타임(일)']
    _xl_title(ws, len(hdrs),
              '※ 커넥터타입(A열)은 수정하지 마세요. 품번·품명·구매처·리드타임만 입력합니다.')
    _xl_header(ws, 2, hdrs, '7030A0')
    fm = get_ferrule_meta(metadata)
    for idx, t in enumerate(FERRULE_TYPES):
        v = fm[t]
        _xl_data(ws, 3+idx, [t, v.get('품번',''), v.get('품명',''), v.get('구매처',''), v.get('리드타임') or ''], idx%2==1)
    _xl_col_widths(ws, [14, 18, 48, 14, 12])
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ─────────────────────────────────────────────────────────────
def parse_meta_excel(file_bytes: bytes) -> tuple:
    """업로드된 품번 관리 Excel 파싱 → (new_meta, errors)
    • 1행(안내) + 2행(헤더) 구조 → header=1 으로 읽음
    • ※ 로 시작하는 예시 행 자동 제외
    """
    import io as _io
    xf = _io.BytesIO(file_bytes)
    new_meta = {}; errors = []

    def _s(v):
        s = str(v).strip() if pd.notna(v) else ''
        return '' if s in ('nan','None') else s

    def _i(v):
        try: return int(float(v)) if pd.notna(v) and str(v).strip() not in ('','nan') else None
        except: return None

    def _skip(v): return (not v) or v.startswith('※')

    try:
        df = pd.read_excel(xf, sheet_name='케이블', header=1, dtype=str)
        cable_dict = {}
        for _, row in df.iterrows():
            pai, ct = _s(row.get('파이')), _s(row.get('케이블종류'))
            if _skip(pai) or _skip(ct): continue
            cable_dict[f"{pai}|{ct}"] = {
                '품번': _s(row.get('품번')), '품명': _s(row.get('품명')),
                '구매처': _s(row.get('구매처')), '리드타임': _i(row.get('리드타임(일)')),
            }
        new_meta['cable'] = cable_dict
    except Exception as e: errors.append(f"케이블 시트 오류: {e}")

    xf.seek(0)
    try:
        df = pd.read_excel(xf, sheet_name='하우징', header=1, dtype=str)
        housing_dict = {}
        for _, row in df.iterrows():
            pai, htype = _s(row.get('파이')), _s(row.get('하우징타입'))
            if _skip(pai) or _skip(htype): continue
            housing_dict.setdefault(f"{pai}|{htype}", []).append({
                '품번': _s(row.get('품번')), '품명': _s(row.get('품명')),
                '구매처': _s(row.get('구매처')), '리드타임': _i(row.get('리드타임(일)')),
            })
        new_meta['housing'] = housing_dict
    except Exception as e: errors.append(f"하우징 시트 오류: {e}")

    xf.seek(0)
    try:
        df = pd.read_excel(xf, sheet_name='페롤', header=1, dtype=str)
        ferrule_dict = {}
        for _, row in df.iterrows():
            t = _s(row.get('커넥터타입'))
            if _skip(t): continue
            ferrule_dict[t] = {
                '품번': _s(row.get('품번')), '품명': _s(row.get('품명')),
                '구매처': _s(row.get('구매처')), '리드타임': _i(row.get('리드타임(일)')),
            }
        new_meta['ferrule'] = ferrule_dict
    except Exception as e: errors.append(f"페롤 시트 오류: {e}")

    return new_meta, errors

# ─────────────────────────────────────────────────────────────
def build_inventory_excel(metadata: dict, inventory: dict) -> bytes:
    """재고 현황 Excel 양식 생성 (예시 행 포함)"""
    import io, openpyxl
    wb = openpyxl.Workbook()
    cable_inv   = inventory.get('cable', {})
    housing_inv = inventory.get('housing', {})

    # ── 케이블 시트 ──────────────────────────────────────────
    ws = wb.active; ws.title = '케이블'
    hdrs = ['파이', '케이블종류', '품번(참고)', '현재고(m)', '기발주(m)']
    _xl_title(ws, len(hdrs),
              '※ 품번(참고) 열은 수정하지 않아도 됩니다. 현재고·기발주만 숫자로 입력하세요.  |  노란 예시 행 삭제 후 입력')
    _xl_header(ws, 2, hdrs, '2E75B6')
    _xl_example(ws, 3, ['2.0mm', '※예시 A1-SP',   'P14-RM-4188', 500, 200])
    _xl_example(ws, 4, ['3.0mm', '※예시 DROP-SP',  'P14-RM-4191', 0,   0  ])
    ri = 5
    for idx, (k, v) in enumerate(sorted(metadata.get('cable', {}).items())):
        pai, ct = k.split('|',1) if '|' in k else (k,'')
        iv = cable_inv.get(k, {})
        _xl_data(ws, ri, [pai, ct, v.get('품번',''), int(iv.get('현재고') or 0), int(iv.get('기발주') or 0)], idx%2==1)
        ri += 1
    _xl_col_widths(ws, [9, 14, 18, 13, 13])
    ws.freeze_panes = 'A3'

    # ── 하우징 시트 ─────────────────────────────────────────
    ws = wb.create_sheet('하우징')
    hdrs = ['파이', '하우징타입', '품번', '현재고(EA)', '기발주(EA)']
    _xl_title(ws, len(hdrs),
              '※ 같은 하우징 타입 내 부품별로 재고를 각각 입력하세요.  |  노란 예시 행 삭제 후 입력')
    _xl_header(ws, 2, hdrs, '375623')
    _xl_example(ws, 3, ['2.0mm', '※예시 LC/PC 청색', 'P14-LS-4228', 1000, 500])
    _xl_example(ws, 4, ['2.0mm', '※예시 LC/PC 청색', 'N94-1-04075',  300,   0])
    ri = 5
    for idx, (k, v) in enumerate(sorted(metadata.get('housing', {}).items())):
        pai, htype = k.split('|',1) if '|' in k else (k,'')
        items    = v if isinstance(v, list) else [v]
        inv_list = housing_inv.get(k, [])
        for i, item in enumerate(items):
            iv = inv_list[i] if i < len(inv_list) else {}
            _xl_data(ws, ri, [pai, htype, item.get('품번',''), int(iv.get('현재고') or 0), int(iv.get('기발주') or 0)], idx%2==1)
            ri += 1
    _xl_col_widths(ws, [9, 16, 18, 13, 13])
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ─────────────────────────────────────────────────────────────
def parse_inventory_excel(file_bytes: bytes) -> tuple:
    """업로드된 재고 현황 Excel 파싱 → (new_inv, errors)
    • 1행(안내) + 2행(헤더) 구조 → header=1 으로 읽음
    • ※ 로 시작하는 예시 행 자동 제외
    """
    import io as _io
    from collections import defaultdict as _dd
    xf = _io.BytesIO(file_bytes)
    new_inv = {'cable': {}, 'housing': {}}; errors = []

    def _s(v):
        s = str(v).strip() if pd.notna(v) else ''
        return '' if s in ('nan','None') else s

    def _i(v):
        try: return int(float(v)) if pd.notna(v) and str(v).strip() not in ('','nan') else 0
        except: return 0

    def _skip(v): return (not v) or v.startswith('※')

    try:
        df = pd.read_excel(xf, sheet_name='케이블', header=1, dtype=str)
        for _, row in df.iterrows():
            pai, ct = _s(row.get('파이')), _s(row.get('케이블종류'))
            if _skip(pai) or _skip(ct): continue
            new_inv['cable'][f"{pai}|{ct}"] = {
                '현재고': _i(row.get('현재고(m)')), '기발주': _i(row.get('기발주(m)')),
            }
    except Exception as e: errors.append(f"케이블 시트 오류: {e}")

    xf.seek(0)
    try:
        df = pd.read_excel(xf, sheet_name='하우징', header=1, dtype=str)
        grouped = _dd(list)
        for _, row in df.iterrows():
            pai, htype = _s(row.get('파이')), _s(row.get('하우징타입'))
            if _skip(pai) or _skip(htype): continue
            grouped[f"{pai}|{htype}"].append({
                '품번': _s(row.get('품번')),
                '현재고': _i(row.get('현재고(EA)')), '기발주': _i(row.get('기발주(EA)')),
            })
        new_inv['housing'] = dict(grouped)
    except Exception as e: errors.append(f"하우징 시트 오류: {e}")

    return new_inv, errors

# ── 페이지 설정 ──────────────────────────────────────────────
st.set_page_config(
    page_title="AJW 발주계획 시스템",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .main-title { font-size: 1.6rem; font-weight: 700; color: #1F3864; margin-bottom: 0.2rem; }
    .sub-title  { font-size: 0.95rem; color: #666; margin-bottom: 1.5rem; }
    .step-box   { background: #f0f4fa; border-left: 4px solid #2E75B6;
                  padding: 1rem 1.2rem; border-radius: 6px; margin-bottom: 1rem; }
    .log-box    { background: #1e1e1e; color: #d4d4d4; font-family: monospace;
                  font-size: 0.82rem; padding: 0.8rem; border-radius: 6px;
                  white-space: pre-wrap; max-height: 200px; overflow-y: auto; }
    .warn       { color: #e67e22; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">📦 AJW 생산자재 발주계획 시스템</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">(주)에이제이월드 SCM팀 — 로우데이터 업로드 후 버튼 클릭으로 Excel 자동 생성</div>', unsafe_allow_html=True)

settings = load_settings()

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📤 STEP 1 — ERP 파일 가공",
    "📈 STEP 2 — 판매 분석",
    "📊 STEP 3 — 발주계획 생성",
    "📋 품번 관리",
    "📦 재고 현황",
    "⚙️ 파라미터 & 양식 설정",
])

# ═══════════════════════════════════════════════════════════
# STEP 1
# ═══════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="step-box"><b>STEP 1 — ERP 파일 가공</b>: 맥산 ERP에서 추출한 <b>구매조회</b> 또는 <b>구매현황</b> 파일을 업로드하면 생산자재_사용내역.xlsx를 자동 생성합니다.<br>두 형식 모두 지원하며 자동으로 감지합니다. 품번·품명·구매처·리드타임은 <b>📋 품번 관리</b> 탭 정보로 자동 채워집니다.</div>', unsafe_allow_html=True)

    meta_info = load_metadata()
    n_cable_meta = len(meta_info.get('cable', {}))
    n_housing_meta = len(meta_info.get('housing', {}))
    col1, col2 = st.columns([2, 1])
    with col1:
        row_file = st.file_uploader(
            "구매조회 / 구매현황 파일 (ERP 원본)",
            type=['xlsx'], key='s1_row',
            help="맥산 ERP → 구매조회 또는 구매현황 시트가 포함된 파일 업로드 (자동 감지)"
        )
    with col2:
        st.info(
            f"저장된 품번\n\n"
            f"케이블 **{n_cable_meta}** 타입 / 하우징 **{n_housing_meta}** 타입\n\n"
            f"리드타임 기본값: **{settings.get('lead_time_default', 60)}일**"
        )
    ojc_file = None

    st.divider()

    btn_col, clr_col = st.columns([5, 1])
    with btn_col:
        run_s1 = st.button("▶ STEP 1 실행 — ERP 파일 가공 & 사용내역 생성", type="primary", disabled=(row_file is None))
    with clr_col:
        if st.button("🗑 초기화", key="s1_clear", disabled=('s1_result' not in st.session_state)):
            del st.session_state['s1_result']
            st.rerun()

    if run_s1:
        if row_file is None:
            st.error("구매조회 / 구매현황 파일을 업로드해주세요.")
        else:
            prog_bar  = st.progress(0)
            prog_text = st.empty()
            import time as _time, threading as _threading, queue as _queue
            _start = _time.time()
            _pq = _queue.Queue()

            def _on_progress(pct: int, msg: str):
                _pq.put((pct, msg))

            try:
                import importlib, step1_core
                importlib.reload(step1_core)
                metadata = load_metadata()
                cable_meta, housing_meta_in = meta_to_step1(metadata)
                _result, _error = [], []

                with tempfile.TemporaryDirectory() as tmp:
                    row_path = save_upload(row_file, tmp, "가공파일_통합_v양식.xlsx")
                    ojc_path = save_upload(ojc_file, tmp, "ojc_ref.xlsx") if ojc_file else None

                    def _worker():
                        try:
                            _result.append(step1_core.run(
                                row_path, cable_meta, housing_meta_in, ojc_path, settings,
                                progress_cb=_on_progress))
                        except Exception as _e:
                            _error.append(_e)

                    _t = _threading.Thread(target=_worker, daemon=True)
                    _t.start()

                    _last_pct, _last_msg = 0, "시작 중..."
                    while _t.is_alive():
                        try:
                            while True:
                                _last_pct, _last_msg = _pq.get_nowait()
                        except _queue.Empty:
                            pass
                        elapsed = _time.time() - _start
                        prog_bar.progress(min(_last_pct, 100))
                        prog_text.markdown(
                            f'<div style="font-size:0.85rem;color:#555;margin-top:4px">'
                            f'⏳ {_last_msg}'
                            f'<span style="float:right;color:#aaa;font-size:0.8rem">'
                            f'경과 {elapsed:.1f}초</span></div>',
                            unsafe_allow_html=True,
                        )
                        _time.sleep(0.1)

                if _error:
                    raise _error[0]

                result_bytes, logs, cable_keys, housing_keys, years, converted_bytes = _result[0]

                # 신규 타입 metadata 자동 추가
                new_types = 0
                for (pai, ct) in cable_keys:
                    k = f"{pai}|{ct}"
                    if k not in metadata.get('cable', {}):
                        metadata.setdefault('cable', {})[k] = {
                            '품번': '', '품명': '', '구매처': '', '리드타임': None}
                        new_types += 1
                for (pai, htype) in housing_keys:
                    k = f"{pai}|{htype}"
                    if k not in metadata.get('housing', {}):
                        metadata.setdefault('housing', {})[k] = [
                            {'품번': '', '품명': '', '구매처': '', '리드타임': None}]
                        new_types += 1
                if new_types:
                    save_metadata(metadata)
                    logs.append(f"ℹ️ 📋 품번 관리 탭에 신규 타입 {new_types}개 추가 — 품번을 입력해주세요.")

                total = _time.time() - _start
                n_yr = len(years)
                st.session_state['s1_result'] = {
                    'result_bytes': result_bytes,
                    'converted_bytes': converted_bytes,
                    'years': years,
                    'logs': logs,
                    'total': total,
                    'fname': f"{n_yr}개년_생산자재_사용내역_{ts()}.xlsx",
                    'fname_conv': f"가공파일_{ts()}.xlsx",
                }
                st.rerun()

            except Exception as e:
                prog_bar.progress(0)
                prog_text.empty()
                st.error(f"오류 발생: {e}")

    # ── 결과 표시 (session_state 유지 — 다운로드 후에도 사라지지 않음) ──
    if 's1_result' in st.session_state:
        r = st.session_state['s1_result']
        st.markdown(
            f'<div style="font-size:0.9rem;font-weight:600;color:#1a7a3c;'
            f'background:#e8f5e9;padding:6px 12px;border-radius:6px;'
            f'border-left:4px solid #1a7a3c;margin-top:6px">'
            f'✅ 처리 완료! &nbsp;&nbsp;'
            f'<span style="font-weight:400;font-size:0.85rem;color:#2e7d32">'
            f'총 소요시간 {r["total"]:.1f}초</span></div>',
            unsafe_allow_html=True,
        )
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                label="⬇ 생산자재_사용내역.xlsx 다운로드",
                data=r['result_bytes'],
                file_name=r['fname'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="s1_dl_result",
            )
        with dl_col2:
            if r['converted_bytes']:
                st.download_button(
                    label="⬇ 가공파일(연도별 시트).xlsx 다운로드",
                    data=r['converted_bytes'],
                    file_name=r['fname_conv'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="s1_dl_conv",
                )
            else:
                st.info("가공파일: 이미 변환된 파일을 입력하여 별도 가공파일 없음")
        log_text = "\n".join(r['logs'])
        st.markdown(f'<div class="log-box">{log_text}</div>', unsafe_allow_html=True)
        warns = [l for l in r['logs'] if '⚠' in l]
        if warns:
            st.markdown(f'<span class="warn">⚠ 주의: {warns[-1]}</span>', unsafe_allow_html=True)
            st.markdown("품번이 없는 항목은 📋 **품번 관리** 탭에서 입력 후 다시 실행하세요.")

# ═══════════════════════════════════════════════════════════
# STEP 3
# ═══════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="step-box"><b>STEP 3 — 발주계획 생성</b>: STEP 1에서 생성한 <b>가공파일</b>을 업로드하면 연간발주계획.xlsx를 생성합니다.<br>현재고·기발주는 <b>📦 재고 현황</b> 탭, 수요 기반 분석은 <b>📈 STEP 2 — 판매 분석</b> 탭에서 사전 실행하세요.</div>', unsafe_allow_html=True)

    _inv_s2   = load_inventory()
    _n_cable_inv  = sum(1 for v in _inv_s2.get('cable',{}).values() if (v.get('현재고') or 0) > 0 or (v.get('기발주') or 0) > 0)
    _n_housing_inv = sum(1 for items in _inv_s2.get('housing',{}).values() for v in (items if isinstance(items,list) else [items]) if (v.get('현재고') or 0) > 0 or (v.get('기발주') or 0) > 0)

    _sales_s2 = load_sales_analysis()
    _n_sales = sum(
        1 for v in _sales_s2.values()
        if any(v.get(yr, {}).get('sales', 0) > 0 for yr in ('23','24','25'))
    )

    col3, col4 = st.columns([2, 1])
    with col3:
        st.markdown("**필수 파일**")
        row_file2 = st.file_uploader(
            "가공파일.xlsx (STEP 1 결과)",
            type=['xlsx'], key='s2_row',
            help="STEP 1에서 다운로드한 가공파일(연도별 시트 포함). ERP 원본 파일도 업로드 가능 (자동 변환)"
        )
    with col4:
        st.info(
            f"재고 현황 탭 입력값\n\n"
            f"케이블 **{_n_cable_inv}** 항목 입력됨 / 하우징 **{_n_housing_inv}** 항목 입력됨\n\n"
            f"현재고·기발주 수정: **📦 재고 현황** 탭"
        )
        if _n_sales > 0:
            st.success(f"📈 STEP 2 판매 분석 데이터 **{_n_sales}** 품목 — 수요 기반 분석 자동 포함")
        else:
            st.warning("📈 판매 분석 데이터 없음\n\n**📈 STEP 2 — 판매 분석** 탭에서 먼저 실행하세요.")
    ojc_file2 = None
    usage_file = None

    st.divider()

    ready2 = (row_file2 is not None)
    btn2_col, clr2_col = st.columns([5, 1])
    with btn2_col:
        run_s2 = st.button("▶ STEP 3 실행 — 발주계획 생성", type="primary", disabled=(not ready2))
    with clr2_col:
        if st.button("🗑 초기화", key="s3_clear", disabled=('s3_result' not in st.session_state)):
            del st.session_state['s3_result']
            st.rerun()

    if run_s2:
        if not ready2:
            st.error("가공파일을 업로드해주세요.")
        else:
            prog_bar2  = st.progress(0)
            prog_text2 = st.empty()
            import time as _time, threading as _threading, queue as _queue
            _start2 = _time.time()
            _pq2 = _queue.Queue()

            def _on_progress2(pct: int, msg: str):
                _pq2.put((pct, msg))

            try:
                import importlib, step2_core
                importlib.reload(step2_core)
                _result2, _error2 = [], []
                _meta2 = load_metadata()
                _inv2  = load_inventory()

                with tempfile.TemporaryDirectory() as tmp:
                    row_path2    = save_upload(row_file2, tmp, "가공파일_통합_v양식.xlsx")
                    ojc_path2    = save_upload(ojc_file2, tmp, "ojc_ref.xlsx") if ojc_file2 else None
                    ferrule_meta = get_ferrule_meta(_meta2)
                    _sales_data  = _sales_s2 if _sales_s2 else None

                    def _worker2():
                        try:
                            _result2.append(step2_core.run(
                                row_path2, None, ojc_path2, None, settings,
                                ferrule_meta=ferrule_meta, progress_cb=_on_progress2,
                                cable_meta_in=_meta2.get('cable', {}),
                                housing_meta_in=_meta2.get('housing', {}),
                                inventory=_inv2,
                                sales_data=_sales_data))
                        except Exception as _e:
                            _error2.append(_e)

                    _t2 = _threading.Thread(target=_worker2, daemon=True)
                    _t2.start()

                    _last_pct2, _last_msg2 = 0, "시작 중..."
                    while _t2.is_alive():
                        try:
                            while True:
                                _last_pct2, _last_msg2 = _pq2.get_nowait()
                        except _queue.Empty:
                            pass
                        elapsed2 = _time.time() - _start2
                        prog_bar2.progress(min(_last_pct2, 100))
                        prog_text2.markdown(
                            f'<div style="font-size:0.85rem;color:#555;margin-top:4px">'
                            f'⏳ {_last_msg2}'
                            f'<span style="float:right;color:#aaa;font-size:0.8rem">'
                            f'경과 {elapsed2:.1f}초</span></div>',
                            unsafe_allow_html=True,
                        )
                        _time.sleep(0.1)

                if _error2:
                    raise _error2[0]

                result_bytes2, logs2 = _result2[0]
                total2 = _time.time() - _start2
                st.session_state['s3_result'] = {
                    'result_bytes': result_bytes2,
                    'logs': logs2,
                    'total': total2,
                    'fname': f"연간발주계획_{ts()}.xlsx",
                }
                st.rerun()

            except Exception as e:
                prog_bar2.progress(0)
                prog_text2.empty()
                st.error(f"오류 발생: {e}")

    # ── 결과 표시 (session_state 유지) ───────────────────────
    if 's3_result' in st.session_state:
        r2 = st.session_state['s3_result']
        st.markdown(
            f'<div style="font-size:0.9rem;font-weight:600;color:#1a7a3c;'
            f'background:#e8f5e9;padding:6px 12px;border-radius:6px;'
            f'border-left:4px solid #1a7a3c;margin-top:6px">'
            f'✅ STEP 3 완료 — 발주계획 생성! &nbsp;&nbsp;'
            f'<span style="font-weight:400;font-size:0.85rem;color:#2e7d32">'
            f'총 소요시간 {r2["total"]:.1f}초</span></div>',
            unsafe_allow_html=True,
        )
        st.download_button(
            label="⬇ 연간발주계획.xlsx 다운로드",
            data=r2['result_bytes'],
            file_name=r2['fname'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="s3_dl_result",
        )
        log_text2 = "\n".join(r2['logs'])
        st.markdown(f'<div class="log-box">{log_text2}</div>', unsafe_allow_html=True)
        st.markdown("💡 노란색 셀(2026 목표 발주량)에 목표량을 입력하면 필요 발주량이 자동 계산됩니다.")

# ═══════════════════════════════════════════════════════════
# 품번 관리
# ═══════════════════════════════════════════════════════════
with tab4:
    st.markdown("### 📋 품번 관리")

    # ── 양식 다운로드 / 업로드 ──────────────────────────────
    with st.expander("📋 Excel 양식으로 일괄 관리", expanded=False):
        st.caption("현재 저장된 데이터를 Excel로 내려받아 수정한 뒤 다시 업로드하면 한 번에 반영됩니다.")
        _meta_for_dl = load_metadata()
        dl_col, ul_col = st.columns([1, 2])
        with dl_col:
            st.download_button(
                "📥 현재 데이터 Excel 다운로드",
                data=build_meta_excel(_meta_for_dl),
                file_name=f"품번관리_{ts()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="meta_dl",
            )
        with ul_col:
            meta_upload = st.file_uploader(
                "작성된 양식 업로드 (케이블 / 하우징 / 페롤 시트)",
                type=['xlsx'], key='meta_upload',
            )
            if meta_upload:
                if st.button("✅ 업로드 내용으로 저장", key="meta_upload_save", type="primary"):
                    parsed, errs = parse_meta_excel(meta_upload.getvalue())
                    if errs:
                        for e in errs: st.warning(e)
                    cur = load_metadata()
                    # 파싱 결과가 빈 dict이면 기존 데이터를 지우지 않고 경고
                    blocked = []
                    for section in ('cable', 'housing', 'ferrule'):
                        if section in parsed and len(parsed[section]) == 0 and len(cur.get(section, {})) > 0:
                            blocked.append(section)
                            parsed.pop(section)
                    if blocked:
                        st.warning(f"⚠️ {'/'.join(blocked)} 시트에서 데이터를 읽지 못했습니다. 기존 데이터를 유지합니다. 앱에서 다운로드한 양식을 사용하고 있는지 확인하세요.")
                    cur.update(parsed)
                    save_metadata(cur)
                    st.toast("✅ 양식 업로드 완료! 품번 데이터가 업데이트됐습니다.", icon="📋")
                    st.rerun()

    st.divider()
    st.info(
        "STEP 1 실행 후 발견된 신규 타입이 자동으로 추가됩니다. "
        "품번·품명·구매처·리드타임을 직접 입력하고 **저장** 버튼을 누르세요. "
        "다음 STEP 1 실행 시 자동으로 적용됩니다."
    )
    st.markdown("""
> **표 사용법**
> - ✏️ **수정**: 셀을 클릭하면 바로 편집할 수 있습니다.
> - ➕ **행 추가**: 표 하단 **＋ Add row** 버튼을 클릭하세요.
> - 🗑️ **행 삭제**: 행 왼쪽 체크박스 선택 → 키보드 **Delete** 키 또는 휴지통 아이콘 클릭.
> - 수정 후 반드시 아래 **💾 품번 저장** 버튼을 눌러야 반영됩니다.
""")

    metadata = load_metadata()

    # ── 케이블 ──────────────────────────────────────────────
    st.markdown("#### 케이블")
    with st.expander("입력 예시 보기"):
        st.dataframe(pd.DataFrame([{
            '파이': '2.0mm', '케이블종류': 'B3-SP',
            '품번': 'N90-29-1100', '품명': 'OPTICAL FIBER CABLE, B3, SP, 2.0MM',
            '구매처': 'CCTC', '리드타임': 40,
        }]), hide_index=True, use_container_width=True)
    cable_rows = []
    for k, v in sorted(metadata.get('cable', {}).items()):
        pai, ct = k.split('|', 1) if '|' in k else (k, '')
        cable_rows.append({
            '파이': pai,
            '케이블종류': ct,
            '품번': v.get('품번') or '',
            '품명': v.get('품명') or '',
            '구매처': v.get('구매처') or '',
            '리드타임': int(v['리드타임']) if v.get('리드타임') else None,
            '상태': '✓' if v.get('품번') else '⚠ 미입력',
        })

    if not cable_rows:
        df_cable = pd.DataFrame(columns=['파이','케이블종류','품번','품명','구매처','리드타임'])
    else:
        df_cable = pd.DataFrame([{k: v for k, v in r.items() if k != '상태'} for r in cable_rows])

    edited_cable = st.data_editor(
        df_cable,
        key='cable_editor',
        column_config={
            '파이':      st.column_config.SelectboxColumn('매칭 하우징 크기', options=['2.0mm','3.0mm','0.9mm'], width='small', required=True),
            '케이블종류': st.column_config.TextColumn('케이블종류', width='medium'),
            '품번':      st.column_config.TextColumn('품번', width='medium'),
            '품명':      st.column_config.TextColumn('품명', width='large'),
            '구매처':    st.column_config.TextColumn('구매처', width='medium'),
            '리드타임':  st.column_config.NumberColumn('리드타임', min_value=1, max_value=365, step=1, format='%d일'),
        },
        num_rows='dynamic',
        hide_index=True,
        use_container_width=True,
    )

    # ── 하우징 ──────────────────────────────────────────────
    st.markdown("#### 하우징")
    st.caption("하나의 하우징 타입에 여러 부품(하우징 본체·스프링·더스트캡 등)이 있으면 행을 추가하세요.")
    with st.expander("입력 예시 보기"):
        st.dataframe(pd.DataFrame([
            {'파이': '2.0mm', '하우징타입': 'LC/PC 청색', '품번': 'N90-29-1213',  '품명': 'FERRULE (W/ FLANGE,LC/PC TYPE)',          '구매처': 'CCTC',     '리드타임': 40},
            {'파이': '2.0mm', '하우징타입': 'LC/PC 청색', '품번': 'P14-LS-4228',  '품명': 'OJC HOUSING KIT (OJC LC/PC 2.0 BLUE)',    '구매처': 'FIBERCAN', '리드타임': 60},
            {'파이': '2.0mm', '하우징타입': 'LC/PC 청색', '품번': 'N94-1-04075',  '품명': 'SPRING',                                   '구매처': '상영스프링','리드타임': 30},
            {'파이': '2.0mm', '하우징타입': 'SC/PC 청색', '품번': 'N90-29-1315',  '품명': 'FERRULE (W/ FLANGE,SC/PC TYPE)',           '구매처': 'CCTC',     '리드타임': 40},
            {'파이': '2.0mm', '하우징타입': 'SC/APC 녹색','품번': 'N90-29-1323',  '품명': 'FERRULE (W/ FLANGE,SC/APC TYPE)',          '구매처': 'CCTC',     '리드타임': 40},
            {'파이': '2.0mm', '하우징타입': 'LC/PC 청색', '품번': 'P15-RM-4021',  '품명': 'DUST CAP(PC)',                             '구매처': 'FIBERCAN', '리드타임': 60},
        ]), hide_index=True, use_container_width=True)
        st.caption("같은 하우징 타입에 부품이 여러 개면 매칭 하우징 크기·하우징타입을 동일하게 입력하고 행을 추가합니다.")
    housing_rows = []
    for k, v in sorted(metadata.get('housing', {}).items()):
        pai, htype = k.split('|', 1) if '|' in k else (k, '')
        items = v if isinstance(v, list) else [v]
        for item in items:
            housing_rows.append({
                '파이': pai,
                '하우징타입': htype,
                '품번': item.get('품번') or '',
                '품명': item.get('품명') or '',
                '구매처': item.get('구매처') or '',
                '리드타임': int(item['리드타임']) if item.get('리드타임') else None,
            })

    if housing_rows:
        df_housing = pd.DataFrame(housing_rows)
    else:
        df_housing = pd.DataFrame(columns=['파이','하우징타입','품번','품명','구매처','리드타임'])

    edited_housing = st.data_editor(
        df_housing,
        key='housing_editor',
        column_config={
            '파이':      st.column_config.SelectboxColumn('파이', options=['2.0mm','3.0mm','0.9mm'], width='small', required=True),
            '하우징타입': st.column_config.TextColumn('하우징타입', width='medium'),
            '품번':      st.column_config.TextColumn('품번', width='medium'),
            '품명':      st.column_config.TextColumn('품명', width='large'),
            '구매처':    st.column_config.TextColumn('구매처', width='medium'),
            '리드타임':  st.column_config.NumberColumn('리드타임', min_value=1, max_value=365, step=1, format='%d일'),
        },
        num_rows='dynamic',
        hide_index=True,
        use_container_width=True,
    )

    # ── 페롤 ────────────────────────────────────────────────
    st.markdown("#### 페롤 (커넥터 타입별 공용)")
    st.caption("매칭 하우징 크기에 관계없이 커넥터 타입이 같으면 동일 페롤을 사용합니다. STEP 2 발주 집계 시 자동 합산됩니다.")
    ferrule_meta_cur = get_ferrule_meta(metadata)
    ferrule_rows = [
        {
            '커넥터타입': t,
            '품번':  ferrule_meta_cur[t].get('품번', ''),
            '품명':  ferrule_meta_cur[t].get('품명', ''),
            '구매처': ferrule_meta_cur[t].get('구매처', ''),
            '리드타임': int(ferrule_meta_cur[t]['리드타임']) if ferrule_meta_cur[t].get('리드타임') else None,
        }
        for t in FERRULE_TYPES
    ]
    edited_ferrule = st.data_editor(
        pd.DataFrame(ferrule_rows),
        key='ferrule_editor',
        column_config={
            '커넥터타입': st.column_config.TextColumn('커넥터타입', width='small', disabled=True),
            '품번':      st.column_config.TextColumn('품번', width='medium'),
            '품명':      st.column_config.TextColumn('품명', width='large'),
            '구매처':    st.column_config.TextColumn('구매처', width='medium'),
            '리드타임':  st.column_config.NumberColumn('리드타임', min_value=1, max_value=365, step=1, format='%d일'),
        },
        num_rows='fixed',
        hide_index=True,
        use_container_width=True,
    )

    st.divider()
    if st.button("💾 품번 저장", type="primary"):
        new_meta = load_metadata()

        if edited_cable is not None and not edited_cable.empty:
            cable_dict = {}
            for _, row in edited_cable.iterrows():
                if not row.get('파이') or not row.get('케이블종류'): continue
                k = f"{row['파이']}|{row['케이블종류']}"
                cable_dict[k] = {
                    '품번':   str(row['품번']).strip()  if pd.notna(row['품번'])   else '',
                    '품명':   str(row['품명']).strip()  if pd.notna(row['품명'])   else '',
                    '구매처': str(row['구매처']).strip() if pd.notna(row['구매처']) else '',
                    '리드타임': int(row['리드타임']) if pd.notna(row['리드타임']) else None,
                }
            new_meta['cable'] = cable_dict

        if edited_housing is not None and not edited_housing.empty:
            housing_dict = {}
            for _, row in edited_housing.iterrows():
                if not row.get('파이') or not row.get('하우징타입'): continue
                k = f"{row['파이']}|{row['하우징타입']}"
                housing_dict.setdefault(k, []).append({
                    '품번':   str(row['품번']).strip()  if pd.notna(row['품번'])   else '',
                    '품명':   str(row['품명']).strip()  if pd.notna(row['품명'])   else '',
                    '구매처': str(row['구매처']).strip() if pd.notna(row['구매처']) else '',
                    '리드타임': int(row['리드타임']) if pd.notna(row['리드타임']) else None,
                })
            new_meta['housing'] = housing_dict

        if edited_ferrule is not None and not edited_ferrule.empty:
            ferrule_dict = {}
            for _, row in edited_ferrule.iterrows():
                t = row.get('커넥터타입', '')
                if not t: continue
                ferrule_dict[t] = {
                    '품번':   str(row['품번']).strip()  if pd.notna(row['품번'])   else '',
                    '품명':   str(row['품명']).strip()  if pd.notna(row['품명'])   else '',
                    '구매처': str(row['구매처']).strip() if pd.notna(row['구매처']) else '',
                    '리드타임': int(row['리드타임']) if pd.notna(row['리드타임']) else None,
                }
            new_meta['ferrule'] = ferrule_dict

        save_metadata(new_meta)
        n_c = len(new_meta.get('cable', {}))
        n_h = len(new_meta.get('housing', {}))
        st.toast(f"✅ 저장 완료! 케이블 {n_c}건 / 하우징 {n_h}건 저장됐습니다.", icon="💾")
        st.rerun()

# ═══════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════
# 재고 현황
# ═══════════════════════════════════════════════════════════
with tab5:
    st.markdown("### 📦 재고 현황")

    # ── 양식 다운로드 / 업로드 ──────────────────────────────
    with st.expander("📋 Excel 양식으로 일괄 관리", expanded=False):
        st.caption("현재 재고 데이터를 Excel로 내려받아 수정한 뒤 다시 업로드하면 한 번에 반영됩니다.")
        _meta_for_inv_dl = load_metadata()
        _inv_for_dl      = load_inventory()
        dl_col2, ul_col2 = st.columns([1, 2])
        with dl_col2:
            st.download_button(
                "📥 현재 재고 Excel 다운로드",
                data=build_inventory_excel(_meta_for_inv_dl, _inv_for_dl),
                file_name=f"재고현황_{ts()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="inv_dl",
            )
        with ul_col2:
            inv_upload = st.file_uploader(
                "작성된 양식 업로드 (케이블 / 하우징 시트)",
                type=['xlsx'], key='inv_upload',
            )
            if inv_upload:
                if st.button("✅ 업로드 내용으로 저장", key="inv_upload_save", type="primary"):
                    parsed_inv, errs_inv = parse_inventory_excel(inv_upload.getvalue())
                    if errs_inv:
                        for e in errs_inv: st.warning(e)
                    save_inventory(parsed_inv)
                    st.toast("✅ 재고 양식 업로드 완료!", icon="📦")
                    st.rerun()

    st.divider()
    st.info(
        "현재고와 기발주 수량을 입력하고 **💾 재고 저장** 버튼을 누르면 "
        "STEP 2 실행 시 자동으로 반영됩니다."
    )

    _meta_inv = load_metadata()
    _inv      = load_inventory()

    # ── 케이블 재고 ─────────────────────────────────────────
    st.markdown("#### 케이블 재고")
    cable_inv_rows = []
    for k, v in sorted(_meta_inv.get('cable', {}).items()):
        pai, ct = k.split('|', 1) if '|' in k else (k, '')
        iv = _inv.get('cable', {}).get(k, {})
        cable_inv_rows.append({
            '파이': pai, '케이블종류': ct,
            '품번': v.get('품번') or '',
            '현재고 (m)': int(iv.get('현재고') or 0),
            '기발주 (m)': int(iv.get('기발주') or 0),
        })

    if cable_inv_rows:
        df_cinv = pd.DataFrame(cable_inv_rows)
        edited_cinv = st.data_editor(
            df_cinv,
            column_config={
                '파이':        st.column_config.TextColumn('파이',       disabled=True, width=75),
                '케이블종류':  st.column_config.TextColumn('케이블종류', disabled=True, width=110),
                '품번':        st.column_config.TextColumn('품번',       disabled=True, width=140),
                '현재고 (m)':  st.column_config.NumberColumn('현재고 (m)', min_value=0, step=1, width=100),
                '기발주 (m)':  st.column_config.NumberColumn('기발주 (m)', min_value=0, step=1, width=100),
            },
            hide_index=True, use_container_width=True, key='inv_cable',
        )
    else:
        st.caption("품번 관리 탭에 케이블 데이터가 없습니다. STEP 1을 먼저 실행하세요.")
        edited_cinv = pd.DataFrame()

    # ── 하우징 재고 ─────────────────────────────────────────
    st.markdown("#### 하우징 재고")
    housing_inv_rows = []
    for k, meta_list in sorted(_meta_inv.get('housing', {}).items()):
        pai, htype = k.split('|', 1) if '|' in k else (k, '')
        items = meta_list if isinstance(meta_list, list) else [meta_list]
        inv_list = _inv.get('housing', {}).get(k, [])
        for i, meta in enumerate(items):
            iv = inv_list[i] if i < len(inv_list) else {}
            housing_inv_rows.append({
                '파이': pai, '하우징타입': htype,
                '품번': meta.get('품번') or '',
                '현재고 (EA)': int(iv.get('현재고') or 0),
                '기발주 (EA)': int(iv.get('기발주') or 0),
            })

    if housing_inv_rows:
        df_hinv = pd.DataFrame(housing_inv_rows)
        edited_hinv = st.data_editor(
            df_hinv,
            column_config={
                '파이':         st.column_config.TextColumn('파이',        disabled=True, width=75),
                '하우징타입':   st.column_config.TextColumn('하우징타입',  disabled=True, width=130),
                '품번':         st.column_config.TextColumn('품번',        disabled=True, width=140),
                '현재고 (EA)':  st.column_config.NumberColumn('현재고 (EA)', min_value=0, step=1, width=100),
                '기발주 (EA)':  st.column_config.NumberColumn('기발주 (EA)', min_value=0, step=1, width=100),
            },
            hide_index=True, use_container_width=True, key='inv_housing',
        )
    else:
        st.caption("품번 관리 탭에 하우징 데이터가 없습니다. STEP 1을 먼저 실행하세요.")
        edited_hinv = pd.DataFrame()

    if st.button("💾 재고 저장", type="primary", key="inv_save"):
        from collections import defaultdict as _dd
        new_inv = {'cable': {}, 'housing': {}}

        if not edited_cinv.empty:
            for _, row in edited_cinv.iterrows():
                k = f"{row['파이']}|{row['케이블종류']}"
                new_inv['cable'][k] = {
                    '현재고': int(row['현재고 (m)'] or 0),
                    '기발주': int(row['기발주 (m)'] or 0),
                }

        if not edited_hinv.empty:
            housing_grouped = _dd(list)
            for _, row in edited_hinv.iterrows():
                k = f"{row['파이']}|{row['하우징타입']}"
                housing_grouped[k].append({
                    '품번': row['품번'],
                    '현재고': int(row['현재고 (EA)'] or 0),
                    '기발주': int(row['기발주 (EA)'] or 0),
                })
            new_inv['housing'] = dict(housing_grouped)

        save_inventory(new_inv)
        st.toast("✅ 재고 저장 완료! STEP 2 실행 시 자동 반영됩니다.", icon="💾")
        st.rerun()

# ═══════════════════════════════════════════════════════════
# 판매 분석
# ═══════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="step-box"><b>STEP 2 — 판매 분석</b>: 전체 판매량에서 OJC 완제품만 필터링하고, 구매관리(맥산) 납품 실적과 합산해 <b>생산비중</b>을 계산합니다.<br>저장된 데이터는 <b>STEP 3 — 발주계획 생성</b> 시 수요 기반 분석(제안량·트렌드·위험도) 컬럼에 자동 반영됩니다.</div>', unsafe_allow_html=True)

    sa_col1, sa_col2 = st.columns(2)
    with sa_col1:
        sales_raw_file = st.file_uploader(
            "① 전체 판매량.xlsx (ERP 원본)",
            type=['xlsx'], key='sa_sales',
            help="판매 ERP에서 추출한 전체 판매량 파일. 컬럼: 년, 월, 일, 품목코드, 품목명, 규격명, 수량, 거래처, 창고"
        )
    with sa_col2:
        purchase_raw_file = st.file_uploader(
            "② 구매관리(맥산).xlsx (ERP 원본)",
            type=['xlsx'], key='sa_purchase',
            help="구매관리 ERP에서 거래처=맥산텔레콤으로 필터링한 파일. 컬럼: 구매처명, 입고일자, 품목코드, 품목명, 수량, ..."
        )

    st.divider()

    sa_btn_col, sa_clr_col = st.columns([5, 1])
    with sa_btn_col:
        sa_run = st.button(
            "▶ STEP 2 실행 — 판매 분석",
            type="primary",
            disabled=(sales_raw_file is None or purchase_raw_file is None),
        )
    with sa_clr_col:
        if st.button("🗑 초기화", key="sa_clear",
                     disabled=('s2_result' not in st.session_state)):
            del st.session_state['s2_result']
            save_sales_analysis({})
            st.rerun()

    if sa_run:
        prog_bar_sa  = st.progress(0)
        prog_text_sa = st.empty()
        import time as _time, threading as _threading, queue as _queue
        _start_sa = _time.time()
        _pq_sa = _queue.Queue()

        def _on_progress_sa(pct, msg=''):
            _pq_sa.put((pct, msg))

        _result_sa, _error_sa = [], []

        def _worker_sa():
            try:
                _result_sa.append(build_sales_analysis(
                    sales_raw_file.getvalue(),
                    purchase_raw_file.getvalue(),
                    progress_cb=_on_progress_sa,
                ))
            except Exception as _e:
                _error_sa.append(_e)

        _t_sa = _threading.Thread(target=_worker_sa, daemon=True)
        _t_sa.start()
        _last_pct_sa, _last_msg_sa = 0, "시작 중..."
        while _t_sa.is_alive():
            try:
                while True:
                    _last_pct_sa, _last_msg_sa = _pq_sa.get_nowait()
            except _queue.Empty:
                pass
            elapsed_sa = _time.time() - _start_sa
            prog_bar_sa.progress(min(_last_pct_sa, 100))
            prog_text_sa.markdown(
                f'<div style="font-size:0.85rem;color:#555;margin-top:4px">'
                f'⏳ {_last_msg_sa}'
                f'<span style="float:right;color:#aaa;font-size:0.8rem">'
                f'경과 {elapsed_sa:.1f}초</span></div>',
                unsafe_allow_html=True,
            )
            _time.sleep(0.1)

        if _error_sa:
            prog_bar_sa.progress(0); prog_text_sa.empty()
            st.error(f"오류: {_error_sa[0]}")
        else:
            analysis, summary, errs, ojc_xl = _result_sa[0]
            if errs:
                for e in errs: st.warning(e)
            if analysis:
                save_sales_analysis(analysis)
                total_sa = _time.time() - _start_sa
                st.session_state['s2_result'] = {
                    'analysis': analysis,
                    'summary': summary,
                    'ojc_xl': ojc_xl,
                    'total': total_sa,
                    'fname': f"OJC_판매량_{ts()}.xlsx",
                }
                st.rerun()

    # ── 결과 표시 (session_state 유지 — 다운로드 후에도 사라지지 않음) ──
    if 's2_result' in st.session_state:
        r_sa = st.session_state['s2_result']
        sm   = r_sa['summary']
        st.markdown(
            f'<div style="font-size:0.9rem;font-weight:600;color:#1a7a3c;'
            f'background:#e8f5e9;padding:6px 12px;border-radius:6px;'
            f'border-left:4px solid #1a7a3c;margin-top:6px">'
            f'✅ STEP 2 완료 — 판매 분석! &nbsp;&nbsp;'
            f'<span style="font-weight:400;font-size:0.85rem;color:#2e7d32">'
            f'OJC {sm["ojc_rows"]:,}건 처리 · 품목 {sm["total"]}개 · '
            f'소요 {r_sa["total"]:.1f}초</span></div>',
            unsafe_allow_html=True,
        )
        if r_sa.get('ojc_xl'):
            st.download_button(
                label="⬇ OJC_판매량_필터링.xlsx 다운로드",
                data=r_sa['ojc_xl'],
                file_name=r_sa['fname'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="s2_dl_ojc",
            )

    # ── 저장된 분석 데이터 미리보기 ─────────────────────────
    _sa_cur = load_sales_analysis()
    if _sa_cur:
        n_sa = len(_sa_cur)
        n_25 = sum(1 for v in _sa_cur.values() if v.get('25', {}).get('sales', 0) > 0)
        st.markdown(f"#### 저장된 분석 데이터 — **{n_sa}** 품목 (2025년 판매 있음: **{n_25}** 품목)")

        preview_rows = []
        for code, data in sorted(_sa_cur.items()):
            for yr in ('23','24','25'):
                yd = data.get(yr, {})
                if yd.get('sales', 0) > 0 or yd.get('production', 0) > 0:
                    preview_rows.append({
                        '품목코드': code,
                        '품목명': data.get('품목명','')[:40],
                        '연도': f"20{yr}",
                        '판매량': yd.get('sales', 0),
                        '납품량(맥산)': yd.get('production', 0),
                        '생산비중': f"{yd.get('ratio', 0)*100:.1f}%",
                    })
        if preview_rows:
            st.dataframe(
                pd.DataFrame(preview_rows),
                hide_index=True,
                use_container_width=True,
                column_config={
                    '판매량':      st.column_config.NumberColumn('판매량',      format='%d'),
                    '납품량(맥산)': st.column_config.NumberColumn('납품량(맥산)', format='%d'),
                },
            )
    else:
        st.caption("아직 저장된 분석 데이터가 없습니다. 위에서 두 파일을 업로드 후 실행하세요.")

# 설정
# ═══════════════════════════════════════════════════════════
with tab6:
    st.markdown("### ⚙️ 파라미터 설정")
    st.markdown("변경 후 **저장** 버튼을 누르면 다음 실행부터 적용됩니다.")
    st.divider()

    col5, col6 = st.columns([1, 1])
    with col5:
        st.markdown("**계산 파라미터**")
        lt_val = st.number_input(
            "리드타임 기본값 (일) — 리드타임 미입력 항목에 적용",
            min_value=1, max_value=365,
            value=int(settings.get('lead_time_default', 60)),
            step=1,
        )
        st.markdown("**안전재고 계산 공식**")
        st.code("안전재고 = 3개년 피크 사용량 평균 × (리드타임 ÷ 30)", language=None)

    with col6:
        st.markdown("**Excel 헤더 색상** (hex 6자리, # 제외)")
        colors = settings.get('colors', DEFAULT_SETTINGS['colors'])
        c_main = st.text_input("메인 헤더 색상", value=colors.get('main_header', '1F3864'), max_chars=6)
        c_23   = st.text_input("2023년 컬럼 색상", value=colors.get('year_23', '2F5597'), max_chars=6)
        c_24   = st.text_input("2024년 컬럼 색상", value=colors.get('year_24', '2E75B6'), max_chars=6)
        c_25   = st.text_input("2025년 컬럼 색상", value=colors.get('year_25', '155480'), max_chars=6)

        st.markdown("미리보기")
        cols_prev = st.columns(4)
        for col_p, label, hex_val in zip(
            cols_prev,
            ['메인', '23년', '24년', '25년'],
            [c_main, c_23, c_24, c_25]
        ):
            col_p.markdown(
                f'<div style="background:#{hex_val};color:#fff;text-align:center;'
                f'padding:6px 2px;border-radius:4px;font-size:0.8rem">{label}</div>',
                unsafe_allow_html=True,
            )

    st.divider()
    if st.button("💾 설정 저장", type="primary"):
        new_settings = {
            "lead_time_default": lt_val,
            "colors": {
                "main_header": c_main,
                "year_23": c_23,
                "year_24": c_24,
                "year_25": c_25,
            }
        }
        save_settings(new_settings)
        settings.update(new_settings)
        st.success("✅ 설정이 저장됐습니다. 다음 STEP 실행부터 적용됩니다.")

    st.divider()
    st.markdown("**현재 저장된 설정**")
    st.json(load_settings())

    st.divider()
    st.markdown("**워크플로우 안내**")
    st.markdown("""
    ```
    ① 맥산 ERP에서 구매조회 또는 구매현황 파일 추출
              ↓
    ② STEP 1 실행 (ERP 원본 파일 업로드)
       → 자동 변환 후 사용내역.xlsx 다운로드
       → 신규 타입은 📋 품번 관리 탭에 자동 추가
              ↓
    ③ 📋 품번 관리 탭에서 품번·품명·구매처·리드타임 입력 후 저장
       (이후 실행부터 자동 적용)
              ↓
    ④ 사용내역.xlsx에 현재고·기발주 직접 입력
              ↓
    ⑤ STEP 2 실행 (동일 ERP 원본 파일 + 수정된 사용내역.xlsx 업로드)
       → 2026_연간발주계획.xlsx 다운로드
              ↓
    ⑥ 노란색 셀에 2026 목표 발주량 입력
    ```
    """)
