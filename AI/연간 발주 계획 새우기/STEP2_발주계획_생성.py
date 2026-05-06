#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STEP 2: 2026_연간발주계획.xlsx 자동 생성
========================================================
실행 방법:
    python STEP2_발주계획_생성.py

필요 파일 (같은 폴더):
    - 가공파일_통합_v양식.xlsx   (ROW 데이터)
    - 3개년_생산자재_사용내역.xlsx (현재고·기발주 입력 완료 버전)

생성 파일:
    - 2026_연간발주계획.xlsx
"""

import openpyxl, re, os, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DIR = os.path.dirname(os.path.abspath(__file__))
ROW_FILE   = os.path.join(DIR, '가공파일_통합_v양식.xlsx')
USAGE_FILE = os.path.join(DIR, '3개년_생산자재_사용내역.xlsx')
REF_FILE   = os.path.join(DIR, '비밀_OJC 3개년 생산자제 사용 내역_SCM팀_박정원_260326.xlsx')  # 피그테일용
OUT_FILE   = os.path.join(DIR, '2026_연간발주계획.xlsx')
YEARS = ['23','24','25']; MM_KINDS={'om1','om1-pigtail','om3'}

print("="*55)
print("  2026 연간 발주 계획 자동 생성")
print("="*55)

missing = [f for f in [ROW_FILE, USAGE_FILE] if not os.path.exists(f)]
if missing:
    print("\n❌ 아래 파일이 없습니다:")
    for f in missing: print(f"   {os.path.basename(f)}")
    print("\n먼저 STEP1을 실행하고 현재고·기발주를 입력해주세요.")
    input("\n엔터를 누르면 종료합니다..."); sys.exit(1)

print(f"✅ ROW 데이터: {os.path.basename(ROW_FILE)}")
print(f"✅ 사용내역:   {os.path.basename(USAGE_FILE)}")

# ── 스타일 헬퍼 ─────────────────────────────────────────────
def fill(h): return PatternFill('solid',start_color=h)
def font(bold=False,size=9,color='000000'): return Font(name='Arial',bold=bold,size=size,color=color)
def bdr(): s=Side(style='thin'); return Border(left=s,right=s,top=s,bottom=s)
BORDER=bdr(); CTR=Alignment(horizontal='center',vertical='center')
CTR_W=Alignment(horizontal='center',vertical='center',wrap_text=True)
LEFT=Alignment(horizontal='left',vertical='center'); RIGHT=Alignment(horizontal='right',vertical='center')
NUM_FMT='#,##0'; PCT_FMT='0.0%;[Red]-0.0%'
PAI_ORDER={'2.0mm':0,'3.0mm':1,'0.9mm':2}

def lt_days(lt):
    if not lt: return 60
    m=re.search(r'(\d+)',str(lt)); return int(m.group(1)) if m else 60
def np(p): return '0.9mm' if str(p).strip()=='0.9' else str(p).strip()
def imc(k): return bool(re.match(r'^(b3|a1)-\d+c$',str(k).lower() if k else ''))
def mc2(kind,core):
    k=str(kind).strip().lower() if kind else ''
    try: c=int(core)
    except: c=1
    sd='SP' if c==1 else('DP' if c==2 else f'{c}C')
    bm={'a1':'A1','b3':'B3','om1':'OM1','om3':'OM3','a1-청':'A1_청','a1-녹':'A1_녹','a1-적':'A1_적','a1-자':'A1_자'}
    if k in bm: return f'{bm[k]}-{sd}'
    if k=='drop': return 'DROP'
    if k in('pigtail','om1-pigtail'): return 'PIGTAIL'
    if k=='a2': return 'Optical cable'
    return k.upper()
def gp(t,kind,pai):
    if not t: return None
    t=str(t).strip(); p=np(pai); k=str(kind).strip().lower() if kind else ''
    if t=='LC/PC': return (p,'LC/PC 베이지MM') if k in MM_KINDS and p=='2.0mm' else (p,'LC/PC 청색')
    return {'LC/APC':(p,'LC/APC 녹색'),'SC/PC':(p,'SC/PC 청색'),'SC/APC':(p,'SC/APC 녹색'),'FC/PC':(p,'FC/PC 흑색'),'FC/APC':(p,'FC/APC 녹색')}.get(t)
def gr(t,pai):
    if not t: return None
    t=str(t).strip(); p=np(pai)
    return {'LC/PC':(p,'LC/PC 적색'),'LC/APC':(p,'LC/APC 적색'),'SC/PC':(p,'SC/PC 적색'),'SC/APC':(p,'SC/APC 적색'),'FC/PC':(p,'FC/PC 적색'),'FC/APC':(p,'FC/APC 적색')}.get(t)

# ── 메타데이터 로드 ─────────────────────────────────────────
print("\n메타데이터 로드 중...", end='', flush=True)
usage_wb=openpyxl.load_workbook(USAGE_FILE,data_only=True)
cable_meta={}
sn=usage_wb.sheetnames
cs_name='케이블 사용내역' if '케이블 사용내역' in sn else None
hs_name='하우징 사용내역' if '하우징 사용내역' in sn else None
if cs_name:
    for row in usage_wb[cs_name].iter_rows(min_row=4,values_only=True):
        pai,ct,bunho,pname=row[1],row[2],row[3],row[4]
        if pai and ct and bunho:
            k=(str(pai).strip(),str(ct).strip())
            if k not in cable_meta:
                # 컬럼 위치: 품번=D(4), 품명=E(5), 현재고=L(12), 구매처=P(16), 리드타임=Q(17)
                g=row[15] if len(row)>15 else None
                lt=row[16] if len(row)>16 else None
                stk=row[11] if len(row)>11 else 0
                cable_meta[k]={'품번':bunho,'품명':pname,'구매처':g,'리드타임':lt,'현재고':stk or 0}

housing_meta=defaultdict(list)
if hs_name:
    for row in usage_wb[hs_name].iter_rows(min_row=4,values_only=True):
        pai,htype,bunho,pname=row[1],row[2],row[3],row[4]
        if pai and htype and bunho:
            k=(str(pai).strip(),str(htype).strip())
            g=row[16] if len(row)>16 else None
            lt=row[17] if len(row)>17 else None
            stk=row[12] if len(row)>12 else 0
            bid=row[13] if len(row)>13 else 0
            housing_meta[k].append({'품번':bunho,'품명':pname,'구매처':g,'리드타임':lt,'현재고':stk or 0,'기발주':bid or 0})

# 피그테일 케이블 (0.9mm) — 원본 참고파일에서
pigtail_cable_meta={}; pigtail_cable_stats={}
COLOR_MAP={'연청':'청록','연등':'분홍'}
def extract_label(pname):
    pname=str(pname) if pname else ''
    if 'MM(OM4)' in pname: t='MM(OM4)'
    elif 'MM(OM3)' in pname: t='MM(OM3)'
    elif 'MM(OM1)' in pname: t='MM(OM1)'
    else: t='SM'
    m=re.search(r'\([^,]+,([^)]+)\)',pname)
    if m:
        kr=m.group(1).strip(); return f'{t}-{COLOR_MAP.get(kr,kr)}'
    return t

if os.path.exists(REF_FILE):
    ref_wb=openpyxl.load_workbook(REF_FILE,data_only=True)
    ref_sn=ref_wb.sheetnames
    pt_sheet='케이블 사용내역' if '케이블 사용내역' in ref_sn else None
    if pt_sheet:
        for row in ref_wb[pt_sheet].iter_rows(min_row=4,values_only=True):
            if row[1] and np(row[1])=='0.9mm' and row[2]=='PIGTAIL' and row[3]:
                label=extract_label(row[4]); k=('0.9mm',f'PIGTAIL-{label}')
                pigtail_cable_meta[k]={'품번':row[3],'품명':row[4],'구매처':row[12],'리드타임':row[13],'현재고':row[11] or 0}
                pigtail_cable_stats[k]={}
                for yr,ai,pi in[('23',5,6),('24',7,8),('25',9,10)]:
                    ann=round(float(row[ai])) if row[ai] else 0
                    peak=round(float(row[pi])) if row[pi] else 0
                    base=[round(ann/12)]*12
                    base[0]+=round(ann)-sum(base)
                    pigtail_cable_stats[k][yr]={'annual':ann,'peak':peak,'monthly':base}

cable_meta[('2.0mm','PIGTAIL')]={'품번':'(확인필요)','품명':'OPTICAL CABLE 2.0mm (2.0mm 자켓 피그테일용)','구매처':'(확인필요)','리드타임':'60일','현재고':0}
print(f" 완료")

# ── ROW 데이터 집계 ─────────────────────────────────────────
print("ROW 데이터 집계 중...", end='', flush=True)
row_wb=openpyxl.load_workbook(ROW_FILE,read_only=True,data_only=True)

cable_agg=defaultdict(lambda:{yr:[0.0]*12 for yr in YEARS})
for yr in YEARS:
    for row in list(row_wb[f'{yr}년_케이블'].iter_rows(values_only=True))[1:]:
        kind=row[3]; pai=row[4]; core=row[5]; length=row[6]
        if not kind or not pai or not length: continue
        ct=mc2(kind,core); p=np(pai)
        if ct=='PIGTAIL' and p=='0.9mm': continue
        for i,q in enumerate(row[9:21]):
            if q: cable_agg[(p,ct)][yr][i]+=float(q)*float(length)

housing_agg=defaultdict(lambda:{yr:[0.0]*12 for yr in YEARS})
for yr in YEARS:
    for row in list(row_wb[f'{yr}년 하우징'].iter_rows(values_only=True))[1:]:
        kind=row[3]; pai=row[4]; core=row[5]; t1=row[7]; t2=row[8]
        if not pai: continue; pai=np(pai)
        try: cps=int(core) if core else 1
        except: cps=1
        is_dp=(cps>=2) and not imc(kind)
        for mi,qty in enumerate(row[9:21]):
            if not qty: continue
            qty=float(qty); ha=gp(t1,kind,pai)
            if ha: housing_agg[ha][yr][mi]+=qty*cps
            if t2:
                hb=gr(t2,pai) if is_dp else gp(t2,kind,pai)
                if hb: housing_agg[hb][yr][mi]+=qty*cps

def fin(agg):
    r={}
    for k,yd in agg.items():
        r[k]={}
        for yr,m in yd.items():
            v=[round(x) for x in m]; r[k][yr]={'monthly':v,'annual':sum(v),'peak':max(v)}
    return r

cable_stats={**fin(cable_agg),**pigtail_cable_stats}
housing_stats=fin(housing_agg)
all_cable_meta={**cable_meta,**pigtail_cable_meta}

# ROW 데이터 누락 케이블 보완: A1-4C, OM4-DP (참고파일에서 직접 로드)
# A1-4C: 4코어 OJC 제품이 ROW에서 b3-4c로 분류되어 A1-4C 누락
# OM4-DP: ROW 집계 안됨 (소량)
OJC_REF = os.path.join(DIR, '비밀_OJC 3개년 생산자제 사용 내역_SCM팀_박정원_260326.xlsx')
REF_ONLY = {('2.0mm','A1-4C'), ('2.0mm','OM4-DP')}
if os.path.exists(OJC_REF):
    try:
        ojc_ref=openpyxl.load_workbook(OJC_REF,data_only=True)
        for row in ojc_ref['케이블 사용내역'].iter_rows(min_row=4,values_only=True):
            pai,kind,bunho,pname=row[1],row[2],row[3],row[4]
            if not pai or not kind: continue
            key=(str(pai).strip(),str(kind).strip())
            if key not in REF_ONLY or key in cable_stats: continue
            def uni_ref(a): b=[round(a/12)]*12 if a else [0]*12; b[0]+=round(a or 0)-sum(b); return b
            cable_stats[key]={}
            for yr,ai,pi in[('23',5,6),('24',7,8),('25',9,10)]:
                ann=round(float(row[ai])) if row[ai] else 0
                peak=round(float(row[pi])) if row[pi] else 0
                cable_stats[key][yr]={'monthly':uni_ref(ann),'annual':ann,'peak':peak}
            if key not in all_cable_meta and bunho:
                all_cable_meta[key]={'품번':bunho,'품명':pname,'구매처':row[12],'리드타임':row[13],'현재고':row[11] or 0}
    except Exception:
        pass

print(f" 완료 (케이블 {len(cable_stats)}타입 / 하우징 {len(housing_stats)}타입)")

# ── Excel 빌드 ──────────────────────────────────────────────
print("Excel 파일 생성 중...", end='', flush=True)
COL_WIDTHS=[5,8,20,16,38,12,10,14,14,14,14,14,14,14,14,9,9,13,12,12,15,14,20]

def write_cell(ws,row,col,val,num=False,is_input=False,bold=False,row_fill=None):
    c=ws.cell(row,col,val)
    if is_input:
        c.font=font(bold=True,size=9,color='0000FF'); c.fill=fill('FFFFC0')
        c.border=Border(left=Side(style='medium',color='C55A11'),right=Side(style='medium',color='C55A11'),top=Side(style='medium',color='C55A11'),bottom=Side(style='medium',color='C55A11'))
    else:
        c.font=font(bold=bold,size=9)
        if row_fill: c.fill=row_fill
        c.border=BORDER
    c.number_format=NUM_FMT if num else 'General'
    c.alignment=RIGHT if num else(CTR if col in[1,2,6,7] else LEFT)

def write_sheet(ws,title,stats_dict,meta_dict,unit,type_label):
    keys=sorted([k for k in stats_dict if any(stats_dict[k][yr]['annual']>0 for yr in YEARS)],key=lambda k:(PAI_ORDER.get(k[0],9),k[1]))
    ws.row_dimensions[1].height=26; ws.merge_cells('A1:W1')
    c=ws['A1']; c.value=title; c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR
    ws.row_dimensions[2].height=18
    for rng,lbl,bg,fg in[('A2:G2','기본 정보','374151','FFFFFF'),('H2:I2','2023년','2F5597','FFFFFF'),('J2:K2','2024년','2E75B6','FFFFFF'),('L2:M2','2025년','9DC3E6','1F3864'),('N2:Q2','📊 트렌드 분석','375623','FFFFFF'),('R2:R2','⚠ 안전재고','C00000','FFFFFF'),('S2:T2','재고 현황','7030A0','FFFFFF'),('U2:V2','✏ 2026 발주 계획','C55A11','FFFFFF'),('W2:W2','비고','595959','FFFFFF')]:
        s,e=rng.split(':')
        if s!=e: ws.merge_cells(rng)
        c=ws[s]; c.value=lbl; c.font=Font(name='Arial',bold=True,size=9,color=fg); c.fill=fill(bg); c.alignment=CTR; c.border=BORDER
    ws.row_dimensions[3].height=42
    hdrs=['NO','파이',type_label,'품번','품명','구매처','리드타임\n(일)',f'연간({unit})',f'피크({unit})',f'연간({unit})',f'피크({unit})',f'연간({unit})',f'피크({unit})','3개년\n평균연간','3개년\n피크평균','23→24\n증감률','24→25\n증감률',f'안전재고\n({unit})',f'현재고\n({unit})','기발주\n(참고)',f'2026목표\n({unit})',f'필요발주\n({unit})','비고']
    for ci,(h,w) in enumerate(zip(hdrs,COL_WIDTHS),1):
        c=ws.cell(3,ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
        c.fill=fill('BDD7EE'); c.alignment=CTR_W; c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ri=4; no=1
    for key in keys:
        pai,ctype=key; d=stats_dict[key]
        raw=meta_dict.get(key)
        comps=raw if isinstance(raw,list) else([raw] if raw else [{}])
        first_u=ri
        for ci2,m in enumerate(comps):
            lt=lt_days(m.get('리드타임',60) if m else 60)
            rf=fill('F5F5F5') if ri%2==0 else None
            write_cell(ws,ri,1,no if ci2==0 else '',row_fill=rf)
            write_cell(ws,ri,2,pai if ci2==0 else '',row_fill=rf)
            write_cell(ws,ri,3,ctype if ci2==0 else '',row_fill=rf)
            write_cell(ws,ri,4,m.get('품번','') if m else '',row_fill=rf)
            write_cell(ws,ri,5,m.get('품명','') if m else '',row_fill=rf)
            write_cell(ws,ri,6,m.get('구매처','') if m else '',row_fill=rf)
            write_cell(ws,ri,7,lt,row_fill=rf)
            if ci2==0:
                for col,yr,typ in[(8,'23','annual'),(9,'23','peak'),(10,'24','annual'),(11,'24','peak'),(12,'25','annual'),(13,'25','peak')]:
                    v=d[yr][typ]; write_cell(ws,ri,col,v if v else None,num=True,row_fill=rf)
                for col,fml,fmt_ in[(14,f'=ROUND(AVERAGE(H{ri},J{ri},L{ri}),0)',NUM_FMT),(15,f'=ROUND(AVERAGE(I{ri},K{ri},M{ri}),0)',NUM_FMT),(16,f'=IFERROR((J{ri}-H{ri})/H{ri},"")',PCT_FMT),(17,f'=IFERROR((L{ri}-J{ri})/J{ri},"")',PCT_FMT)]:
                    c2=ws.cell(ri,col,fml); c2.font=font(size=9); c2.border=BORDER; c2.number_format=fmt_
                    c2.alignment=RIGHT if col<16 else CTR
                    if rf: c2.fill=rf
            else:
                for col in range(8,18):
                    cl=get_column_letter(col)
                    c2=ws.cell(ri,col,f'={cl}{first_u}'); c2.font=font(size=9); c2.border=BORDER
                    c2.number_format=NUM_FMT if col<16 else PCT_FMT; c2.alignment=RIGHT if col<16 else CTR
                    if rf: c2.fill=rf
            c2=ws.cell(ri,18,f'=ROUND(O{ri}*G{ri}/30,0)'); c2.font=font(bold=True,size=9); c2.border=BORDER; c2.number_format=NUM_FMT; c2.alignment=RIGHT; c2.fill=fill('FFF2CC')
            write_cell(ws,ri,19,m.get('현재고') if m else None,num=True,row_fill=rf)
            write_cell(ws,ri,20,m.get('기발주') if m else None,num=True,row_fill=rf)
            if ci2==0:
                write_cell(ws,ri,21,None,num=True,is_input=True)
            else:
                c2=ws.cell(ri,21,f'=U{first_u}'); c2.font=font(bold=True,size=9,color='0000FF'); c2.fill=fill('EBF3FB'); c2.border=BORDER; c2.number_format=NUM_FMT; c2.alignment=RIGHT
            c2=ws.cell(ri,22,f'=IFERROR(U{ri}-IFERROR(S{ri},0)-IFERROR(T{ri},0),"")'); c2.font=font(bold=True,size=9,color='C00000'); c2.border=BORDER; c2.number_format=NUM_FMT; c2.alignment=RIGHT
            if rf: c2.fill=rf
            ws.cell(ri,23,'').border=BORDER
            if rf: ws.cell(ri,23).fill=rf
            ws.row_dimensions[ri].height=17; ri+=1
        no+=1
    last=ri-1; tr=ri+1; ws.row_dimensions[tr].height=18
    c=ws.cell(tr,3,'합  계'); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR; c.border=BORDER
    for ci in range(8,16):
        cl=get_column_letter(ci); c=ws.cell(tr,ci,f'=SUM({cl}4:{cl}{last})')
        c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('1F3864'); c.number_format=NUM_FMT; c.alignment=RIGHT; c.border=BORDER
    ws.freeze_panes='A4'; ws.sheet_view.showGridLines=False
    return keys

wb=openpyxl.Workbook(); wb.remove(wb.active)
ws_c=wb.create_sheet('케이블 사용내역')
cable_keys=write_sheet(ws_c,'2026 연간 발주 계획 — 케이블 사용내역  (단위 : m)',cable_stats,all_cable_meta,'m','케이블 종류')
ws_h=wb.create_sheet('하우징 사용내역')
housing_keys=write_sheet(ws_h,'2026 연간 발주 계획 — 하우징 사용내역  (단위 : EA)',housing_stats,housing_meta,'EA','하우징 타입')

# 월별 발주계획 시트
ws_m=wb.create_sheet('2026 월별 발주계획'); ws_m.sheet_view.showGridLines=False
ws_m.row_dimensions[1].height=26; ws_m.merge_cells('A1:T1')
c=ws_m['A1']; c.value='2026 월별 발주 계획 (과거 계절 패턴 기반 자동 분배)'
c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR
ws_m.row_dimensions[2].height=18
for rng,lbl,bg,fg in[('A2:F2','기본정보','374151','FFFFFF'),('G2:G2','연간목표','C55A11','FFFFFF'),('H2:S2','월별 발주 (연간목표 × 과거 패턴)','2E75B6','FFFFFF'),('T2:T2','합계검증','375623','FFFFFF')]:
    s,e=rng.split(':')
    if s!=e: ws_m.merge_cells(rng)
    c=ws_m[s]; c.value=lbl; c.font=Font(name='Arial',bold=True,size=9,color=fg); c.fill=fill(bg); c.alignment=CTR; c.border=BORDER
ws_m.row_dimensions[3].height=40
for ci,h in enumerate(['NO','분류','파이','종류','품번','단위','연간목표']+[f'{i}월' for i in range(1,13)]+['합계검증'],1):
    c=ws_m.cell(3,ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
    c.fill=fill('BDD7EE'); c.alignment=CTR_W; c.border=BORDER
for ci,w in enumerate([5,10,8,20,16,6,14]+[9]*12+[10],1):
    ws_m.column_dimensions[get_column_letter(ci)].width=w
ri=4
all_items=[('케이블',k,cable_stats[k]) for k in cable_keys]+[('하우징',k,housing_stats[k]) for k in housing_keys]
for idx,(cat,key,d) in enumerate(all_items):
    pai,ctype=key; rf=fill('F5F5F5') if ri%2==0 else None
    combined=[0.0]*12; total=0.0
    for yr in YEARS:
        for i,v in enumerate(d[yr]['monthly']): combined[i]+=v; total+=v
    ratios=[round(v/total,6) if total>0 else round(1/12,6) for v in combined]
    unit='m' if cat=='케이블' else 'EA'
    for ci2,v in enumerate([idx+1,cat,pai,ctype,'',unit],1):
        c=ws_m.cell(ri,ci2,v); c.font=font(size=9); c.alignment=CTR if ci2<=3 else LEFT; c.border=BORDER
        if rf: c.fill=rf
    c=ws_m.cell(ri,7,None); c.font=font(bold=True,size=9,color='0000FF'); c.fill=fill('FFFFC0')
    c.border=Border(left=Side(style='medium',color='C55A11'),right=Side(style='medium',color='C55A11'),top=Side(style='medium',color='C55A11'),bottom=Side(style='medium',color='C55A11'))
    c.number_format=NUM_FMT; c.alignment=RIGHT
    for mi,ratio in enumerate(ratios):
        c=ws_m.cell(ri,8+mi,f'=IFERROR(ROUND($G{ri}*{ratio},0),"")'); c.font=font(size=9); c.border=BORDER; c.number_format=NUM_FMT; c.alignment=RIGHT
        if rf: c.fill=rf
    c=ws_m.cell(ri,20,f'=IFERROR(SUM(H{ri}:S{ri}),"")'); c.font=font(size=9,color='375623',bold=True); c.border=BORDER; c.number_format=NUM_FMT; c.alignment=RIGHT
    if rf: c.fill=rf
    ws_m.row_dimensions[ri].height=17; ri+=1
ws_m.freeze_panes='A4'

# 이상항목 시트
ws_a=wb.create_sheet('⚠ 이상항목 검토'); ws_a.sheet_view.showGridLines=False
ws_a.row_dimensions[1].height=26; ws_a.merge_cells('A1:D1')
c=ws_a['A1']; c.value='데이터 이상 항목 검토 (자동 분석)'; c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF'); c.fill=fill('C00000'); c.alignment=CTR
ws_a.row_dimensions[3].height=30
for ci,h in enumerate(['구분','항목','품번','내용 및 조치 권고'],1):
    c=ws_a.cell(3,ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('BDD7EE'); c.alignment=CTR_W; c.border=BORDER
ws_a.column_dimensions['A'].width=10; ws_a.column_dimensions['B'].width=22; ws_a.column_dimensions['C'].width=16; ws_a.column_dimensions['D'].width=65
anomalies=[
    ('주의','2.0mm 자켓 피그테일','(확인필요)','참고파일에 품번 미등재. 사용 케이블 품번 확인 필요.','orange'),
    ('정보','OM4 피그테일 케이블','P14-RM-417K','3개년 사용 없음. 재고 보유. 단종 검토 필요.','red'),
    ('정보','OM3 피그테일 케이블','P14-RM-417H','23년 648m → 24~25년 0m. 미사용 추세.','orange'),
    ('정보','피그테일 전체','(전 색상)','23년 대비 25년 약 81% 급감. 2026 목표량 보수적 설정 권고.','blue'),
]
cm={'red':'FFD7D7','orange':'FFE6C8','blue':'D7E8FF'}
for ri2,(_t,item,bunho,desc,color) in enumerate(anomalies,start=4):
    rf=PatternFill('solid',start_color=cm.get(color,'FFFFFF'))
    for ci,v in enumerate([_t,item,bunho,desc],1):
        c=ws_a.cell(ri2,ci,v); c.font=Font(name='Arial',size=9,bold=(ci==1)); c.fill=rf; c.border=BORDER
        c.alignment=CTR if ci==1 else(Alignment(horizontal='left',vertical='center',wrap_text=True) if ci==4 else LEFT)
    ws_a.row_dimensions[ri2].height=36
ws_a.freeze_panes='A4'


wb.save(OUT_FILE)
print(f" 완료!")

# ── 수요 기반 분석 (생산_판매_비교.xlsx 있을 때 자동 포함) ──────
# 필요 파일: 가공파일_통합_v양식.xlsx (이미 로드됨) + 생산_판매_비교.xlsx
CMP_FILE = os.path.join(DIR, '생산_판매_비교.xlsx')
if os.path.exists(CMP_FILE):
    print("\n수요 기반 분석 추가 중...", end='', flush=True)
    try:
        cmp_wb2=openpyxl.load_workbook(CMP_FILE,read_only=True,data_only=True)
        prod_data2={}
        for row in list(cmp_wb2['연간_요약'].iter_rows(values_only=True))[3:]:
            if not row[1]: continue
            code=str(row[1]).strip()
            prod_data2[code]={'23':{'sales':row[4] or 0,'ratio':row[7] or 0},'24':{'sales':row[8] or 0,'ratio':row[11] or 0},'25':{'sales':row[12] or 0,'ratio':row[15] or 0}}
        row_wb2=openpyxl.load_workbook(ROW_FILE,read_only=True,data_only=True)
        cable_use2={}; housing_use2={}
        for yr2,cs2,hs2 in[('25','25년_케이블','25년 하우징'),('24','24년_케이블','24년 하우징'),('23','23년_케이블','23년 하우징')]:
            for row in list(row_wb2[cs2].iter_rows(values_only=True))[1:]:
                code=str(row[0]) if row[0] else ''
                if not code or code in cable_use2: continue
                kind=row[3]; pai=row[4]; core=row[5]; length=row[6]
                if not kind or not pai or not length: continue
                ct=mc2(kind,core); p=np(pai)
                if ct=='PIGTAIL' and p=='0.9mm': continue
                cable_use2[code]=(p,ct,float(length))
            for row in list(row_wb2[hs2].iter_rows(values_only=True))[1:]:
                code=str(row[0]) if row[0] else ''
                if not code or code in housing_use2: continue
                kind=row[3]; pai=row[4]; core=row[5]; t1=row[7]; t2=row[8]
                if not pai: continue
                p=np(pai)
                try: cps=int(core) if core else 1
                except: cps=1
                is_dp2=(cps>=2) and not imc(kind); h={}
                ha=gp(t1,kind,p)
                if ha: h[ha]=h.get(ha,0)+cps
                if t2:
                    hb=gr(t2,p) if is_dp2 else gp(t2,kind,p)
                    if hb: h[hb]=h.get(hb,0)+cps
                if h: housing_use2[code]=h
        def calc_trend2(d):
            s23=d['23']['sales']; s25=d['25']['sales']
            if s23>0 and s25>0: return (s25/s23)**0.5-1
            if s25>0 and d['24']['sales']>0: return s25/d['24']['sales']-1
            return 0.0
        cp2=defaultdict(lambda:[0,0,0,0,0]); hp2=defaultdict(lambda:[0,0,0,0,0])
        for code,(pai,ct,length) in cable_use2.items():
            if code not in prod_data2: continue
            d=prod_data2[code]; trend=min(max(calc_trend2(d),-0.5),1.0)
            ratio=min(d['25']['ratio'],1.5); s25=d['25']['sales']
            if s25==0: continue
            key=(pai,ct); p2=cp2[key]
            p2[0]+=s25*(1+trend)*ratio*length; p2[1]+=s25*ratio; p2[2]+=s25; p2[3]+=s25*trend; p2[4]+=s25
        for code,h_map in housing_use2.items():
            if code not in prod_data2: continue
            d=prod_data2[code]; trend=min(max(calc_trend2(d),-0.5),1.0)
            ratio=min(d['25']['ratio'],1.5); s25=d['25']['sales']
            if s25==0: continue
            for (pai,htype),qpu in h_map.items():
                key=(pai,htype); p2=hp2[key]
                p2[0]+=s25*(1+trend)*ratio*qpu; p2[1]+=s25*ratio; p2[2]+=s25; p2[3]+=s25*trend; p2[4]+=s25
        def fp2(p): return round(p[0]),(p[1]/p[2] if p[2]>0 else None),(p[3]/p[4] if p[4]>0 else None)
        cf2={k:fp2(v) for k,v in cp2.items()}; hf2={k:fp2(v) for k,v in hp2.items()}
        pwb2=openpyxl.load_workbook(OUT_FILE)
        def adc(ws2,dd2):
            sc=24; N2=14
            ws2.merge_cells(f'{get_column_letter(sc)}2:{get_column_letter(sc+4)}2')
            c=ws2[f'{get_column_letter(sc)}2']; c.value='📊 수요 기반 분석 (판매량 + 생산비중)'
            c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('843C0C'); c.alignment=CTR; c.border=BORDER
            for ci,h in enumerate(['2026 제안량\n(수요기반)','vs 3개년평균\n비교','평균\n생산비중','판매\n트렌드','수입의존\n위험도']):
                c=ws2.cell(3,sc+ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('843C0C')
                c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BORDER
                ws2.column_dimensions[get_column_letter(sc+ci)].width=13
            ck=None; cf=None
            for ri in range(4,ws2.max_row+1):
                pv=ws2.cell(ri,2).value; tv=ws2.cell(ri,3).value
                if str(tv or '').strip()=='합  계': continue
                if pv and tv: ck=(str(pv).strip(),str(tv).strip()); cf=ri
                if not ck: continue
                proj,ratio,trend=dd2.get(ck,(None,None,None))
                rf=fill('F5F5F5') if ri%2==0 else None
                c=ws2.cell(ri,sc,round(proj) if proj else None)
                c.font=Font(name='Arial',bold=True,size=9,color='1F3864'); c.border=BORDER; c.number_format='#,##0'; c.alignment=RIGHT; c.fill=fill('FFF2CC')
                expr=f'=IFERROR({get_column_letter(sc)}{ri}/{get_column_letter(N2)}{ri}-1,"")' if ri==cf else f'={get_column_letter(sc+1)}{cf}'
                c=ws2.cell(ri,sc+1,expr); c.font=font(size=9); c.border=BORDER; c.number_format='0%;[Red]-0%'; c.alignment=CTR; c.fill=fill('EBF3FB')
                rv=round(ratio,3) if ratio else None
                c=ws2.cell(ri,sc+2,rv); c.font=font(size=9); c.border=BORDER; c.number_format='0%'; c.alignment=CTR
                c.fill=fill('D7F0D7') if rv and rv>=0.8 else(fill('FFF2CC') if rv and rv>=0.5 else(fill('FFD7D7') if rv else(rf if rf else PatternFill())))
                tv2=round(trend,4) if trend is not None else None
                c=ws2.cell(ri,sc+3,tv2)
                c.font=Font(name='Arial',size=9,color='1A5C1A' if (tv2 or 0)>0 else 'C00000')
                c.border=BORDER; c.number_format='+0.0%;[Red]-0.0%'; c.alignment=CTR; c.fill=fill('D7E8FF')
                if rv is not None:
                    if rv<0.3: risk='🔴 고위험'; rf2=fill('FFD7D7')
                    elif rv<0.7: risk='🟠 주의'; rf2=fill('FFE6C8')
                    else: risk='🟢 안전'; rf2=fill('D7F0D7')
                else: risk='⬜ 데이터없음'; rf2=rf
                c=ws2.cell(ri,sc+4,risk); c.font=Font(name='Arial',bold=True,size=8); c.border=BORDER; c.alignment=CTR
                if rf2: c.fill=rf2
        adc(pwb2['케이블 사용내역'],cf2); adc(pwb2['하우징 사용내역'],hf2)
        pwb2.save(OUT_FILE)
        print(" 완료! (제안량·생산비중·트렌드·위험도 자동 포함)")
    except Exception as e:
        print(f" 건너뜀 ({e})")
else:
    print("\nℹ 수요 기반 분석 제외 — 생산_판매_비교.xlsx를 같은 폴더에 추가하면 자동 포함됩니다")
print()
print(f"✅ 저장 완료: {os.path.basename(OUT_FILE)}")
print(f"   케이블: {len(cable_keys)}타입 / 하우징: {len(housing_keys)}타입")
print()
print("다음 단계:")
print("  1. 2026_연간발주계획.xlsx 열기")
print("  2. 노란색 셀(2026 목표 발주량)에 목표량 입력")
print("  3. 필요 발주량이 자동 계산됩니다")
input("\n엔터를 누르면 종료합니다...")
