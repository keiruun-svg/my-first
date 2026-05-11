#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STEP2 core — 웹앱에서 호출 가능한 형태
run(row_path, usage_path, ojc_ref_path, cmp_path, settings) -> (bytes, logs)
"""
import openpyxl, re, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

YEARS = ['23', '24', '25']
MM_KINDS = {'om1', 'om1-pigtail', 'om3'}
YR_PALETTE = ['2F5597', '2E75B6', '155480', '375623', '7030A0', '833C00', '1F3864', '595959']

def _fill(h): return PatternFill('solid', start_color=h)
def _font(bold=False, size=9, color='000000'): return Font(name='Arial', bold=bold, size=size, color=color)
def _bdr():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = _bdr()
CTR   = Alignment(horizontal='center', vertical='center')
CTR_W = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left',   vertical='center')
RIGHT = Alignment(horizontal='right',  vertical='center')
NUM_FMT = '#,##0'; PCT_FMT = '0.0%;[Red]-0.0%'
PAI_ORDER = {'2.0mm': 0, '3.0mm': 1, '0.9mm': 2}

def _np(p): return '0.9mm' if str(p).strip() == '0.9' else str(p).strip()
def _imc(k): return bool(re.match(r'^(b3|a1)-\d+c$', str(k).lower() if k else ''))

def _mc2(kind, core):
    k = str(kind).strip().lower() if kind else ''
    try: c = int(core)
    except: c = 1
    sd = 'SP' if c == 1 else ('DP' if c == 2 else f'{c}C')
    bm = {'a1': 'A1', 'b3': 'B3', 'om1': 'OM1', 'om3': 'OM3',
          'a1-청': 'A1_청', 'a1-녹': 'A1_녹', 'a1-적': 'A1_적', 'a1-자': 'A1_자'}
    if k in bm: return f'{bm[k]}-{sd}'
    if k == 'drop': return f'DROP-{sd}'
    if k in ('pigtail', 'om1-pigtail'): return 'PIGTAIL'
    if k == 'a2': return 'Optical cable'
    return k.upper()

def _gp(t, kind, pai):
    if not t: return None
    t = str(t).strip(); p = _np(pai); k = str(kind).strip().lower() if kind else ''
    if t == 'LC/PC':
        return (p, 'LC/PC 베이지MM') if k in MM_KINDS and p == '2.0mm' else (p, 'LC/PC 청색')
    return {'LC/APC': (p, 'LC/APC 녹색'), 'SC/PC': (p, 'SC/PC 청색'),
            'SC/APC': (p, 'SC/APC 녹색'), 'FC/PC': (p, 'FC/PC 흑색'),
            'FC/APC': (p, 'FC/APC 녹색')}.get(t)

def _gr(t, pai):
    if not t: return None
    t = str(t).strip(); p = _np(pai)
    return {'LC/PC': (p, 'LC/PC 적색'), 'LC/APC': (p, 'LC/APC 적색'),
            'SC/PC': (p, 'SC/PC 적색'), 'SC/APC': (p, 'SC/APC 적색'),
            'FC/PC': (p, 'FC/PC 적색'), 'FC/APC': (p, 'FC/APC 적색')}.get(t)


def run(row_path: str, usage_path: str = None, ojc_ref_path: str = None,
        cmp_path: str = None, settings: dict = None, ferrule_meta: dict = None,
        progress_cb=None,
        cable_meta_in: dict = None, housing_meta_in: dict = None,
        inventory: dict = None,
        sales_data: dict = None) -> tuple:
    """
    STEP2 실행: ROW + 사용내역 → 2026_연간발주계획 xlsx bytes 반환
    Returns: (xlsx_bytes: bytes, logs: list[str])
    """
    def _prog(pct, msg=''):
        if progress_cb:
            progress_cb(pct, msg)

    logs = []
    settings = settings or {}
    colors = settings.get('colors', {})
    C_MAIN = colors.get('main_header', '1F3864')
    C_23   = colors.get('year_23',    '2F5597')
    C_24   = colors.get('year_24',    '2E75B6')
    C_25   = colors.get('year_25',    '9DC3E6')
    lt_default = int(settings.get('lead_time_default', 60))

    def _lt_days(lt):
        if not lt: return lt_default
        m = re.search(r'(\d+)', str(lt))
        return int(m.group(1)) if m else lt_default

    # ── 메타데이터 로드 ──────────────────────────────────────
    _prog(5, "메타데이터 로드 중...")
    logs.append("메타데이터 로드 중...")
    cable_meta = {}; housing_meta = defaultdict(list)

    if usage_path:
        # 기존 방식: 사용내역 Excel에서 읽기 (하위 호환)
        usage_wb = openpyxl.load_workbook(usage_path, data_only=True)
        sn = usage_wb.sheetnames
        if '케이블 사용내역' in sn:
            for row in usage_wb['케이블 사용내역'].iter_rows(min_row=4, values_only=True):
                pai, ct, bunho, pname = row[1], row[2], row[3], row[4]
                if pai and ct and bunho:
                    k = (str(pai).strip(), str(ct).strip())
                    if k not in cable_meta:
                        cable_meta[k] = {
                            '품번': bunho, '품명': pname,
                            '구매처': row[15] if len(row) > 15 else None,
                            '리드타임': row[16] if len(row) > 16 else None,
                            '현재고': row[11] or 0 if len(row) > 11 else 0,
                        }
        if '하우징 사용내역' in sn:
            for row in usage_wb['하우징 사용내역'].iter_rows(min_row=4, values_only=True):
                pai, htype, bunho, pname = row[1], row[2], row[3], row[4]
                if pai and htype and bunho:
                    k = (str(pai).strip(), str(htype).strip())
                    housing_meta[k].append({
                        '품번': bunho, '품명': pname,
                        '구매처': row[16] if len(row) > 16 else None,
                        '리드타임': row[17] if len(row) > 17 else None,
                        '현재고': row[12] or 0 if len(row) > 12 else 0,
                        '기발주': row[13] or 0 if len(row) > 13 else 0,
                    })
        usage_wb.close()
    else:
        # 신규 방식: metadata.json + inventory.json에서 직접 구성
        _cinv  = (inventory or {}).get('cable', {})
        _hinv  = (inventory or {}).get('housing', {})
        for key_str, meta in (cable_meta_in or {}).items():
            parts = key_str.split('|', 1)
            if len(parts) != 2: continue
            pai, ct = parts
            iv = _cinv.get(key_str, {})
            cable_meta[(pai, ct)] = {
                '품번': meta.get('품번') or '',
                '품명': meta.get('품명') or '',
                '구매처': meta.get('구매처') or '',
                '리드타임': meta.get('리드타임'),
                '현재고': int(iv.get('현재고') or 0),
            }
        for key_str, meta_list in (housing_meta_in or {}).items():
            parts = key_str.split('|', 1)
            if len(parts) != 2: continue
            pai, htype = parts
            items    = meta_list if isinstance(meta_list, list) else [meta_list]
            inv_list = _hinv.get(key_str, [])
            for i, meta in enumerate(items):
                if not meta.get('품번'): continue
                iv = inv_list[i] if i < len(inv_list) else {}
                housing_meta[(pai, htype)].append({
                    '품번': meta.get('품번') or '',
                    '품명': meta.get('품명') or '',
                    '구매처': meta.get('구매처') or '',
                    '리드타임': meta.get('리드타임'),
                    '현재고': int(iv.get('현재고') or 0),
                    '기발주': int(iv.get('기발주') or 0),
                })

    # 피그테일 (0.9mm) — OJC 참고파일에서
    pigtail_cable_meta = {}; pigtail_cable_stats = {}
    COLOR_MAP = {'연청': '청록', '연등': '분홍'}

    def _extract_label(pname):
        pname = str(pname) if pname else ''
        if 'MM(OM4)' in pname: t = 'MM(OM4)'
        elif 'MM(OM3)' in pname: t = 'MM(OM3)'
        elif 'MM(OM1)' in pname: t = 'MM(OM1)'
        else: t = 'SM'
        m = re.search(r'\([^,]+,([^)]+)\)', pname)
        if m:
            kr = m.group(1).strip()
            return f'{t}-{COLOR_MAP.get(kr, kr)}'
        return t

    if ojc_ref_path:
        try:
            ref_wb = openpyxl.load_workbook(ojc_ref_path, data_only=True)
            ref_sn = ref_wb.sheetnames
            if '케이블 사용내역' in ref_sn:
                for row in ref_wb['케이블 사용내역'].iter_rows(min_row=4, values_only=True):
                    if row[1] and _np(row[1]) == '0.9mm' and row[2] == 'PIGTAIL' and row[3]:
                        label = _extract_label(row[4])
                        k = ('0.9mm', f'PIGTAIL-{label}')
                        pigtail_cable_meta[k] = {
                            '품번': row[3], '품명': row[4],
                            '구매처': row[12], '리드타임': row[13], '현재고': row[11] or 0,
                        }
                        pigtail_cable_stats[k] = {}
                        for yr, ai, pi in [('23',5,6),('24',7,8),('25',9,10)]:
                            ann  = round(float(row[ai])) if row[ai] else 0
                            peak = round(float(row[pi])) if row[pi] else 0
                            base = [round(ann/12)]*12; base[0] += round(ann) - sum(base)
                            pigtail_cable_stats[k][yr] = {'annual': ann, 'peak': peak, 'monthly': base}
            ref_wb.close()
        except Exception as e:
            logs.append(f"OJC 참고파일 로드 건너뜀: {e}")

    cable_meta[('2.0mm', 'PIGTAIL')] = {
        '품번': '(확인필요)', '품명': 'OPTICAL CABLE 2.0mm (2.0mm 자켓 피그테일용)',
        '구매처': '(확인필요)', '리드타임': f'{lt_default}일', '현재고': 0,
    }
    logs.append("메타데이터 로드 완료")

    # ── ROW 데이터 집계 ──────────────────────────────────────
    _prog(20, "ERP 파일 로드 중...")
    logs.append("ROW 데이터 집계 중...")
    import io as _io
    row_wb = openpyxl.load_workbook(row_path, read_only=True, data_only=True)
    row_sheets = set(row_wb.sheetnames)

    if '구매조회' in row_sheets or '구매현황' in row_sheets:
        row_wb.close()
        logs.append("ERP 원본 파일 감지 (구매조회/구매현황) — 자동 변환 시작")
        _prog(25, "ERP 원본 파일 변환 중...")
        import convert_core as _cc
        converted_wb = _cc.preprocess(row_path, logs)
        buf_tmp = _io.BytesIO()
        converted_wb.save(buf_tmp)
        converted_wb.close()
        buf_tmp.seek(0)
        row_wb = openpyxl.load_workbook(buf_tmp, read_only=True, data_only=True)
        row_sheets = set(row_wb.sheetnames)

    # 연도 자동 감지 (케이블 시트 기준)
    import re as _re
    active_years = sorted([
        _re.match(r'(\d{2})년[_ ]케이블', s).group(1)
        for s in row_sheets if _re.match(r'\d{2}년[_ ]케이블', s)
    ])
    if not active_years:
        if '케이블 사용내역' in row_sheets or '하우징 사용내역' in row_sheets:
            row_wb.close()
            raise ValueError(
                "잘못된 파일 형식입니다.\n"
                "'생산자재_사용내역.xlsx' 대신 '가공파일(연도별 시트).xlsx'를 STEP 3에 업로드하세요.\n"
                "또는 ERP 원본 파일(구매조회/구매현황)을 직접 업로드할 수 있습니다."
            )
        active_years = YEARS
        logs.append("연도 감지 실패 — 기본값 23/24/25 사용")
    else:
        logs.append(f"감지된 연도: {', '.join('20'+y for y in active_years)}년")

    cable_agg   = defaultdict(lambda: {yr: [0.0]*12 for yr in active_years})
    housing_agg = defaultdict(lambda: {yr: [0.0]*12 for yr in active_years})

    def _find_row_sheet(wb, yr, kind):
        for name in [f'{yr}년_{kind}', f'{yr}년 {kind}']:
            if name in set(wb.sheetnames): return name
        return None

    for _yi, yr in enumerate(active_years):
        _prog(35 + int(15 * _yi / max(len(active_years), 1)), f"20{yr}년 데이터 집계 중...")
        cs = _find_row_sheet(row_wb, yr, '케이블')
        hs = _find_row_sheet(row_wb, yr, '하우징')
        if cs:
            for row in list(row_wb[cs].iter_rows(values_only=True))[1:]:
                kind, pai, core, length = row[3], row[4], row[5], row[6]
                if not kind or not pai or not length: continue
                ct = _mc2(kind, core); p = _np(pai)
                if ct == 'PIGTAIL' and p == '0.9mm': continue
                for i, q in enumerate(row[9:21]):
                    if q: cable_agg[(p, ct)][yr][i] += float(q) * float(length)
        if hs:
            for row in list(row_wb[hs].iter_rows(values_only=True))[1:]:
                kind, pai, core, t1, t2 = row[3], row[4], row[5], row[7], row[8]
                if not pai: continue
                pai = _np(pai)
                try: cps = int(core) if core else 1
                except: cps = 1
                is_dp = (cps >= 2) and not _imc(kind)
                for mi, qty in enumerate(row[9:21]):
                    if not qty: continue
                    qty = float(qty)
                    ha = _gp(t1, kind, pai)
                    if ha: housing_agg[ha][yr][mi] += qty * cps
                    if t2:
                        hb = _gr(t2, pai) if is_dp else _gp(t2, kind, pai)
                        if hb: housing_agg[hb][yr][mi] += qty * cps

    def _fin(agg):
        r = {}
        for k, yd in agg.items():
            r[k] = {}
            for yr, m in yd.items():
                v = [round(x) for x in m]
                r[k][yr] = {'monthly': v, 'annual': sum(v), 'peak': max(v)}
        return r

    row_wb.close()
    cable_stats  = {**_fin(cable_agg), **pigtail_cable_stats}
    housing_stats = _fin(housing_agg)
    all_cable_meta = {**cable_meta, **pigtail_cable_meta}

    # ROW 누락 케이블 보완 (A1-4C, OM4-DP)
    REF_ONLY = {('2.0mm', 'A1-4C'), ('2.0mm', 'OM4-DP')}
    if ojc_ref_path:
        try:
            ojc_ref = openpyxl.load_workbook(ojc_ref_path, data_only=True)
            for row in ojc_ref['케이블 사용내역'].iter_rows(min_row=4, values_only=True):
                pai, kind, bunho, pname = row[1], row[2], row[3], row[4]
                if not pai or not kind: continue
                key = (str(pai).strip(), str(kind).strip())
                if key not in REF_ONLY or key in cable_stats: continue
                def _uni_ref(a):
                    b = [round(a/12)]*12 if a else [0]*12
                    b[0] += round(a or 0) - sum(b); return b
                cable_stats[key] = {}
                for yr, ai, pi in [('23',5,6),('24',7,8),('25',9,10)]:
                    ann  = round(float(row[ai])) if row[ai] else 0
                    peak = round(float(row[pi])) if row[pi] else 0
                    cable_stats[key][yr] = {'monthly': _uni_ref(ann), 'annual': ann, 'peak': peak}
                if key not in all_cable_meta and bunho:
                    all_cable_meta[key] = {
                        '품번': bunho, '품명': pname,
                        '구매처': row[12], '리드타임': row[13], '현재고': row[11] or 0,
                    }
            ojc_ref.close()
        except Exception as e:
            logs.append(f"OJC 보완 건너뜀: {e}")

    _prog(52, "통계 계산 완료 — Excel 빌드 시작...")
    logs.append(f"집계 완료 — 케이블 {len(cable_stats)}타입 / 하우징 {len(housing_stats)}타입")

    # ── Excel 빌드 ───────────────────────────────────────────
    COL_WIDTHS = [5,8,20,16,38,12,10,14,14,14,14,14,14,14,14,9,9,13,12,12,15,14,20]

    def _write_cell(ws, row, col, val, num=False, is_input=False, bold=False, row_fill=None):
        c = ws.cell(row, col, val)
        if is_input:
            c.font = _font(bold=True, size=9, color='0000FF'); c.fill = _fill('FFFFC0')
            c.border = Border(
                left=Side(style='medium', color='C55A11'), right=Side(style='medium', color='C55A11'),
                top=Side(style='medium', color='C55A11'), bottom=Side(style='medium', color='C55A11'),
            )
        else:
            c.font = _font(bold=bold, size=9)
            if row_fill: c.fill = row_fill
            c.border = BORDER
        c.number_format = NUM_FMT if num else 'General'
        c.alignment = RIGHT if num else (CTR if col in [1,2,6,7] else LEFT)

    def _write_sheet(ws, title, stats_dict, meta_dict, unit, type_label):
        keys = sorted(
            [k for k in stats_dict if any(stats_dict[k].get(yr, {}).get('annual', 0) > 0 for yr in active_years)],
            key=lambda k: (PAI_ORDER.get(k[0], 9), k[1])
        )
        ws.row_dimensions[1].height = 26; ws.merge_cells('A1:W1')
        c = ws['A1']; c.value = title
        c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR

        ws.row_dimensions[2].height = 18
        for rng, lbl, bg, fg in [
            ('A2:G2', '기본 정보', '374151', 'FFFFFF'),
            ('H2:I2', '2023년', C_23, 'FFFFFF'),
            ('J2:K2', '2024년', C_24, 'FFFFFF'),
            ('L2:M2', '2025년', C_25, '1F3864'),
            ('N2:Q2', '📊 트렌드 분석', '375623', 'FFFFFF'),
            ('R2:R2', '⚠ 안전재고', 'C00000', 'FFFFFF'),
            ('S2:T2', '재고 현황', '7030A0', 'FFFFFF'),
            ('U2:V2', '✏ 2026 발주 계획', 'C55A11', 'FFFFFF'),
            ('W2:W2', '비고', '595959', 'FFFFFF'),
        ]:
            s, e = rng.split(':')
            if s != e: ws.merge_cells(rng)
            c = ws[s]; c.value = lbl
            c.font = Font(name='Arial', bold=True, size=9, color=fg)
            c.fill = _fill(bg); c.alignment = CTR; c.border = BORDER

        ws.row_dimensions[3].height = 42
        hdrs = ['NO','파이',type_label,'품번','품명','구매처','리드타임\n(일)',
                f'연간({unit})',f'피크({unit})',f'연간({unit})',f'피크({unit})',
                f'연간({unit})',f'피크({unit})','3개년\n평균연간','3개년\n피크평균',
                '23→24\n증감률','24→25\n증감률',f'안전재고\n({unit})',
                f'현재고\n({unit})','기발주\n(참고)',f'2026목표\n({unit})',f'필요발주\n({unit})','비고']
        for ci, (h, w) in enumerate(zip(hdrs, COL_WIDTHS), 1):
            c = ws.cell(3, ci, h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill('BDD7EE'); c.alignment = CTR_W; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w

        ri = 4; no = 1
        for key in keys:
            pai, ctype = key; d = stats_dict[key]
            raw = meta_dict.get(key)
            comps = raw if isinstance(raw, list) else ([raw] if raw else [{}])
            first_u = ri
            for ci2, m in enumerate(comps):
                lt = _lt_days(m.get('리드타임', lt_default) if m else lt_default)
                rf = _fill('F5F5F5') if ri % 2 == 0 else None
                _write_cell(ws, ri, 1, no if ci2==0 else '', row_fill=rf)
                _write_cell(ws, ri, 2, pai if ci2==0 else '', row_fill=rf)
                _write_cell(ws, ri, 3, ctype if ci2==0 else '', row_fill=rf)
                _write_cell(ws, ri, 4, m.get('품번','') if m else '', row_fill=rf)
                _write_cell(ws, ri, 5, m.get('품명','') if m else '', row_fill=rf)
                _write_cell(ws, ri, 6, m.get('구매처','') if m else '', row_fill=rf)
                _write_cell(ws, ri, 7, lt, row_fill=rf)
                if ci2 == 0:
                    for col, yr, typ in [(8,'23','annual'),(9,'23','peak'),(10,'24','annual'),
                                         (11,'24','peak'),(12,'25','annual'),(13,'25','peak')]:
                        v = d[yr][typ]; _write_cell(ws, ri, col, v if v else None, num=True, row_fill=rf)
                    for col, fml, fmt_ in [
                        (14, f'=ROUND(AVERAGE(H{ri},J{ri},L{ri}),0)', NUM_FMT),
                        (15, f'=ROUND(AVERAGE(I{ri},K{ri},M{ri}),0)', NUM_FMT),
                        (16, f'=IFERROR((J{ri}-H{ri})/H{ri},"")', PCT_FMT),
                        (17, f'=IFERROR((L{ri}-J{ri})/J{ri},"")', PCT_FMT),
                    ]:
                        c2 = ws.cell(ri, col, fml); c2.font = _font(size=9); c2.border = BORDER
                        c2.number_format = fmt_; c2.alignment = RIGHT if col < 16 else CTR
                        if rf: c2.fill = rf
                else:
                    for col in range(8, 18):
                        cl = get_column_letter(col)
                        c2 = ws.cell(ri, col, f'={cl}{first_u}'); c2.font = _font(size=9); c2.border = BORDER
                        c2.number_format = NUM_FMT if col < 16 else PCT_FMT
                        c2.alignment = RIGHT if col < 16 else CTR
                        if rf: c2.fill = rf

                c2 = ws.cell(ri, 18, f'=ROUND(O{ri}*G{ri}/30,0)')
                c2.font = _font(bold=True, size=9); c2.border = BORDER
                c2.number_format = NUM_FMT; c2.alignment = RIGHT; c2.fill = _fill('FFF2CC')

                _write_cell(ws, ri, 19, m.get('현재고') if m else None, num=True, row_fill=rf)
                _write_cell(ws, ri, 20, m.get('기발주') if m else None, num=True, row_fill=rf)

                if ci2 == 0:
                    _write_cell(ws, ri, 21, None, num=True, is_input=True)
                else:
                    c2 = ws.cell(ri, 21, f'=U{first_u}')
                    c2.font = _font(bold=True, size=9, color='0000FF')
                    c2.fill = _fill('EBF3FB'); c2.border = BORDER
                    c2.number_format = NUM_FMT; c2.alignment = RIGHT

                c2 = ws.cell(ri, 22, f'=IFERROR(U{ri}-IFERROR(S{ri},0)-IFERROR(T{ri},0),"")')
                c2.font = _font(bold=True, size=9, color='C00000'); c2.border = BORDER
                c2.number_format = NUM_FMT; c2.alignment = RIGHT
                if rf: c2.fill = rf

                ws.cell(ri, 23, '').border = BORDER
                if rf: ws.cell(ri, 23).fill = rf
                ws.row_dimensions[ri].height = 17; ri += 1
            no += 1

        last = ri - 1; tr = ri + 1; ws.row_dimensions[tr].height = 18
        c = ws.cell(tr, 3, '합  계')
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR; c.border = BORDER
        for ci in range(8, 16):
            cl = get_column_letter(ci)
            c = ws.cell(tr, ci, f'=SUM({cl}4:{cl}{last})')
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(C_MAIN); c.number_format = NUM_FMT; c.alignment = RIGHT; c.border = BORDER
        ws.freeze_panes = 'A4'; ws.sheet_view.showGridLines = False
        return keys

    def _write_bunho_sheet(ws):
        """품번별 발주 집계 — 동일 품번이 여러 하우징 타입에 걸쳐 있을 때 자동 합산"""
        # 품번별 집계
        bunho_data = {}
        for (pai, htype), comps in housing_meta.items():
            if (pai, htype) not in housing_stats: continue
            h_st = housing_stats[(pai, htype)]
            items = comps if isinstance(comps, list) else [comps]
            for item in items:
                bn = (item.get('품번') or '').strip()
                if not bn: continue
                if bn not in bunho_data:
                    bunho_data[bn] = {
                        '품명': item.get('품명', ''),
                        '구매처': item.get('구매처', ''),
                        '리드타임': _lt_days(item.get('리드타임')),
                        '현재고': 0, '기발주': 0,
                        'stats': {yr: [0.0]*12 for yr in active_years},
                    }
                bunho_data[bn]['현재고'] += item.get('현재고', 0) or 0
                bunho_data[bn]['기발주'] += item.get('기발주', 0) or 0
                for yr in active_years:
                    for i, v in enumerate(h_st.get(yr, {}).get('monthly', [0]*12)):
                        bunho_data[bn]['stats'][yr][i] += v

        n_yr = len(active_years)
        # 컬럼 레이아웃: NO 품번 품명 구매처 리드타임 | 연도별연간×n | n개년평균 안전재고 현재고 기발주 2026목표 필요발주 비고
        C_YRS  = 6                      # 첫 연도 연간 컬럼
        C_AVG  = C_YRS + n_yr           # n개년 평균
        C_SS   = C_AVG + 1              # 안전재고
        C_CUR  = C_SS  + 1              # 현재고
        C_ORD  = C_CUR + 1              # 기발주
        C_TGT  = C_ORD + 1              # 2026목표 (입력)
        C_REQ  = C_TGT + 1              # 필요발주
        C_NOTE = C_REQ + 1              # 비고
        total_cols = C_NOTE

        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 26
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        c = ws['A1']; c.value = '2026 연간 발주 계획 — 품번별 집계 (하우징 공용 부품 합산)'
        c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR

        ws.row_dimensions[2].height = 18
        bands = [('A2', 'E2', '기본 정보', '374151', 'FFFFFF')]
        for i, yr in enumerate(active_years):
            cl = get_column_letter(C_YRS + i)
            bands.append((f'{cl}2', f'{cl}2', f'20{yr}년 연간', YR_PALETTE[i % len(YR_PALETTE)], 'FFFFFF'))
        bands += [
            (get_column_letter(C_AVG)+'2',  get_column_letter(C_AVG)+'2',  f'{n_yr}개년 평균', '375623', 'FFFFFF'),
            (get_column_letter(C_SS)+'2',   get_column_letter(C_SS)+'2',   '안전재고',         'C00000', 'FFFFFF'),
            (get_column_letter(C_CUR)+'2',  get_column_letter(C_ORD)+'2',  '재고 현황',        '7030A0', 'FFFFFF'),
            (get_column_letter(C_TGT)+'2',  get_column_letter(C_REQ)+'2',  '✏ 2026 발주 계획', 'C55A11', 'FFFFFF'),
            (get_column_letter(C_NOTE)+'2', get_column_letter(C_NOTE)+'2', '비고',             '595959', 'FFFFFF'),
        ]
        for s, e, lbl, bg, fg in bands:
            if s != e: ws.merge_cells(f'{s}:{e}')
            c = ws[s]; c.value = lbl
            c.font = Font(name='Arial', bold=True, size=9, color=fg)
            c.fill = _fill(bg); c.alignment = CTR; c.border = BORDER

        ws.row_dimensions[3].height = 42
        hdrs = ['NO', '품번', '품명', '구매처', f'리드타임\n(일)']
        for yr in active_years: hdrs.append(f'20{yr}년\n연간(EA)')
        hdrs += [f'{n_yr}개년\n평균(EA)', f'안전재고\n(EA)', f'현재고\n(EA)', f'기발주\n(참고)',
                 f'2026목표\n(EA)', f'필요발주\n(EA)', '비고']
        col_widths = [5, 16, 38, 14, 10] + [12]*n_yr + [12, 12, 10, 10, 12, 12, 20]
        for ci, (h, w) in enumerate(zip(hdrs, col_widths), 1):
            c = ws.cell(3, ci, h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill('BDD7EE'); c.alignment = CTR_W; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w

        ri = 4; no = 1
        for bn, d in sorted(bunho_data.items()):
            rf = _fill('F5F5F5') if ri % 2 == 0 else None
            lt = d['리드타임']
            ann_list = [round(sum(d['stats'].get(yr, {yr: [0]*12} if False else [0]*12)
                                  if isinstance(d['stats'].get(yr), list) else d['stats'].get(yr, [0]*12)))
                        for yr in active_years]
            # 연도별 연간합계 계산
            ann_vals = [round(sum(d['stats'].get(yr, [0]*12))) for yr in active_years]

            for ci, v in enumerate([no, bn, d['품명'], d['구매처'], lt], 1):
                c = ws.cell(ri, ci, v)
                c.font = _font(size=9, bold=(ci==2))
                c.border = BORDER
                c.alignment = CTR if ci in [1,5] else LEFT
                c.number_format = NUM_FMT if ci == 5 else 'General'
                if rf: c.fill = rf

            for i, (yr, av) in enumerate(zip(active_years, ann_vals)):
                col = C_YRS + i
                c = ws.cell(ri, col, av if av else None)
                c.font = _font(size=9); c.border = BORDER
                c.number_format = NUM_FMT; c.alignment = RIGHT
                if rf: c.fill = rf

            yr_cols = ','.join(f'{get_column_letter(C_YRS+i)}{ri}' for i in range(n_yr))
            c = ws.cell(ri, C_AVG, f'=ROUND(AVERAGE({yr_cols}),0)')
            c.font = _font(size=9); c.border = BORDER
            c.number_format = NUM_FMT; c.alignment = RIGHT
            if rf: c.fill = rf

            c = ws.cell(ri, C_SS, f'=ROUND({get_column_letter(C_AVG)}{ri}*E{ri}/30,0)')
            c.font = _font(bold=True, size=9); c.border = BORDER
            c.number_format = NUM_FMT; c.alignment = RIGHT; c.fill = _fill('FFF2CC')

            for col, val in [(C_CUR, d['현재고']), (C_ORD, d['기발주'])]:
                c = ws.cell(ri, col, val if val else None)
                c.font = _font(size=9); c.border = BORDER
                c.number_format = NUM_FMT; c.alignment = RIGHT
                if rf: c.fill = rf

            c = ws.cell(ri, C_TGT, None)
            c.font = _font(bold=True, size=9, color='0000FF'); c.fill = _fill('FFFFC0')
            c.border = Border(
                left=Side(style='medium', color='C55A11'), right=Side(style='medium', color='C55A11'),
                top=Side(style='medium', color='C55A11'), bottom=Side(style='medium', color='C55A11'),
            )
            c.number_format = NUM_FMT; c.alignment = RIGHT

            tgt = get_column_letter(C_TGT); cur = get_column_letter(C_CUR)
            ord_ = get_column_letter(C_ORD); ss = get_column_letter(C_SS)
            c = ws.cell(ri, C_REQ,
                f'=IFERROR(MAX({tgt}{ri}-{cur}{ri}-{ord_}{ri}+{ss}{ri},0),"")')
            c.font = _font(bold=True, size=9, color='C00000'); c.border = BORDER
            c.number_format = NUM_FMT; c.alignment = RIGHT
            if rf: c.fill = rf

            ws.cell(ri, C_NOTE, '').border = BORDER
            if rf: ws.cell(ri, C_NOTE).fill = rf
            ws.row_dimensions[ri].height = 17; ri += 1; no += 1

        # ── 페롤 섹션 ────────────────────────────────────────
        if ferrule_meta:
            # 커넥터 타입별 수량 집계 (파이 무관 합산)
            conn_stats = {}
            for (pai, htype), h_st in housing_stats.items():
                m = re.match(r'^(LC/PC|LC/APC|SC/PC|SC/APC|FC/PC|FC/APC)', str(htype))
                if not m: continue
                ct = m.group(1)
                if ct not in conn_stats:
                    conn_stats[ct] = {yr: [0.0]*12 for yr in active_years}
                for yr in active_years:
                    for i, v in enumerate(h_st.get(yr, {}).get('monthly', [0]*12)):
                        conn_stats[ct][yr][i] += v

            # 구분선 행
            ri += 1
            ws.merge_cells(f'A{ri}:{get_column_letter(total_cols)}{ri}')
            c = ws.cell(ri, 1, '▼ 페롤 (커넥터 타입별 공용 — 파이 무관 합산)')
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill('375623'); c.alignment = CTR; c.border = BORDER
            ws.row_dimensions[ri].height = 16; ri += 1

            for ct, fv in ferrule_meta.items():
                if ct not in conn_stats: continue
                rf = _fill('F0FFF4') if ri % 2 == 0 else None
                lt = _lt_days(fv.get('리드타임'))
                ann_vals = [round(sum(conn_stats[ct].get(yr, [0]*12))) for yr in active_years]

                for ci, v in enumerate([no, fv.get('품번',''), fv.get('품명',''), fv.get('구매처',''), lt], 1):
                    c = ws.cell(ri, ci, v)
                    c.font = _font(size=9, bold=(ci==2))
                    c.border = BORDER
                    c.alignment = CTR if ci in [1,5] else LEFT
                    c.number_format = NUM_FMT if ci == 5 else 'General'
                    if rf: c.fill = rf

                for i, av in enumerate(ann_vals):
                    col = C_YRS + i
                    c = ws.cell(ri, col, av if av else None)
                    c.font = _font(size=9); c.border = BORDER
                    c.number_format = NUM_FMT; c.alignment = RIGHT
                    if rf: c.fill = rf

                yr_cols = ','.join(f'{get_column_letter(C_YRS+i)}{ri}' for i in range(n_yr))
                c = ws.cell(ri, C_AVG, f'=ROUND(AVERAGE({yr_cols}),0)')
                c.font = _font(size=9); c.border = BORDER
                c.number_format = NUM_FMT; c.alignment = RIGHT
                if rf: c.fill = rf

                c = ws.cell(ri, C_SS, f'=ROUND({get_column_letter(C_AVG)}{ri}*E{ri}/30,0)')
                c.font = _font(bold=True, size=9); c.border = BORDER
                c.number_format = NUM_FMT; c.alignment = RIGHT; c.fill = _fill('FFF2CC')

                for col in [C_CUR, C_ORD]:
                    ws.cell(ri, col, None).border = BORDER

                c = ws.cell(ri, C_TGT, None)
                c.font = _font(bold=True, size=9, color='0000FF'); c.fill = _fill('FFFFC0')
                c.border = Border(
                    left=Side(style='medium', color='C55A11'), right=Side(style='medium', color='C55A11'),
                    top=Side(style='medium', color='C55A11'), bottom=Side(style='medium', color='C55A11'),
                )
                c.number_format = NUM_FMT; c.alignment = RIGHT

                tgt = get_column_letter(C_TGT); ss = get_column_letter(C_SS)
                c = ws.cell(ri, C_REQ, f'=IFERROR(MAX({tgt}{ri}+{ss}{ri},0),"")')
                c.font = _font(bold=True, size=9, color='C00000'); c.border = BORDER
                c.number_format = NUM_FMT; c.alignment = RIGHT
                if rf: c.fill = rf

                # 커넥터 타입 비고 표시
                c = ws.cell(ri, C_NOTE, ct)
                c.font = _font(size=8, color='375623'); c.border = BORDER; c.alignment = LEFT
                if rf: c.fill = rf

                ws.row_dimensions[ri].height = 17; ri += 1; no += 1

        ws.freeze_panes = 'A4'

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _prog(60, "케이블 사용내역 시트 작성 중...")
    ws_c = wb.create_sheet('케이블 사용내역')
    cable_keys = _write_sheet(ws_c, '2026 연간 발주 계획 — 케이블 사용내역  (단위 : m)',
                              cable_stats, all_cable_meta, 'm', '케이블 종류')
    _prog(72, "하우징 사용내역 시트 작성 중...")
    ws_h = wb.create_sheet('하우징 사용내역')
    housing_keys = _write_sheet(ws_h, '2026 연간 발주 계획 — 하우징 사용내역  (단위 : EA)',
                                housing_stats, housing_meta, 'EA', '하우징 타입')
    _prog(82, "품번별 발주 집계 시트 작성 중...")
    ws_b = wb.create_sheet('📦 품번별 발주 집계')
    _write_bunho_sheet(ws_b)

    # ── 월별 발주계획 시트 ───────────────────────────────────
    _prog(90, "월별 발주계획 시트 작성 중...")
    ws_m = wb.create_sheet('2026 월별 발주계획'); ws_m.sheet_view.showGridLines = False
    ws_m.row_dimensions[1].height = 26; ws_m.merge_cells('A1:T1')
    c = ws_m['A1']; c.value = '2026 월별 발주 계획 (과거 계절 패턴 기반 자동 분배)'
    c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    c.fill = _fill(C_MAIN); c.alignment = CTR

    ws_m.row_dimensions[2].height = 18
    for rng, lbl, bg, fg in [
        ('A2:F2', '기본정보', '374151', 'FFFFFF'),
        ('G2:G2', '연간목표', 'C55A11', 'FFFFFF'),
        ('H2:S2', '월별 발주 (연간목표 × 과거 패턴)', '2E75B6', 'FFFFFF'),
        ('T2:T2', '합계검증', '375623', 'FFFFFF'),
    ]:
        s, e = rng.split(':')
        if s != e: ws_m.merge_cells(rng)
        c = ws_m[s]; c.value = lbl
        c.font = Font(name='Arial', bold=True, size=9, color=fg)
        c.fill = _fill(bg); c.alignment = CTR; c.border = BORDER

    ws_m.row_dimensions[3].height = 40
    for ci, h in enumerate(['NO','분류','파이','종류','품번','단위','연간목표'] + [f'{i}월' for i in range(1,13)] + ['합계검증'], 1):
        c = ws_m.cell(3, ci, h)
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = _fill('BDD7EE'); c.alignment = CTR_W; c.border = BORDER
    for ci, w in enumerate([5,10,8,20,16,6,14]+[9]*12+[10], 1):
        ws_m.column_dimensions[get_column_letter(ci)].width = w

    ri = 4
    all_items = ([('케이블', k, cable_stats[k]) for k in cable_keys] +
                 [('하우징', k, housing_stats[k]) for k in housing_keys])
    for idx, (cat, key, d) in enumerate(all_items):
        pai, ctype = key; rf = _fill('F5F5F5') if ri % 2 == 0 else None
        combined = [0.0]*12; total = 0.0
        for yr in active_years:
            for i, v in enumerate(d.get(yr, {}).get('monthly', [0]*12)): combined[i] += v; total += v
        ratios = [round(v/total, 6) if total > 0 else round(1/12, 6) for v in combined]
        unit = 'm' if cat == '케이블' else 'EA'
        for ci2, v in enumerate([idx+1, cat, pai, ctype, '', unit], 1):
            c = ws_m.cell(ri, ci2, v)
            c.font = _font(size=9); c.alignment = CTR if ci2 <= 3 else LEFT; c.border = BORDER
            if rf: c.fill = rf
        c = ws_m.cell(ri, 7, None)
        c.font = _font(bold=True, size=9, color='0000FF'); c.fill = _fill('FFFFC0')
        c.border = Border(
            left=Side(style='medium', color='C55A11'), right=Side(style='medium', color='C55A11'),
            top=Side(style='medium', color='C55A11'), bottom=Side(style='medium', color='C55A11'),
        )
        c.number_format = NUM_FMT; c.alignment = RIGHT
        for mi, ratio in enumerate(ratios):
            c = ws_m.cell(ri, 8+mi, f'=IFERROR(ROUND($G{ri}*{ratio},0),"")')
            c.font = _font(size=9); c.border = BORDER; c.number_format = NUM_FMT; c.alignment = RIGHT
            if rf: c.fill = rf
        c = ws_m.cell(ri, 20, f'=IFERROR(SUM(H{ri}:S{ri}),"")')
        c.font = _font(size=9, color='375623', bold=True); c.border = BORDER
        c.number_format = NUM_FMT; c.alignment = RIGHT
        if rf: c.fill = rf
        ws_m.row_dimensions[ri].height = 17; ri += 1
    ws_m.freeze_panes = 'A4'

    # ── 이상항목 시트 ────────────────────────────────────────
    ws_a = wb.create_sheet('⚠ 이상항목 검토'); ws_a.sheet_view.showGridLines = False
    ws_a.row_dimensions[1].height = 26; ws_a.merge_cells('A1:D1')
    c = ws_a['A1']; c.value = '데이터 이상 항목 검토 (자동 분석)'
    c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    c.fill = _fill('C00000'); c.alignment = CTR
    ws_a.row_dimensions[3].height = 30
    for ci, h in enumerate(['구분', '항목', '품번', '내용 및 조치 권고'], 1):
        c = ws_a.cell(3, ci, h)
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = _fill('BDD7EE'); c.alignment = CTR_W; c.border = BORDER
    ws_a.column_dimensions['A'].width = 10; ws_a.column_dimensions['B'].width = 22
    ws_a.column_dimensions['C'].width = 16; ws_a.column_dimensions['D'].width = 65
    anomalies = [
        ('주의', '2.0mm 자켓 피그테일', '(확인필요)', '참고파일에 품번 미등재. 사용 케이블 품번 확인 필요.', 'orange'),
        ('정보', 'OM4 피그테일 케이블', 'P14-RM-417K', '3개년 사용 없음. 재고 보유. 단종 검토 필요.', 'red'),
        ('정보', 'OM3 피그테일 케이블', 'P14-RM-417H', '23년 648m → 24~25년 0m. 미사용 추세.', 'orange'),
        ('정보', '피그테일 전체', '(전 색상)', '23년 대비 25년 약 81% 급감. 2026 목표량 보수적 설정 권고.', 'blue'),
    ]
    cm = {'red': 'FFD7D7', 'orange': 'FFE6C8', 'blue': 'D7E8FF'}
    for ri2, (_t, item, bunho, desc, color) in enumerate(anomalies, start=4):
        rf2 = PatternFill('solid', start_color=cm.get(color, 'FFFFFF'))
        for ci, v in enumerate([_t, item, bunho, desc], 1):
            c = ws_a.cell(ri2, ci, v)
            c.font = Font(name='Arial', size=9, bold=(ci==1)); c.fill = rf2; c.border = BORDER
            c.alignment = CTR if ci == 1 else (Alignment(horizontal='left', vertical='center', wrap_text=True) if ci == 4 else LEFT)
        ws_a.row_dimensions[ri2].height = 36
    ws_a.freeze_panes = 'A4'

    # ── 수요 기반 분석 (sales_data dict 또는 생산_판매_비교.xlsx) ─────────
    _prog(96, "파일 저장 중...")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    if sales_data or cmp_path:
        logs.append("수요 기반 분석 추가 중...")
        try:
            if sales_data:
                # 웹앱 판매 분석 탭에서 저장된 dict 직접 사용
                prod_data2 = {}
                for code, data in sales_data.items():
                    prod_data2[str(code).strip()] = {
                        yr: {
                            'sales': data.get(yr, {}).get('sales', 0),
                            'ratio': data.get(yr, {}).get('ratio', 0),
                        }
                        for yr in ('23', '24', '25')
                    }
                cmp_wb2 = None
            else:
                cmp_wb2 = openpyxl.load_workbook(cmp_path, read_only=True, data_only=True)
                prod_data2 = {}
                for row in list(cmp_wb2['연간_요약'].iter_rows(values_only=True))[3:]:
                    if not row[1]: continue
                    code = str(row[1]).strip()
                    prod_data2[code] = {
                        '23': {'sales': row[4] or 0, 'ratio': row[7] or 0},
                        '24': {'sales': row[8] or 0, 'ratio': row[11] or 0},
                        '25': {'sales': row[12] or 0, 'ratio': row[15] or 0},
                    }
            row_wb2 = openpyxl.load_workbook(row_path, read_only=True, data_only=True)
            cable_use2 = {}; housing_use2 = {}
            for yr2, cs2, hs2 in [('25','25년_케이블','25년 하우징'),('24','24년_케이블','24년 하우징'),('23','23년_케이블','23년 하우징')]:
                for row in list(row_wb2[cs2].iter_rows(values_only=True))[1:]:
                    code = str(row[0]) if row[0] else ''
                    if not code or code in cable_use2: continue
                    kind, pai, core, length = row[3], row[4], row[5], row[6]
                    if not kind or not pai or not length: continue
                    ct = _mc2(kind, core); p = _np(pai)
                    if ct == 'PIGTAIL' and p == '0.9mm': continue
                    cable_use2[code] = (p, ct, float(length))
                for row in list(row_wb2[hs2].iter_rows(values_only=True))[1:]:
                    code = str(row[0]) if row[0] else ''
                    if not code or code in housing_use2: continue
                    kind, pai, core, t1, t2 = row[3], row[4], row[5], row[7], row[8]
                    if not pai: continue
                    p = _np(pai)
                    try: cps = int(core) if core else 1
                    except: cps = 1
                    is_dp2 = (cps >= 2) and not _imc(kind); h = {}
                    ha = _gp(t1, kind, p)
                    if ha: h[ha] = h.get(ha, 0) + cps
                    if t2:
                        hb = _gr(t2, p) if is_dp2 else _gp(t2, kind, p)
                        if hb: h[hb] = h.get(hb, 0) + cps
                    if h: housing_use2[code] = h

            def _calc_trend2(d):
                s23 = d['23']['sales']; s25 = d['25']['sales']
                if s23 > 0 and s25 > 0: return (s25/s23)**0.5 - 1
                if s25 > 0 and d['24']['sales'] > 0: return s25/d['24']['sales'] - 1
                return 0.0

            cp2 = defaultdict(lambda: [0,0,0,0,0]); hp2 = defaultdict(lambda: [0,0,0,0,0])
            for code, (pai, ct, length) in cable_use2.items():
                if code not in prod_data2: continue
                d = prod_data2[code]; trend = min(max(_calc_trend2(d), -0.5), 1.0)
                ratio = min(d['25']['ratio'], 1.5); s25 = d['25']['sales']
                if s25 == 0: continue
                key = (pai, ct); p2 = cp2[key]
                p2[0] += s25*(1+trend)*ratio*length; p2[1] += s25*ratio; p2[2] += s25; p2[3] += s25*trend; p2[4] += s25
            for code, h_map in housing_use2.items():
                if code not in prod_data2: continue
                d = prod_data2[code]; trend = min(max(_calc_trend2(d), -0.5), 1.0)
                ratio = min(d['25']['ratio'], 1.5); s25 = d['25']['sales']
                if s25 == 0: continue
                for (pai, htype), qpu in h_map.items():
                    key = (pai, htype); p2 = hp2[key]
                    p2[0] += s25*(1+trend)*ratio*qpu; p2[1] += s25*ratio; p2[2] += s25; p2[3] += s25*trend; p2[4] += s25

            def _fp2(p): return round(p[0]), (p[1]/p[2] if p[2] > 0 else None), (p[3]/p[4] if p[4] > 0 else None)
            cf2 = {k: _fp2(v) for k, v in cp2.items()}; hf2 = {k: _fp2(v) for k, v in hp2.items()}

            pwb2 = openpyxl.load_workbook(buf)
            buf.seek(0)

            def _adc(ws2, dd2):
                sc = 24; N2 = 14
                ws2.merge_cells(f'{get_column_letter(sc)}2:{get_column_letter(sc+4)}2')
                c = ws2[f'{get_column_letter(sc)}2']; c.value = '📊 수요 기반 분석 (판매량 + 생산비중)'
                c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                c.fill = _fill('843C0C'); c.alignment = CTR; c.border = BORDER
                for ci, h in enumerate(['2026 제안량\n(수요기반)','vs 3개년평균\n비교','평균\n생산비중','판매\n트렌드','수입의존\n위험도']):
                    c = ws2.cell(3, sc+ci, h)
                    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                    c.fill = _fill('843C0C'); c.alignment = CTR_W; c.border = BORDER
                    ws2.column_dimensions[get_column_letter(sc+ci)].width = 13
                ck = None; cf = None
                for ri3 in range(4, ws2.max_row + 1):
                    pv = ws2.cell(ri3, 2).value; tv = ws2.cell(ri3, 3).value
                    if str(tv or '').strip() == '합  계': continue
                    if pv and tv: ck = (str(pv).strip(), str(tv).strip()); cf = ri3
                    if not ck: continue
                    proj, ratio, trend = dd2.get(ck, (None, None, None))
                    rf3 = _fill('F5F5F5') if ri3 % 2 == 0 else None
                    c = ws2.cell(ri3, sc, round(proj) if proj else None)
                    c.font = Font(name='Arial', bold=True, size=9, color='1F3864')
                    c.border = BORDER; c.number_format = '#,##0'; c.alignment = RIGHT; c.fill = _fill('FFF2CC')
                    expr = f'=IFERROR({get_column_letter(sc)}{ri3}/{get_column_letter(N2)}{ri3}-1,"")' if ri3 == cf else f'={get_column_letter(sc+1)}{cf}'
                    c = ws2.cell(ri3, sc+1, expr); c.font = _font(size=9); c.border = BORDER
                    c.number_format = '0%;[Red]-0%'; c.alignment = CTR; c.fill = _fill('EBF3FB')
                    rv = round(ratio, 3) if ratio else None
                    c = ws2.cell(ri3, sc+2, rv); c.font = _font(size=9); c.border = BORDER; c.number_format = '0%'; c.alignment = CTR
                    c.fill = _fill('D7F0D7') if rv and rv >= 0.8 else (_fill('FFF2CC') if rv and rv >= 0.5 else (_fill('FFD7D7') if rv else (rf3 if rf3 else PatternFill())))
                    tv2 = round(trend, 4) if trend is not None else None
                    c = ws2.cell(ri3, sc+3, tv2)
                    c.font = Font(name='Arial', size=9, color='1A5C1A' if (tv2 or 0) > 0 else 'C00000')
                    c.border = BORDER; c.number_format = '+0.0%;[Red]-0.0%'; c.alignment = CTR; c.fill = _fill('D7E8FF')
                    if rv is not None:
                        if rv < 0.3: risk = '🔴 고위험'; rf4 = _fill('FFD7D7')
                        elif rv < 0.7: risk = '🟠 주의'; rf4 = _fill('FFE6C8')
                        else: risk = '🟢 안전'; rf4 = _fill('D7F0D7')
                    else: risk = '⬜ 데이터없음'; rf4 = rf3
                    c = ws2.cell(ri3, sc+4, risk)
                    c.font = Font(name='Arial', bold=True, size=8); c.border = BORDER; c.alignment = CTR
                    if rf4: c.fill = rf4

            _adc(pwb2['케이블 사용내역'], cf2); _adc(pwb2['하우징 사용내역'], hf2)
            if cmp_wb2: cmp_wb2.close()
            row_wb2.close()
            buf2 = io.BytesIO(); pwb2.save(buf2); pwb2.close(); buf2.seek(0)
            logs.append("수요 기반 분석 완료 (제안량·생산비중·트렌드·위험도 포함)")
            return buf2.read(), logs
        except Exception as e:
            logs.append(f"수요 기반 분석 건너뜀: {e}")

    _prog(100, "완료")
    logs.append(f"Excel 생성 완료 — 케이블 {len(cable_keys)}타입 / 하우징 {len(housing_keys)}타입")
    return buf.read(), logs
