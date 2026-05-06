#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STEP 1: 3개년_생산자재_사용내역.xlsx 자동 생성
사용법: python STEP1_사용내역_생성.py
필요 파일: 가공파일_통합_v양식.xlsx (같은 폴더)
선택 파일: 기존 3개년_생산자재_사용내역.xlsx (품번 자동 매칭)
           비밀_OJC 3개년 생산자제 사용 내역_*.xlsx (A1-4C 등 보완용)
"""
import openpyxl, re, os, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DIR = os.path.dirname(os.path.abspath(__file__))
ROW_FILE = os.path.join(DIR, '가공파일_통합_v양식.xlsx')
REF_FILE = os.path.join(DIR, '3개년_생산자재_사용내역.xlsx')
OUT_FILE = os.path.join(DIR, '3개년_생산자재_사용내역.xlsx')
OJC_REF_FILE = os.path.join(DIR, '비밀_OJC 3개년 생산자제 사용 내역_SCM팀_박정원_260326.xlsx')

print("="*55)
print("  3개년 생산자재 사용내역 자동 생성")
print("="*55)
if not os.path.exists(ROW_FILE):
    print(f"\n파일 없음: {ROW_FILE}")
    print("  가공파일_통합_v양식.xlsx를 같은 폴더에 놓아주세요.")
    input("\n엔터를 누르면 종료합니다..."); sys.exit(1)

print(f"ROW 데이터 발견")
print(f"참고파일: {'있음' if os.path.exists(REF_FILE) else '없음 (빈칸 생성)'}")
print()

YEARS=['23','24','25']; MM={'om1','om1-pigtail','om3'}
PAI_ORDER={'2.0mm':0,'3.0mm':1,'0.9mm':2}
REF_ONLY_KEYS={('2.0mm','A1-4C'),('2.0mm','OM4-DP')}

def fill(h): return PatternFill('solid',start_color=h)
def font(bold=False,size=9,color='000000'): return Font(name='Arial',bold=bold,size=size,color=color)
def bdr(): s=Side(style='thin'); return Border(left=s,right=s,top=s,bottom=s)
BORDER=bdr(); CTR=Alignment(horizontal='center',vertical='center')
CTR_W=Alignment(horizontal='center',vertical='center',wrap_text=True)
LEFT=Alignment(horizontal='left',vertical='center'); RIGHT=Alignment(horizontal='right',vertical='center')

def np(p): return '0.9mm' if str(p).strip()=='0.9' else str(p).strip()
def imc(k): return bool(re.match(r'^(b3|a1)-\d+c$',str(k).lower() if k else ''))
def mc2(kind,core):
    k=str(kind).strip().lower() if kind else ''
    try: c=int(core)
    except: c=1
    sd='SP' if c==1 else('DP' if c==2 else f'{c}C')
    bm={'a1':'A1','b3':'B3','om1':'OM1','om3':'OM3','a1-청':'A1_청','a1-녹':'A1_녹','a1-적':'A1_적','a1-자':'A1_자'}
    if k in bm: return f'{bm[k]}-{sd}'
    if k=='drop': return 'DROP'
    if k in('pigtail','om1-pigtail'): return 'PIGTAIL'
    if k=='a2': return 'Optical cable'
    return k.upper()
def gp(t,kind,pai):
    if not t: return None
    t=str(t).strip(); p=np(pai); k=str(kind).strip().lower() if kind else ''
    if t=='LC/PC': return (p,'LC/PC 베이지MM') if k in MM and p=='2.0mm' else (p,'LC/PC 청색')
    return {'LC/APC':(p,'LC/APC 녹색'),'SC/PC':(p,'SC/PC 청색'),'SC/APC':(p,'SC/APC 녹색'),'FC/PC':(p,'FC/PC 흑색'),'FC/APC':(p,'FC/APC 녹색')}.get(t)
def gr(t,pai):
    if not t: return None
    t=str(t).strip(); p=np(pai)
    return {'LC/PC':(p,'LC/PC 적색'),'LC/APC':(p,'LC/APC 적색'),'SC/PC':(p,'SC/PC 적색'),'SC/APC':(p,'SC/APC 적색'),'FC/PC':(p,'FC/PC 적색'),'FC/APC':(p,'FC/APC 적색')}.get(t)

print("ROW 데이터 집계 중...", end='', flush=True)
row_wb=openpyxl.load_workbook(ROW_FILE,read_only=True,data_only=True)
cable_agg=defaultdict(lambda:{yr:[0.0]*12 for yr in YEARS})
housing_agg=defaultdict(lambda:{yr:[0.0]*12 for yr in YEARS})
for yr in YEARS:
    for row in list(row_wb[f'{yr}년_케이블'].iter_rows(values_only=True))[1:]:
        kind=row[3];pai=row[4];core=row[5];length=row[6]
        if not kind or not pai or not length: continue
        ct=mc2(kind,core); p_c=np(pai)
        if ct=='PIGTAIL' and p_c=='0.9mm': continue
        for i,q in enumerate(row[9:21]):
            if q: cable_agg[(p_c,ct)][yr][i]+=float(q)*float(length)
    for row in list(row_wb[f'{yr}년 하우징'].iter_rows(values_only=True))[1:]:
        kind=row[3];pai=row[4];core=row[5];t1=row[7];t2=row[8]
        if not pai: continue
        p_h=np(pai)
        try: cps=int(core) if core else 1
        except: cps=1
        is_dp=(cps>=2) and not imc(kind)
        for mi,qty in enumerate(row[9:21]):
            if not qty: continue
            qty=float(qty)
            ha=gp(t1,kind,p_h)
            if ha: housing_agg[ha][yr][mi]+=qty*cps
            if t2:
                hb=gr(t2,p_h) if is_dp else gp(t2,kind,p_h)
                if hb: housing_agg[hb][yr][mi]+=qty*cps

def fin(agg):
    r={}
    for k,yd in agg.items():
        r[k]={}
        for yr,m in yd.items():
            v=[round(x) for x in m]; r[k][yr]={'monthly':v,'annual':sum(v),'peak':max(v)}
    return r
cs=fin(cable_agg); hs=fin(housing_agg)
print(f" 완료 (케이블 {len(cs)}타입 / 하우징 {len(hs)}타입)")

# ROW 누락 케이블 보완 (A1-4C, OM4-DP)
if os.path.exists(OJC_REF_FILE):
    try:
        ojc_ref=openpyxl.load_workbook(OJC_REF_FILE,data_only=True)
        for row in ojc_ref['케이블 사용내역'].iter_rows(min_row=4,values_only=True):
            pai,kind=row[1],row[2]
            if not pai or not kind: continue
            key=(str(pai).strip(),str(kind).strip())
            if key not in REF_ONLY_KEYS or key in cs: continue
            def uni2(a):
                b=[round(a/12)]*12 if a else [0]*12
                b[0]+=round(a or 0)-sum(b); return b
            cs[key]={}
            for yr,ai,pi in[('23',5,6),('24',7,8),('25',9,10)]:
                ann=round(float(row[ai])) if row[ai] else 0
                peak=round(float(row[pi])) if row[pi] else 0
                cs[key][yr]={'monthly':uni2(ann),'annual':ann,'peak':peak}
        print("  A1-4C/OM4-DP 보완 적용")
    except Exception:
        pass

cable_meta={}; housing_meta=defaultdict(list)
if os.path.exists(REF_FILE):
    print("참고파일 매칭 중...", end='', flush=True)
    try:
        ref_wb=openpyxl.load_workbook(REF_FILE,data_only=True)
        sn=ref_wb.sheetnames
        cs_name='케이블 사용내역' if '케이블 사용내역' in sn else None
        hs_name='하우징 사용내역' if '하우징 사용내역' in sn else None
        if cs_name:
            for row in ref_wb[cs_name].iter_rows(min_row=4,values_only=True):
                pai,ct,bunho,pname=row[1],row[2],row[3],row[4]
                if pai and ct and bunho:
                    k=(str(pai).strip(),str(ct).strip())
                    if k not in cable_meta:
                        cable_meta[k]={'품번':bunho,'품명':pname,'구매처':row[15] if len(row)>15 else None,'리드타임':row[16] if len(row)>16 else None,'현재고':row[11] or 0 if len(row)>11 else 0}
        if hs_name:
            for row in ref_wb[hs_name].iter_rows(min_row=4,values_only=True):
                pai,htype,bunho,pname=row[1],row[2],row[3],row[4]
                if pai and htype and bunho:
                    k=(str(pai).strip(),str(htype).strip())
                    housing_meta[k].append({'품번':bunho,'품명':pname,'구매처':row[16] if len(row)>16 else None,'리드타임':row[17] if len(row)>17 else None,'현재고':row[12] or 0 if len(row)>12 else 0,'기발주':row[13] or 0 if len(row)>13 else 0})
        print(" 완료")
    except Exception as e:
        print(f" 건너뜀 ({e})")

if os.path.exists(OJC_REF_FILE):
    try:
        ojc_ref2=openpyxl.load_workbook(OJC_REF_FILE,data_only=True)
        for row in ojc_ref2['케이블 사용내역'].iter_rows(min_row=4,values_only=True):
            pai,kind,bunho,pname=row[1],row[2],row[3],row[4]
            if not pai or not kind or not bunho: continue
            key=(str(pai).strip(),str(kind).strip())
            if key in REF_ONLY_KEYS and key not in cable_meta:
                cable_meta[key]={'품번':bunho,'품명':pname,'구매처':row[12],'리드타임':row[13],'현재고':row[11] or 0}
    except Exception:
        pass

wb=openpyxl.Workbook(); wb.remove(wb.active)

def write_cable_sheet(ws):
    ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=24; ws.merge_cells('A1:R1')
    c=ws['A1']; c.value='2023~2025년 케이블 생산자재 사용내역 (ROW 데이터 자동 생성)'
    c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR
    ws.row_dimensions[2].height=18
    for rng,lbl,bg in[('A2:E2','기본정보','374151'),('F2:G2','23년','2F5597'),('H2:I2','24년','2E75B6'),('J2:K2','25년','155480'),('L2:L2','현재고','7030A0'),('M2:M2','기발주','7030A0'),('N2:O2','3개년 분석','375623'),('P2:P2','구매처','595959'),('Q2:Q2','리드타임','595959'),('R2:R2','비고','595959')]:
        s,e=rng.split(':')
        if s!=e: ws.merge_cells(rng)
        c=ws[s]; c.value=lbl; c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill(bg); c.alignment=CTR; c.border=BORDER
    ws.row_dimensions[3].height=40
    hdrs=['NO','파이','케이블종류','품번','품명','연간(m)','피크(m)','연간(m)','피크(m)','연간(m)','피크(m)','현재고(m)','기발주(m)','3개년평균(m)','3개년피크평균(m)','구매처','리드타임(일)','비고']
    hfill=['374151']*5+['2F5597']*2+['2E75B6']*2+['155480']*2+['7030A0']*2+['375623']*2+['595959']*3
    wds=[5,8,18,16,36,14,14,14,14,14,14,12,10,14,14,14,10,20]
    for ci,(h,bf,w) in enumerate(zip(hdrs,hfill,wds),1):
        c=ws.cell(3,ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
        c.fill=fill(bf); c.alignment=CTR_W; c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ri=4; no=0; unmatched=[]
    for key in sorted(cs,key=lambda k:(PAI_ORDER.get(k[0],9),k[1])):
        pai,ct=key; d=cs[key]
        if not any(d[yr]['annual']>0 for yr in YEARS): continue
        no+=1; meta=cable_meta.get(key,{}); matched=bool(meta)
        if not matched: unmatched.append(key)
        is_even=(ri%2==0); rf=fill('F5F5F5') if is_even else None
        ws.row_dimensions[ri].height=17
        for ci,val in enumerate([no,pai,ct,meta.get('품번',''),meta.get('품명','')],1):
            c=ws.cell(ri,ci,val); c.font=font(size=9); c.border=BORDER
            c.alignment=RIGHT if ci>=6 else(LEFT if ci>=2 else CTR)
            if rf: c.fill=rf
        for col,yr,typ in[(6,'23','annual'),(7,'23','peak'),(8,'24','annual'),(9,'24','peak'),(10,'25','annual'),(11,'25','peak')]:
            v=d[yr][typ]; c=ws.cell(ri,col,v if v else None); c.font=font(size=9); c.border=BORDER; c.number_format='#,##0'; c.alignment=RIGHT
            if rf: c.fill=rf
        ws.cell(ri,12,meta.get('현재고') or None).font=font(size=9); ws.cell(ri,12).border=BORDER; ws.cell(ri,12).number_format='#,##0'; ws.cell(ri,12).alignment=RIGHT
        if rf: ws.cell(ri,12).fill=rf
        ws.cell(ri,13,None).border=BORDER
        if rf: ws.cell(ri,13).fill=rf
        c=ws.cell(ri,14,f'=ROUND(AVERAGE(F{ri},H{ri},J{ri}),0)'); c.font=font(size=9); c.border=BORDER; c.number_format='#,##0'; c.alignment=RIGHT
        if rf: c.fill=rf
        c=ws.cell(ri,15,f'=ROUND(AVERAGE(G{ri},I{ri},K{ri}),0)'); c.font=font(size=9); c.border=BORDER; c.number_format='#,##0'; c.alignment=RIGHT
        if rf: c.fill=rf
        ws.cell(ri,16,meta.get('구매처','')).font=font(size=9); ws.cell(ri,16).border=BORDER; ws.cell(ri,16).alignment=LEFT
        if rf: ws.cell(ri,16).fill=rf
        ws.cell(ri,17,meta.get('리드타임','')).font=font(size=9); ws.cell(ri,17).border=BORDER; ws.cell(ri,17).alignment=CTR
        if rf: ws.cell(ri,17).fill=rf
        from_ref=key in REF_ONLY_KEYS
        nv='⚠ 참고파일 매칭 필요' if not matched else ('※ 참고파일 기준' if from_ref else '')
        c=ws.cell(ri,18,nv); c.font=Font(name='Arial',size=8,color='C00000' if not matched else '595959')
        c.border=BORDER; c.alignment=LEFT
        if not matched: c.fill=fill('FFD7D7')
        elif rf: c.fill=rf
        ri+=1
    last=ri-1; tr=ri+1; ws.row_dimensions[tr].height=18
    c=ws.cell(tr,2,'합  계'); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR; c.border=BORDER
    for ci in [6,7,8,9,10,11,14,15]:
        cl=get_column_letter(ci)
        c=ws.cell(tr,ci,f'=SUM({cl}4:{cl}{last})')
        c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill('1F3864'); c.number_format='#,##0'; c.alignment=RIGHT; c.border=BORDER
    ws.freeze_panes='D4'
    return no, unmatched

def write_housing_sheet(ws):
    ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=24; ws.merge_cells('A1:S1')
    c=ws['A1']; c.value='2023~2025년 하우징 생산자재 사용내역 (ROW 데이터 자동 생성)'
    c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF'); c.fill=fill('1F3864'); c.alignment=CTR
    ws.row_dimensions[2].height=18
    for rng,lbl,bg in[('A2:F2','기본정보','374151'),('G2:H2','23년','2F5597'),('I2:J2','24년','2E75B6'),('K2:L2','25년','155480'),('M2:M2','현재고','7030A0'),('N2:N2','기발주','7030A0'),('O2:P2','3개년 분석','375623'),('Q2:Q2','구매처','595959'),('R2:R2','리드타임','595959'),('S2:S2','비고','595959')]:
        s,e=rng.split(':')
        if s!=e: ws.merge_cells(rng)
        c=ws[s]; c.value=lbl; c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF'); c.fill=fill(bg); c.alignment=CTR; c.border=BORDER
    ws.row_dimensions[3].height=40
    hdrs2=['NO','파이','하우징타입','품번','품명','구매용도','연간','피크','연간','피크','연간','피크','현재고','기발주','3개년평균','3개년피크평균','구매처','리드타임(일)','비고']
    hfill2=['374151']*6+['2F5597']*2+['2E75B6']*2+['155480']*2+['7030A0']*2+['375623']*2+['595959']*3
    wds2=[5,8,16,16,36,14,12,12,12,12,12,12,10,10,12,12,14,10,20]
    for ci,(h,bf,w) in enumerate(zip(hdrs2,hfill2,wds2),1):
        c=ws.cell(3,ci,h); c.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
        c.fill=fill(bf); c.alignment=CTR_W; c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ri2=4; no2=0
    for key in sorted(hs,key=lambda k:(PAI_ORDER.get(k[0],9),k[1])):
        pai,htype=key; d=hs[key]
        if not any(d.get(yr,{}).get('annual',0)>0 for yr in YEARS): continue
        comps=housing_meta.get(key,[{}]); first_u=ri2
        for ci2,comp in enumerate(comps):
            no2+=1; matched=bool(comp and comp.get('품번'))
            is_even=(ri2%2==0); rf=fill('F5F5F5') if is_even else None
            ws.row_dimensions[ri2].height=17
            for col,val in [(1,no2 if ci2==0 else ''),(2,pai if ci2==0 else ''),(3,htype if ci2==0 else '')]:
                c=ws.cell(ri2,col,val); c.font=font(size=9); c.border=BORDER; c.alignment=CTR
                if rf: c.fill=rf
            for ci3,val in enumerate([comp.get('품번','') if comp else '',comp.get('품명','') if comp else ''],4):
                ws.cell(ri2,ci3,val).font=font(size=9); ws.cell(ri2,ci3).border=BORDER; ws.cell(ri2,ci3).alignment=LEFT
                if rf: ws.cell(ri2,ci3).fill=rf
            pn=comp.get('품명','') if comp else ''; usage=''
            if 'SPRING' in str(pn).upper(): usage='스프링'
            elif 'HOUSING' in str(pn).upper(): usage='하우징 본체'
            elif 'DUST CAP' in str(pn).upper(): usage='더스트캡'
            elif 'STOPPER' in str(pn).upper(): usage='스토퍼'
            elif 'FRAME' in str(pn).upper(): usage='플러그 프레임'
            elif 'BOOT' in str(pn).upper(): usage='부트'
            elif 'KIT' in str(pn).upper(): usage='하우징 키트'
            ws.cell(ri2,6,usage).font=font(size=8,color='595959'); ws.cell(ri2,6).border=BORDER; ws.cell(ri2,6).alignment=CTR
            if rf: ws.cell(ri2,6).fill=rf
            if ci2==0:
                for col,yr,typ in[(7,'23','annual'),(8,'23','peak'),(9,'24','annual'),(10,'24','peak'),(11,'25','annual'),(12,'25','peak')]:
                    v=d.get(yr,{}).get(typ,0) or None
                    c3=ws.cell(ri2,col,v); c3.font=font(size=9); c3.border=BORDER; c3.number_format='#,##0'; c3.alignment=RIGHT
                    if rf: c3.fill=rf
                for col,fml in[(15,f'=ROUND(AVERAGE(G{ri2},I{ri2},K{ri2}),0)'),(16,f'=ROUND(AVERAGE(H{ri2},J{ri2},L{ri2}),0)')]:
                    c3=ws.cell(ri2,col,fml); c3.font=font(size=9); c3.border=BORDER; c3.number_format='#,##0'; c3.alignment=RIGHT
                    if rf: c3.fill=rf
            else:
                for col in range(7,17):
                    cl=get_column_letter(col)
                    c3=ws.cell(ri2,col,f'={cl}{first_u}'); c3.font=font(size=9); c3.border=BORDER; c3.number_format='#,##0'; c3.alignment=RIGHT
                    if rf: c3.fill=rf
            for col,fld in[(13,'현재고'),(14,'기발주')]:
                v=comp.get(fld) if comp and comp.get(fld) else None
                ws.cell(ri2,col,v).font=font(size=9); ws.cell(ri2,col).border=BORDER; ws.cell(ri2,col).number_format='#,##0'; ws.cell(ri2,col).alignment=RIGHT
                if rf: ws.cell(ri2,col).fill=rf
            ws.cell(ri2,17,comp.get('구매처','') if comp else '').font=font(size=9); ws.cell(ri2,17).border=BORDER; ws.cell(ri2,17).alignment=LEFT
            if rf: ws.cell(ri2,17).fill=rf
            ws.cell(ri2,18,comp.get('리드타임','') if comp else '').font=font(size=9); ws.cell(ri2,18).border=BORDER; ws.cell(ri2,18).alignment=CTR
            if rf: ws.cell(ri2,18).fill=rf
            nv='' if matched else '⚠ 참고파일 매칭 필요'
            if pai=='0.9mm' and not nv: nv='※ 참고파일 기준'
            ws.cell(ri2,19,nv).font=Font(name='Arial',size=8,color='C00000' if '⚠' in nv else '595959')
            ws.cell(ri2,19).border=BORDER; ws.cell(ri2,19).alignment=LEFT
            if '⚠' in nv: ws.cell(ri2,19).fill=fill('FFD7D7')
            elif rf: ws.cell(ri2,19).fill=rf
            ri2+=1
    ws.freeze_panes='D4'
    return ri2-4

print("Excel 파일 생성 중...", end='', flush=True)
ws_cable=wb.create_sheet('케이블 사용내역')
n_cable, unmatched = write_cable_sheet(ws_cable)
ws_housing=wb.create_sheet('하우징 사용내역')
n_housing = write_housing_sheet(ws_housing)
wb.save(OUT_FILE)
print(f" 완료!")
print()
print(f"저장 완료: {os.path.basename(OUT_FILE)}")
print(f"   케이블: {n_cable}행 / 하우징: {n_housing}행")
if unmatched:
    print(f"\n⚠ {len(unmatched)}개 항목 품번 직접 입력 필요:")
    for k in unmatched: print(f"   - {k}")
print()
print("다음 단계: 현재고·기발주 입력 후 STEP2_발주계획_생성.py 실행")
input("\n엔터를 누르면 종료합니다...")
