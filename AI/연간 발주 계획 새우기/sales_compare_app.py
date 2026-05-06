"""
AJW 생산_판매_비교.xlsx 자동 생성기
버전: v1 (2026-04-27)
실행: python sales_compare_app.py
빌드: pyinstaller --onefile --windowed sales_compare_app.py
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 계산 로직
# ============================================================
MONTHS = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']
YEARS  = ['23', '24', '25']
PROD_SHEETS = {'23':'23년_케이블', '24':'24년_케이블', '25':'25년_케이블'}
SALE_SHEETS = {'23':'23년_완제품', '24':'24년_완제품', '25':'25년_완제품'}

C_HEADER = '1F3864'
C_YEAR23 = '2F5597'
C_YEAR24 = '2E75B6'
C_YEAR25 = '155480'
C_SALE   = 'D6E4F0'
C_PROD   = 'E2EFDA'
C_IMP    = 'FCE4D6'
C_RATIO  = 'FFF2CC'
C_EVEN   = 'F5F5F5'
C_DROP   = 'EDEDED'
C_RED    = 'C00000'
C_SUB    = '243F60'

def fill(h): return PatternFill('solid', start_color=h, end_color=h)
def font(bold=False, size=9, color='000000'):
    return Font(name='Arial', bold=bold, size=size, color=color)
def bdr():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)
BORDER = bdr()
CTR   = Alignment(horizontal='center', vertical='center')
CTR_W = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left',   vertical='center')
RIGHT = Alignment(horizontal='right',  vertical='center')

def is_drop(kind):
    return str(kind).strip().lower() == 'drop' if kind else False

def safe_int(v):
    try: return int(v) if v else 0
    except: return 0

def sc(cell, value, bold=False, bg=None, fg='000000', align=CTR, size=9):
    cell.value = value
    cell.font = font(bold=bold, size=size, color=fg)
    if bg: cell.fill = fill(bg)
    cell.alignment = align
    cell.border = BORDER

def load_production(path, log):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for yr, sname in PROD_SHEETS.items():
        result[yr] = {}
        if sname not in wb.sheetnames:
            log(f'  [경고] 생산 시트 없음: {sname}')
            continue
        ws = wb[sname]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]: continue
            key = (str(row[0]).strip(), str(row[1]).strip() if row[1] else '', str(row[2]).strip() if row[2] else '')
            kind = str(row[3]).strip() if row[3] else ''
            monthly = [safe_int(row[9+i]) for i in range(12)]
            if key not in result[yr]:
                result[yr][key] = {'monthly': [0]*12, 'kind': kind}
            for i in range(12):
                result[yr][key]['monthly'][i] += monthly[i]
            result[yr][key]['kind'] = kind
    wb.close()
    return result

def load_sales(path, log):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for yr, sname in SALE_SHEETS.items():
        result[yr] = {}
        if sname not in wb.sheetnames:
            log(f'  [경고] 판매 시트 없음: {sname}')
            continue
        ws = wb[sname]
        for row in list(ws.iter_rows(values_only=True))[2:]:
            if not row[1]: continue
            key = (str(row[1]).strip(), str(row[2]).strip() if row[2] else '', str(row[3]).strip() if row[3] else '')
            result[yr][key] = [safe_int(row[4+i]) for i in range(12)]
    wb.close()
    return result

def collect_items(prod, sale):
    seen = {}
    for yr in YEARS:
        for key in sale.get(yr, {}):
            if key not in seen: seen[key] = ''
        for key, v in prod.get(yr, {}).items():
            if key not in seen: seen[key] = v.get('kind','')
    kinds = {}
    for yr in YEARS:
        for key, v in prod.get(yr, {}).items():
            if key not in kinds: kinds[key] = v.get('kind','')
    ordered = list(seen.keys())
    drop_items  = [(k, kinds.get(k,'')) for k in ordered if is_drop(kinds.get(k,''))]
    other_items = [(k, kinds.get(k,'')) for k in ordered if not is_drop(kinds.get(k,''))]
    return drop_items + other_items

def calc_import(sale_m, prod_m, drop):
    if drop: return [None]*12
    return [max(0, s - p) if (s or p) else None for s, p in zip(sale_m, prod_m)]

def build_annual_sheet(wb_out, items, prod, sale):
    ws = wb_out.create_sheet('연간_요약')
    ws.merge_cells('A1:Q1')
    c = ws['A1']
    c.value = '생산 vs 판매 — 연간 요약  ※ DROP-CABLE: 전량 생산 처리'
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = fill(C_HEADER)
    c.alignment = CTR

    yr_info = [('2023년', 5, C_YEAR23), ('2024년', 9, C_YEAR24), ('2025년', 13, C_YEAR25)]
    for label, sc_col, color in yr_info:
        ws.merge_cells(start_row=2, start_column=sc_col, end_row=2, end_column=sc_col+3)
        c = ws.cell(row=2, column=sc_col)
        c.value = label
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = fill(color); c.alignment = CTR; c.border = BORDER

    headers = ['NO','품목코드','품목명','규격명','판매','생산','수입추정','생산비중',
               '판매','생산','수입추정','생산비중','판매','생산','수입추정','생산비중','비고']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = fill(C_SUB); c.alignment = CTR_W; c.border = BORDER
    ws.row_dimensions[3].height = 28

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 8
    for ci in range(5, 18):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.column_dimensions['Q'].width = 16

    yr_map = {'23': 5, '24': 9, '25': 13}
    for ri, (key, kind) in enumerate(items, 1):
        row = ri + 3
        drop = is_drop(kind)
        bg_row = C_DROP if drop else (C_EVEN if ri % 2 == 0 else None)
        sc(ws.cell(row=row, column=1), ri, align=CTR, bg=bg_row)
        sc(ws.cell(row=row, column=2), key[0], align=LEFT, bg=bg_row)
        sc(ws.cell(row=row, column=3), key[1], align=LEFT, bg=bg_row)
        sc(ws.cell(row=row, column=4), key[2], align=CTR, bg=bg_row)
        for yr, base in yr_map.items():
            sale_m = sale.get(yr, {}).get(key, [0]*12)
            prod_d = prod.get(yr, {}).get(key, {'monthly':[0]*12})
            prod_m = prod_d['monthly']
            s_ann = sum(sale_m); p_ann = sum(prod_m)
            imp_m = calc_import(sale_m, prod_m, drop)
            i_ann = sum(v for v in imp_m if v is not None) if not drop else None
            ratio = f'{p_ann/s_ann:.1%}' if s_ann else (None if not p_ann else None)
            sc(ws.cell(row=row, column=base),   s_ann or None, align=RIGHT, bg=C_SALE)
            sc(ws.cell(row=row, column=base+1), p_ann or None, align=RIGHT, bg=C_PROD)
            sc(ws.cell(row=row, column=base+2), i_ann if not drop else None, align=RIGHT,
               bg=C_IMP, fg=C_RED if (i_ann and i_ann > 0) else '000000')
            sc(ws.cell(row=row, column=base+3), ratio, align=CTR, bg=C_RATIO)
        sc(ws.cell(row=row, column=17), '전량생산(DROP-CABLE)' if drop else '', align=LEFT, bg=bg_row)
    ws.freeze_panes = 'E4'

def build_monthly_sheet(wb_out, yr, items, prod, sale):
    yr_label = f'20{yr}년'
    ws = wb_out.create_sheet(f'{yr}년_월별비교')
    total_cols = 4 + 12*3 + 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws['A1']
    c.value = f'{yr_label} 월별 판매 vs 생산 vs 수입추정  ※ DROP-CABLE: 전량 생산'
    yr_color = {'23': C_YEAR23, '24': C_YEAR24, '25': C_YEAR25}[yr]
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = fill(yr_color); c.alignment = CTR

    for ci, h in enumerate(['NO','품목코드','품목명','규격명'], 1):
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
        c = ws.cell(row=2, column=ci)
        c.value = h; c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = fill(C_SUB); c.alignment = CTR_W; c.border = BORDER

    for mi, m in enumerate(MONTHS):
        base = 5 + mi*3
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base+2)
        c = ws.cell(row=2, column=base)
        c.value = m; c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = fill(C_SUB); c.alignment = CTR; c.border = BORDER
        for si, (lbl, bg) in enumerate([('판매',C_SALE),('생산',C_PROD),('수입추정',C_IMP)]):
            col = base + si
            c2 = ws.cell(row=3, column=col)
            c2.value = lbl; c2.font = Font(name='Arial', bold=True, size=8)
            c2.fill = fill(bg); c2.alignment = CTR; c2.border = BORDER

    for ci, lbl in enumerate(['연간판매','연간생산','연간수입'], 5+36):
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
        c = ws.cell(row=2, column=ci)
        c.value = lbl; c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = fill(C_HEADER); c.alignment = CTR_W; c.border = BORDER

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 7
    for ci in range(5, total_cols+1):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    ws.row_dimensions[3].height = 20

    for ri, (key, kind) in enumerate(items, 1):
        row = ri + 3; drop = is_drop(kind)
        bg_row = C_DROP if drop else (C_EVEN if ri % 2 == 0 else None)
        sale_m = sale.get(yr, {}).get(key, [0]*12)
        prod_d = prod.get(yr, {}).get(key, {'monthly':[0]*12})
        prod_m = prod_d['monthly']
        imp_m = calc_import(sale_m, prod_m, drop)
        sc(ws.cell(row=row, column=1), ri, align=CTR, bg=bg_row)
        sc(ws.cell(row=row, column=2), key[0], align=LEFT, bg=bg_row)
        sc(ws.cell(row=row, column=3), key[1], align=LEFT, bg=bg_row)
        sc(ws.cell(row=row, column=4), key[2], align=CTR, bg=bg_row)
        for mi in range(12):
            base = 5 + mi*3
            iv = imp_m[mi]
            sc(ws.cell(row=row, column=base),   sale_m[mi] or None, align=RIGHT, bg=C_SALE)
            sc(ws.cell(row=row, column=base+1), prod_m[mi] or None, align=RIGHT, bg=C_PROD)
            sc(ws.cell(row=row, column=base+2), iv if not drop else None, align=RIGHT,
               bg=C_IMP, fg=C_RED if (iv and iv > 0) else '000000')
        ann_base = 5 + 36
        s_ann = sum(sale_m); p_ann = sum(prod_m)
        i_ann = sum(v for v in imp_m if v is not None) if not drop else None
        sc(ws.cell(row=row, column=ann_base),   s_ann or None, align=RIGHT, bold=True, bg=C_SALE)
        sc(ws.cell(row=row, column=ann_base+1), p_ann or None, align=RIGHT, bold=True, bg=C_PROD)
        sc(ws.cell(row=row, column=ann_base+2), i_ann if not drop else None, align=RIGHT,
           bold=True, bg=C_IMP, fg=C_RED if (i_ann and i_ann > 0) else '000000')
    ws.freeze_panes = 'E4'

def build_compare(prod_path, sale_path, out_path, log):
    log('생산 데이터 로드 중...')
    prod = load_production(prod_path, log)
    log('판매 데이터 로드 중...')
    sale = load_sales(sale_path, log)
    items = collect_items(prod, sale)
    drop_cnt = sum(1 for _, k in items if is_drop(k))
    log(f'총 {len(items)}개 품목 확인 (DROP-CABLE {drop_cnt}개)')
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    log('연간_요약 시트 생성 중...')
    build_annual_sheet(wb_out, items, prod, sale)
    for yr in YEARS:
        log(f'20{yr}년_월별비교 시트 생성 중...')
        build_monthly_sheet(wb_out, yr, items, prod, sale)
    wb_out.save(out_path)
    # 수입추정 발생 품목 통계
    wb_chk = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws_ann = wb_chk['연간_요약']
    rows_ann = list(ws_ann.iter_rows(values_only=True))[3:]
    stats = {}
    for yr_i, col in enumerate([6, 10, 14]):
        cnt = sum(1 for r in rows_ann if r[col] and r[col] > 0)
        stats[YEARS[yr_i]] = cnt
    wb_chk.close()
    log(f'\n✅ 완료!')
    log(f'   수입추정 발생: 23년 {stats["23"]}개 / 24년 {stats["24"]}개 / 25년 {stats["25"]}개')
    log(f'   저장 위치: {out_path}')
    return True

# ============================================================
# GUI
# ============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('AJW 생산_판매_비교 생성기  v1')
        self.resizable(False, False)
        self.configure(bg='#F0F4F8')
        self._build_ui()
        self.center()

    def center(self):
        self.update_idletasks()
        w, h = 580, 480
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    def _build_ui(self):
        # ── 헤더 ──
        hdr = tk.Frame(self, bg='#1F3864', height=56)
        hdr.pack(fill='x')
        tk.Label(hdr, text='  AJW  생산 vs 판매 비교 생성기',
                 font=('Arial', 13, 'bold'), fg='white', bg='#1F3864').pack(side='left', pady=12)
        tk.Label(hdr, text='v1 · 2026', font=('Arial', 9), fg='#8BAFD4', bg='#1F3864').pack(side='right', padx=12)

        pad = dict(padx=20, pady=6)

        # ── 파일 선택 ──
        grp = tk.LabelFrame(self, text=' 입력 파일 ', font=('Arial', 9, 'bold'),
                            bg='#F0F4F8', fg='#1F3864', bd=1)
        grp.pack(fill='x', **pad)

        self.prod_var = tk.StringVar()
        self.sale_var = tk.StringVar()
        self.out_var  = tk.StringVar()

        for label, var, cmd in [
            ('생산 ROW 데이터\n(가공파일_통합_v양식.xlsx)', self.prod_var, self._pick_prod),
            ('판매량 정리 파일\n(판매량_정리.xlsx)',         self.sale_var, self._pick_sale),
        ]:
            row = tk.Frame(grp, bg='#F0F4F8')
            row.pack(fill='x', padx=8, pady=4)
            tk.Label(row, text=label, width=22, justify='left',
                     font=('Arial', 9), bg='#F0F4F8', fg='#333').pack(side='left')
            tk.Entry(row, textvariable=var, width=34, font=('Arial', 9),
                     state='readonly', relief='flat', bg='white').pack(side='left', padx=4)
            tk.Button(row, text='찾기', command=cmd, width=5,
                      font=('Arial', 9), bg='#2E75B6', fg='white',
                      relief='flat', cursor='hand2').pack(side='left')

        # ── 출력 파일 ──
        grp2 = tk.LabelFrame(self, text=' 출력 파일 ', font=('Arial', 9, 'bold'),
                             bg='#F0F4F8', fg='#1F3864', bd=1)
        grp2.pack(fill='x', **pad)
        row2 = tk.Frame(grp2, bg='#F0F4F8')
        row2.pack(fill='x', padx=8, pady=4)
        tk.Label(row2, text='저장 위치', width=22, font=('Arial', 9),
                 bg='#F0F4F8', fg='#333').pack(side='left')
        tk.Entry(row2, textvariable=self.out_var, width=34, font=('Arial', 9),
                 relief='flat', bg='white').pack(side='left', padx=4)
        tk.Button(row2, text='찾기', command=self._pick_out, width=5,
                  font=('Arial', 9), bg='#2E75B6', fg='white',
                  relief='flat', cursor='hand2').pack(side='left')

        # ── 실행 버튼 ──
        self.run_btn = tk.Button(self, text='▶  생성 시작', command=self._run,
                                 font=('Arial', 11, 'bold'), bg='#1F3864', fg='white',
                                 relief='flat', cursor='hand2', height=2)
        self.run_btn.pack(fill='x', padx=20, pady=8)

        # ── 로그 ──
        grp3 = tk.LabelFrame(self, text=' 진행 상황 ', font=('Arial', 9, 'bold'),
                             bg='#F0F4F8', fg='#1F3864', bd=1)
        grp3.pack(fill='both', expand=True, padx=20, pady=(0,16))
        self.log_text = tk.Text(grp3, height=8, font=('Consolas', 9),
                                bg='#1A1A2E', fg='#A8D8EA', relief='flat',
                                state='disabled', wrap='word')
        self.log_text.pack(fill='both', expand=True, padx=6, pady=6)

    # ── 파일 선택 ──
    def _pick_prod(self):
        p = filedialog.askopenfilename(
            title='가공파일_통합_v양식.xlsx 선택',
            filetypes=[('Excel 파일', '*.xlsx')])
        if p:
            self.prod_var.set(p)
            self._auto_out()

    def _pick_sale(self):
        p = filedialog.askopenfilename(
            title='판매량_정리.xlsx 선택',
            filetypes=[('Excel 파일', '*.xlsx')])
        if p: self.sale_var.set(p)

    def _pick_out(self):
        p = filedialog.asksaveasfilename(
            title='저장 위치 선택',
            defaultextension='.xlsx',
            initialfile='생산_판매_비교.xlsx',
            filetypes=[('Excel 파일', '*.xlsx')])
        if p: self.out_var.set(p)

    def _auto_out(self):
        prod = self.prod_var.get()
        if prod and not self.out_var.get():
            self.out_var.set(os.path.join(os.path.dirname(prod), '생산_판매_비교.xlsx'))

    # ── 로그 출력 ──
    def log(self, msg):
        self.log_text.configure(state='normal')
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.log_text.configure(state='disabled')
        self.update_idletasks()

    # ── 실행 ──
    def _run(self):
        prod = self.prod_var.get().strip()
        sale = self.sale_var.get().strip()
        out  = self.out_var.get().strip()

        if not prod or not os.path.exists(prod):
            messagebox.showerror('오류', '가공파일_통합_v양식.xlsx를 선택하세요.')
            return
        if not sale or not os.path.exists(sale):
            messagebox.showerror('오류', '판매량_정리.xlsx를 선택하세요.')
            return
        if not out:
            messagebox.showerror('오류', '저장 위치를 지정하세요.')
            return

        self.run_btn.configure(state='disabled', text='처리 중...')
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.configure(state='disabled')

        def worker():
            try:
                build_compare(prod, sale, out, self.log)
                self.after(0, lambda: self.run_btn.configure(state='normal', text='▶  생성 시작'))
                self.after(0, lambda: messagebox.showinfo('완료', f'생성 완료!\n\n{out}'))
            except Exception as e:
                self.after(0, lambda: self.log(f'\n❌ 오류: {e}'))
                self.after(0, lambda: self.run_btn.configure(state='normal', text='▶  생성 시작'))
                self.after(0, lambda: messagebox.showerror('오류', str(e)))

        threading.Thread(target=worker, daemon=True).start()

if __name__ == '__main__':
    App().mainloop()
