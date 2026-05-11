#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
구매조회 / 구매현황 → 연도별 시트(가공파일) 변환 코어
converter_app.py의 GUI 없는 버전 — step1_core.py에서 임포트
"""
import re, io
import pandas as pd
from openpyxl import Workbook, load_workbook

CONN    = r'(?:SC|LC|FC)'
FERR    = r'(?:PC|APC)'
TYPE_PAT = f'({CONN}/{FERR})'

# ── 파싱 함수 ────────────────────────────────────────────────

def extract_types(p):
    p = str(p).strip()
    if p.startswith('DROP'):
        m = re.search(r'\((.+?)\)', p)
        if m:
            t = re.findall(TYPE_PAT, m.group(1))
            if len(t) >= 2: return t[0], t[1]
            if len(t) == 1: return t[0], t[0]
        return '', ''
    if p.startswith('PIGTAIL'):
        t = re.findall(TYPE_PAT, p)
        return (t[0], '') if t else ('', '')
    if p.startswith('OJC-'):
        parts = p.split('-')
        conn = next((x for x in parts if re.match(r'^(SC|LC|FC)/(SC|LC|FC)$', x)), None)
        ferr = next((x for x in parts if re.match(r'^(PC|APC)/(PC|APC)$', x)), None)
        if conn and ferr:
            cA, cB = conn.split('/'); fA, fB = ferr.split('/')
            return f'{cA}/{fA}', f'{cB}/{fB}'
        return '', ''
    if p.startswith(('SOJC', 'DOJC', 'MOJC')):
        t = re.findall(TYPE_PAT, p)
        if len(t) >= 2: return t[0], t[1]
        if len(t) == 1: return t[0], t[0]
        return '', ''
    if p.startswith('Optical Cable Parts'):
        t = re.findall(TYPE_PAT, p)
        if len(t) >= 2: return t[0], t[1]
        if len(t) == 1: return t[0], t[0]
        return '', ''
    return '', ''

def derive_kind(p, g, core):
    if p.startswith('PIGTAIL'):
        return 'om1-pigtail' if re.search(r'-MM\b|-MM-|MM\(OM3\)', p) else 'pigtail'
    if p.startswith('DROP'):           return 'drop'
    if p.startswith('Optical Cable Parts'): return 'a2'
    if p.startswith('MOJC'):           return f'b3-{core}c'
    if 'OJC-C2' in p:                 return f'b3-{core}c' if core == 4 else f'a1-{core}c'
    if re.search(r'\bB3\b', g.upper()): return 'b3'
    if 'MM(OM3)' in p or re.search(r'\bOM3\b', g.upper()) or '-OM3' in p: return 'om3'
    if re.search(r'-MM\b|-MM-', p):   return 'om1'
    for c, l in [('청','a1-청'),('적','a1-적'),('녹','a1-녹'),('자','a1-자')]:
        if c in g: return l
    for c, l in [('청','a1-청'),('적','a1-적'),('녹','a1-녹'),('자','a1-자')]:
        if c in p: return l
    return 'a1'

def derive_pai(p, g, kind):
    if p.startswith('PIGTAIL'):
        if '(0.9mm)' in p or '0.9mm' in p: return 0.9
        if '(2.0mm)' in p or re.search(r'-MM\b|-MM-', p) or 'MM(OM3)' in p: return '2.0mm'
        return 0.9
    if p.startswith('DROP'): return '3.0mm'
    if re.search(r'-MM\b|-MM-', p) or 'MM(OM3)' in p or 'OM3' in p: return '2.0mm'
    if p.startswith('MOJC') or 'OJC-C2' in p or 'OJC-A1' in p: return '2.0mm'
    if p.startswith('Optical Cable Parts'): return '2.0mm'
    if p.startswith(('SOJC', 'DOJC')):
        if '3.0mm' in g or '3.0MM' in g.upper() or '3MM' in g.upper(): return '3.0mm'
        return '2.0mm'
    if '3.0mm' in g or '3.0MM' in g.upper(): return '3.0mm'
    return '2.0mm'

def derive_core(p):
    m = re.search(r'MOJC-(?:SM|MM)-(\d+)C', p)
    if m: return int(m.group(1))
    m = re.search(r'PIGTAIL-[A-Z/()]+-[A-Z]+-(\d+)C\b', p)
    if m: return int(m.group(1))
    if p.startswith('PIGTAIL'):
        m = re.search(r'-(\d+)C\b', p); return int(m.group(1)) if m else 1
    m = re.search(r'OJC-C2-.*-(\d+)C', p)
    if m: return int(m.group(1))
    if p.startswith('Optical Cable Parts'):
        m = re.search(r'(\d+)Core', p); return int(m.group(1)) if m else 1
    if p.startswith('DROP') and '2C' in p: return 2
    if p.startswith('DROP'):  return 1
    if p.startswith('DOJC'):  return 2
    if p.startswith('SOJC'):  return 1
    if p.endswith('-SP'):     return 1
    if p.endswith('-DP'):     return 2
    m = re.search(r'-(\d+)C$', p)
    if m: return int(m.group(1))
    return 1

def derive_length(p, g):
    if p.startswith('Optical Cable Parts'):
        m = re.search(r'-(\d+(?:\.\d+)?)m-', p, re.IGNORECASE)
        if m: return float(m.group(1))
    if 'OJC-A1' in p or 'OJC-C2' in p:
        m = re.search(r'-(?:SM|MM)(?:\(OM3\))?-(\d+(?:\.\d+)?)-(?:PC|APC)', p)
        if m: return float(m.group(1))
    m = re.search(r'(\d+(?:\.\d+)?)\s*[mM]\b', g)
    if m: return float(m.group(1))
    m = re.search(r'(\d+(?:\.\d+)?)\s*[mM]', g)
    if m: return float(m.group(1))
    m = re.search(r'\[(\d+(?:\.\d+)?)M', p, re.IGNORECASE)
    if m: return float(m.group(1))
    return None

MM_KINDS     = {'om1', 'om1-pigtail', 'om3'}
PRIMARY_20   = {'LC/PC':23,'LC/APC':25,'SC/PC':27,'SC/APC':29,'FC/PC':31,'FC/APC':33}
SECONDARY_20 = {'LC/PC':24,'LC/APC':26,'SC/PC':28,'SC/APC':30,'FC/PC':32,'FC/APC':34}
BEIGE_COL    = 35
PRIMARY_30   = {'LC/PC':36,'LC/APC':38,'SC/PC':40,'SC/APC':42,'FC/PC':44,'FC/APC':46}
SECONDARY_30 = {'LC/PC':37,'LC/APC':39,'SC/PC':41,'SC/APC':43,'FC/PC':45,'FC/APC':47}

def is_multicore(kind):
    return bool(re.match(r'^(b3|a1)-\d+c$', kind))

def calc_housing(row):
    t1 = row['타입1']; t2 = row['타입2']; pai = row['파이']
    kind = row['케이블종류']; core = int(row['코어수'])
    qty  = sum(int(row[mm]) for mm in range(1, 13) if row[mm])
    if qty == 0 or not t1: return {}
    is_mm = kind in MM_KINDS
    use_red = (core >= 2) and not is_multicore(kind)
    is_20 = (pai == '2.0mm'); is_30 = (pai == '3.0mm'); cps = core
    res = {}
    def add(col, val):
        if col and val: res[col] = res.get(col, 0) + val
    def add_primary(t, amount):
        if t == 'LC/PC' and is_mm and is_20: add(BEIGE_COL, amount)
        elif is_20 and t in PRIMARY_20:       add(PRIMARY_20[t], amount)
        elif is_30 and t in PRIMARY_30:       add(PRIMARY_30[t], amount)
    if t1: add_primary(t1, cps * qty)
    if t2:
        if use_red:
            if is_20 and t2 in SECONDARY_20: add(SECONDARY_20[t2], cps * qty)
            elif is_30 and t2 in SECONDARY_30: add(SECONDARY_30[t2], cps * qty)
        else:
            add_primary(t2, cps * qty)
    return res

def build_merged(df_y):
    pivot = df_y.pivot_table(
        index=['품목코드','품목명','규격명'], columns='월',
        values='수량', aggfunc='sum', fill_value=0).reset_index()
    for mm in range(1, 13):
        if mm not in pivot.columns: pivot[mm] = 0
    pivot['코어수']     = pivot['품목명'].apply(derive_core)
    pivot['케이블종류'] = pivot.apply(lambda r: derive_kind(r['품목명'], r['규격명'], r['코어수']), axis=1)
    pivot['파이']       = pivot.apply(lambda r: derive_pai(r['품목명'], r['규격명'], r['케이블종류']), axis=1)
    pivot['케이블길이'] = pivot.apply(lambda r: derive_length(r['품목명'], r['규격명']), axis=1)
    pivot[['타입1','타입2']] = pivot['품목명'].apply(lambda x: pd.Series(extract_types(x)))
    return pivot.sort_values('품목코드').reset_index(drop=True)

def write_sheets(wb_out, merged, YY):
    C1H = ['품목코드','품목명','규격명','케이블 종류','파이','코어수','케이블 길이','타입 1','타입2'] + \
          [f'{YY}년{m:02d}월' for m in range(1,13)] + ['케이블 사용량','최고제작량','최고판매 케이블 소요량']
    C2H = ['품목코드','품목명','규격명','케이블 종류','파이','코어수','케이블 길이','타입 1','타입2',
           *[f'{YY}년{m:02d}월' for m in range(1,13)],'합계',
           '2.0MM - LC/PC(청색)','2.0MM - LC/PC(적색)',
           '2.0MM - LC/APC(녹색)','2.0MM - LC/APC(적색)',
           '2.0MM - SC/PC(청색)','2.0MM - SC/PC(적색)',
           '2.0MM - SC/APC(녹색)','2.0MM - SC/APC(적색)',
           '2.0MM - FC/PC(흑색)','2.0MM - FC/PC(적색)',
           '2.0MM - FC/APC(녹색)','2.0MM - FC/APC(적색)',
           '2.0MM - BEIGE(OM1·OM3)',
           '3.0MM - LC/PC(청색)','3.0MM - LC/PC(적색)',
           '3.0MM - LC/APC(녹색)','3.0MM - LC/APC(적색)',
           '3.0MM - SC/PC(청색)','3.0MM - SC/PC(적색)',
           '3.0MM - SC/APC(녹색)','3.0MM - SC/APC(적색)',
           '3.0MM - FC/PC(흑색)','3.0MM - FC/PC(적색)',
           '3.0MM - FC/APC(녹색)','3.0MM - FC/APC(적색)',
           '검증','계산수량']

    ws_c = wb_out.create_sheet(f'{YY}년_케이블')
    ws_h = wb_out.create_sheet(f'{YY}년 하우징')
    for i, h in enumerate(C1H, 1): ws_c.cell(1, i, h)
    for i, h in enumerate(C2H, 1): ws_h.cell(1, i, h)

    for idx in range(len(merged)):
        r = idx + 2; row = merged.iloc[idx]
        for ws in [ws_c, ws_h]:
            ws.cell(r,1, row['품목코드']); ws.cell(r,2, row['품목명']); ws.cell(r,3, row['규격명'])
            ws.cell(r,4, row['케이블종류']); ws.cell(r,5, row['파이']); ws.cell(r,6, row['코어수'])
            ws.cell(r,7, row['케이블길이'] if pd.notna(row['케이블길이']) else None)
            ws.cell(r,8, row['타입1'] or None); ws.cell(r,9, row['타입2'] or None)
            for mm in range(1, 13):
                v = row[mm] if mm in row else 0
                ws.cell(r, 9+mm, int(v) if v else None)
        ws_c.cell(r,22, f"=SUM(J{r}:U{r})*G{r}")
        ws_c.cell(r,23, f"=MAX(J{r}:U{r})")
        ws_c.cell(r,24, f"=W{r}*G{r}")
        ws_h.cell(r,22, f"=SUM(J{r}:U{r})")
        for col, val in calc_housing(row).items():
            ws_h.cell(r, col, val)
        ws_h.cell(r,48, f"=IF(F{r}=1,V{r}*2,IF(F{r}=2,V{r}*F{r}*2,IF(F{r}>2,F{r}*2*V{r})))")
        ws_h.cell(r,49, f"=SUM(W{r}:AU{r})")


def _parse_buy_date(buy_no):
    """날짜 셀 값 → (year_str, month_int). datetime 객체 및 다양한 문자열 형식 처리."""
    if not buy_no:
        return None, None
    if hasattr(buy_no, 'year'):
        return str(buy_no.year), buy_no.month
    s = str(buy_no).strip()
    # YYYY/MM/DD ("2024/09/05 -3" 형식)
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})', s)
    if m: return m.group(1), int(m.group(2))
    # YYYY-MM-DD ("2025-01-09 00:00:00" datetime 문자열)
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m: return m.group(1), int(m.group(2))
    # YYYYMMDD ("20250109-2" 형식)
    m = re.match(r'^(\d{4})(\d{2})(\d{2})', s)
    if m: return m.group(1), int(m.group(2))
    # YY/MM/DD ("24/09/09-1" 발주번호 형식)
    m = re.match(r'^(\d{2})/(\d{1,2})/(\d{1,2})', s)
    if m: return '20' + m.group(1), int(m.group(2))
    return None, None


def _parse_rows_from_sheet(ws, sheet_type: str, logs: list = None) -> list:
    """워크시트 → rows 리스트 변환. sheet_type: 'ojc'(구매조회) or 'purchase'(구매현황)"""
    rows = []
    skipped_no_key = 0
    skipped_date   = 0
    sample_raw     = []   # 진단용 첫 3행 원본

    for cells in ws.iter_rows(min_row=3, values_only=True):
        if len(sample_raw) < 3:
            sample_raw.append(cells[:8])

        if sheet_type == 'ojc':
            # 구매조회 컬럼: 거래처|발주번호|발주일자|품목코드|품목명|납기지점|규격명|수량
            buy_no = cells[2] if len(cells) > 2 else None   # 발주일자 YYYY/MM/DD
            pc     = cells[3] if len(cells) > 3 else None   # 품목코드
            pr     = cells[4] if len(cells) > 4 else None   # 품목명
            gy_raw = cells[6] if len(cells) > 6 else None   # 규격명 (없을 수 있음)
            qty    = cells[7] if len(cells) > 7 else None   # 수량
        else:
            # 구매현황 컬럼: 날짜|...|...|...|품목코드|품목명[규격명]|수량
            buy_no = cells[0] if len(cells) > 0 else None
            pc     = cells[4] if len(cells) > 4 else None
            pr     = cells[5] if len(cells) > 5 else None
            gy_raw = None
            qty    = cells[6] if len(cells) > 6 else None

        if not buy_no or not pr or not qty:
            skipped_no_key += 1
            continue

        year, month = _parse_buy_date(buy_no)
        if not year:
            skipped_date += 1
            continue

        s  = re.sub(r'\s*외\s*\d+건\s*$', '', str(pr).strip())
        bm = re.match(r'^(.+?)\s*\[(.+)\]\s*$', s)
        pn = bm.group(1).strip() if bm else s
        if bm:
            gy = bm.group(2).strip()
        elif gy_raw:
            gy = str(gy_raw).strip()
        else:
            gy = ''

        try: q = int(float(str(qty)))
        except: continue
        if q <= 0: continue

        rows.append({'연도': year, '월': month,
                     '품목코드': str(pc).strip() if pc else '',
                     '품목명': pn, '규격명': gy, '수량': q})

    if logs is not None and not rows:
        logs.append(f"[진단] 샘플 행(첫 3행): {sample_raw}")
        logs.append(f"[진단] 필수 값 없음으로 제외: {skipped_no_key}행 / 날짜 형식 불일치: {skipped_date}행")

    return rows


def _detect_col_format(ws) -> str:
    """헤더 행(row 2)으로 컬럼 형식 판별.
    첫 컬럼이 거래처/구매처 계열이면 'ojc' 파서(맥산 납품 형식), 아니면 'purchase' 파서.
    """
    for cells in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        h0 = str(cells[0] or '').strip() if cells else ''
        if any(kw in h0 for kw in ('거래처', '구매처', '공급처', '납품처')):
            return 'ojc'
    return 'purchase'


def preprocess(path: str, logs: list = None) -> Workbook:
    """
    구매조회 / 구매현황 파일 → 연도별 시트 Workbook 반환
    step1_core.py에서 호출용
    """
    if logs is None: logs = []
    wb_in  = load_workbook(path, data_only=True, read_only=True)
    sheets = wb_in.sheetnames
    wb_out = Workbook(); wb_out.remove(wb_out.active)

    if '구매조회' in sheets:
        logs.append("구매조회 형식 감지 → 자동 변환 시작")
        rows = _parse_rows_from_sheet(wb_in['구매조회'], 'ojc', logs)
    elif '구매현황' in sheets:
        ws_h = wb_in['구매현황']
        fmt  = _detect_col_format(ws_h)
        if fmt == 'ojc':
            logs.append("구매현황 파일 (맥산 납품 형식: 거래처|납고번호|날짜|품목코드|...) 감지 → 자동 변환 시작")
        else:
            logs.append("구매현황 형식 감지 → 자동 변환 시작")
        rows = _parse_rows_from_sheet(ws_h, fmt, logs)
    else:
        wb_in.close()
        raise ValueError(f"지원하지 않는 파일 형식입니다. 시트 목록: {', '.join(sheets)}")

    wb_in.close()

    if not rows:
        diag = ' | '.join(l for l in logs if '[진단]' in l)
        msg  = "파싱된 데이터가 없습니다. 파일 형식을 확인해주세요."
        if diag:
            msg += f"\n{diag}"
        raise ValueError(msg)

    df = pd.DataFrame(rows)
    for year in sorted(df['연도'].unique()):
        YY  = year[-2:]
        dfY = df[df['연도'] == year].copy()
        merged = build_merged(dfY)
        write_sheets(wb_out, merged, YY)
        logs.append(f"  {year}년: {len(merged)}개 품목 변환 완료")

    return wb_out
