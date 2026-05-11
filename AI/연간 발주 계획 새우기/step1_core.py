#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STEP1 core — 웹앱에서 호출 가능한 형태
run(row_path, cable_meta, housing_meta_in, ojc_ref_path, settings)
  -> (bytes, logs, cable_keys, housing_keys, years)
연도는 워크북 시트 이름에서 자동 감지 — 1개년/2개년/5개년 모두 동작
"""
import openpyxl, re, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

PAI_ORDER   = {'2.0mm': 0, '3.0mm': 1, '0.9mm': 2}
REF_ONLY_KEYS = {('2.0mm', 'A1-4C'), ('2.0mm', 'OM4-DP')}
MM          = {'om1', 'om1-pigtail', 'om3'}

# 연도별 컬러 팔레트 (최대 8개년 대응)
YR_PALETTE  = ['2F5597', '2E75B6', '155480', '375623', '7030A0', '833C00', '1F3864', '595959']

def _fill(h): return PatternFill('solid', start_color=h)
def _font(bold=False, size=9, color='000000'): return Font(name='Arial', bold=bold, size=size, color=color)
def _bdr():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = _bdr()
CTR    = Alignment(horizontal='center', vertical='center')
CTR_W  = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

def _np(p): return '0.9mm' if str(p).strip() == '0.9' else str(p).strip()
def _imc(k): return bool(re.match(r'^(b3|a1)-\d+c$', str(k).lower() if k else ''))

def _mc2(kind, core):
    k = str(kind).strip().lower() if kind else ''
    try: c = int(core)
    except: c = 1
    sd = 'SP' if c == 1 else ('DP' if c == 2 else f'{c}C')
    bm = {'a1': 'A1', 'b3': 'B3', 'om1': 'OM1', 'om3': 'OM3',
          'a1-청': 'A1_청', 'a1-녹': 'A1_녹', 'a1-적': 'A1_적', 'a1-자': 'A1_자'}
    if k in bm: return f'{bm[k]}-{sd}'
    if k == 'drop': return f'DROP-{sd}'
    if k in ('pigtail', 'om1-pigtail'): return 'PIGTAIL'
    if k == 'a2': return 'Optical cable'
    return k.upper()

def _gp(t, kind, pai):
    if not t: return None
    t = str(t).strip(); p = _np(pai); k = str(kind).strip().lower() if kind else ''
    if t == 'LC/PC':
        return (p, 'LC/PC 베이지MM') if k in MM and p == '2.0mm' else (p, 'LC/PC 청색')
    return {'LC/APC': (p, 'LC/APC 녹색'), 'SC/PC': (p, 'SC/PC 청색'),
            'SC/APC': (p, 'SC/APC 녹색'), 'FC/PC': (p, 'FC/PC 흑색'),
            'FC/APC': (p, 'FC/APC 녹색')}.get(t)

def _gr(t, pai):
    if not t: return None
    t = str(t).strip(); p = _np(pai)
    return {'LC/PC': (p, 'LC/PC 적색'), 'LC/APC': (p, 'LC/APC 적색'),
            'SC/PC': (p, 'SC/PC 적색'), 'SC/APC': (p, 'SC/APC 적색'),
            'FC/PC': (p, 'FC/PC 적색'), 'FC/APC': (p, 'FC/APC 적색')}.get(t)


def run(row_path: str, cable_meta: dict = None, housing_meta_in: dict = None,
        ojc_ref_path: str = None, settings: dict = None,
        progress_cb=None) -> tuple:
    """
    Returns: (xlsx_bytes, logs, cable_keys, housing_keys, years)
    years: 감지된 연도 목록 (예: ['23','24','25'])
    progress_cb: optional callable(pct: int, msg: str) — 0~100
    """
    def _prog(pct, msg=''):
        if progress_cb:
            progress_cb(pct, msg)

    logs = []
    settings = settings or {}
    colors  = settings.get('colors', {})
    C_MAIN  = colors.get('main_header', '1F3864')
    converted_bytes = None  # ERP 변환 시 가공파일 바이트 저장

    # ── ERP 원본 파일 자동 감지 및 변환 ──────────────────────
    _prog(5, "파일 로드 중...")
    row_wb = openpyxl.load_workbook(row_path, read_only=True, data_only=True)
    sheets = set(row_wb.sheetnames)

    if '구매조회' in sheets or '구매현황' in sheets:
        row_wb.close()
        logs.append("ERP 원본 파일 감지 (구매조회/구매현황) — 자동 변환 시작")
        _prog(15, "ERP 원본 파일 변환 중...")
        import convert_core as _cc
        converted_wb = _cc.preprocess(row_path, logs)
        _prog(35, "변환 완료 — 연도별 시트 인식 중...")
        buf_tmp = io.BytesIO()
        converted_wb.save(buf_tmp)
        converted_wb.close()
        converted_bytes = buf_tmp.getvalue()   # 가공파일 보존
        buf_tmp.seek(0)
        row_wb = openpyxl.load_workbook(buf_tmp, read_only=True, data_only=True)
        sheets = set(row_wb.sheetnames)
    else:
        _prog(20, "연도별 시트 인식 중...")

    logs.append(f"파일 내 시트 목록: {', '.join(sorted(sheets))}")

    def _find_sheet(yr, kind):
        """연도·종류로 시트 이름 검색 — 언더스코어/공백/전체연도 모두 시도"""
        for name in [
            f'{yr}년_{kind}', f'{yr}년 {kind}',
            f'20{yr}년_{kind}', f'20{yr}년 {kind}',
            f'{yr}년_{kind.lower()}', f'{yr}년 {kind.lower()}',
        ]:
            if name in sheets:
                return name
        return None

    yr_sheet_map = {}
    for y in range(2015, 2031):
        yr = str(y)[-2:]
        cs = _find_sheet(yr, '케이블')
        if cs:
            yr_sheet_map[yr] = {'cable': cs, 'housing': _find_sheet(yr, '하우징')}

    YEARS = sorted(yr_sheet_map.keys())
    if not YEARS:
        logs.append(f"⚠ 케이블 시트를 찾지 못했습니다. 위 시트 목록을 확인해주세요.")
        row_wb.close()
        return b'', logs, [], [], [], None
    logs.append(f"감지된 연도: {', '.join('20'+y for y in YEARS)}년")
    _prog(40, f"데이터 집계 중 ({', '.join('20'+y for y in YEARS)})...")

    n = len(YEARS)
    yr_label = f"20{YEARS[0]}년" if n == 1 else f"20{YEARS[0]}~20{YEARS[-1]}년"
    n_label  = f"{n}개년"

    # 연도별 컬러 (설정값 우선, 부족하면 팔레트 순환)
    yr_palette = [
        colors.get('year_23', YR_PALETTE[0]),
        colors.get('year_24', YR_PALETTE[1]),
        colors.get('year_25', YR_PALETTE[2]),
    ] + YR_PALETTE[3:]
    yr_colors = {yr: yr_palette[i % len(yr_palette)] for i, yr in enumerate(YEARS)}

    # ── 케이블 컬럼 레이아웃 (1-indexed) ─────────────────────
    # A-E: 기본정보(5) | F...: 연간+피크 × n | 현재고 기발주 평균 피크평균 구매처 리드타임 비고
    C_YR0    = 6                  # 첫 연도 연간 컬럼
    C_CURR   = C_YR0 + 2 * n     # 현재고
    C_ORD    = C_CURR + 1         # 기발주
    C_AVG    = C_ORD  + 1         # N개년평균
    C_PAVG   = C_AVG  + 1         # N개년피크평균
    C_VEN    = C_PAVG + 1         # 구매처
    C_LT     = C_VEN  + 1         # 리드타임
    C_NOTE   = C_LT   + 1         # 비고

    # ── 하우징 컬럼 레이아웃 ─────────────────────────────────
    # A-F: 기본정보(6) | G...: 연간+피크 × n | 현재고 기발주 평균 피크평균 구매처 리드타임 비고
    H_YR0    = 7
    H_CURR   = H_YR0 + 2 * n
    H_ORD    = H_CURR + 1
    H_AVG    = H_ORD  + 1
    H_PAVG   = H_AVG  + 1
    H_VEN    = H_PAVG + 1
    H_LT     = H_VEN  + 1
    H_NOTE   = H_LT   + 1

    # ── ROW 데이터 집계 ──────────────────────────────────────
    logs.append("ROW 데이터 집계 중...")
    cable_agg   = defaultdict(lambda: {yr: [0.0]*12 for yr in YEARS})
    housing_agg = defaultdict(lambda: {yr: [0.0]*12 for yr in YEARS})

    for yr_i, yr in enumerate(YEARS):
        _prog(40 + int(20 * yr_i / max(len(YEARS), 1)), f"20{yr}년 데이터 집계 중...")
        cable_sheet = yr_sheet_map[yr]['cable']
        hous_sheet  = yr_sheet_map[yr].get('housing')
        for row in list(row_wb[cable_sheet].iter_rows(values_only=True))[1:]:
            kind, pai, core, length = row[3], row[4], row[5], row[6]
            if not kind or not pai or not length: continue
            ct = _mc2(kind, core); p_c = _np(pai)
            if ct == 'PIGTAIL' and p_c == '0.9mm': continue
            for i, q in enumerate(row[9:21]):
                if q: cable_agg[(p_c, ct)][yr][i] += float(q) * float(length)

        if not hous_sheet:
            logs.append(f"하우징 시트 없음 ({yr}년) — 건너뜀")
            continue
        for row in list(row_wb[hous_sheet].iter_rows(values_only=True))[1:]:
            kind, pai, core, t1, t2 = row[3], row[4], row[5], row[7], row[8]
            if not pai: continue
            p_h = _np(pai)
            try: cps = int(core) if core else 1
            except: cps = 1
            is_dp = (cps >= 2) and not _imc(kind)
            for mi, qty in enumerate(row[9:21]):
                if not qty: continue
                qty = float(qty)
                ha = _gp(t1, kind, p_h)
                if ha: housing_agg[ha][yr][mi] += qty * cps
                if t2:
                    hb = _gr(t2, p_h) if is_dp else _gp(t2, kind, p_h)
                    if hb: housing_agg[hb][yr][mi] += qty * cps

    def _fin(agg):
        r = {}
        for k, yd in agg.items():
            r[k] = {}
            for yr, m in yd.items():
                v = [round(x) for x in m]
                r[k][yr] = {'monthly': v, 'annual': sum(v), 'peak': max(v)}
        return r

    cs = _fin(cable_agg); hs = _fin(housing_agg)
    row_wb.close()
    logs.append(f"집계 완료 — 케이블 {len(cs)}타입 / 하우징 {len(hs)}타입")
    _prog(62, "집계 완료 — Excel 파일 생성 준비 중...")

    # ── OJC 참고파일 보완 (A1-4C, OM4-DP) ───────────────────
    if ojc_ref_path:
        try:
            ojc_ref = openpyxl.load_workbook(ojc_ref_path, data_only=True)
            for row in ojc_ref['케이블 사용내역'].iter_rows(min_row=4, values_only=True):
                pai, kind = row[1], row[2]
                if not pai or not kind: continue
                key = (str(pai).strip(), str(kind).strip())
                if key not in REF_ONLY_KEYS or key in cs: continue
                def _uni2(a):
                    b = [round(a/12)]*12 if a else [0]*12
                    b[0] += round(a or 0) - sum(b); return b
                ojc_yr_map = {yr: (5+2*i, 6+2*i) for i, yr in enumerate(['23','24','25'])}
                cs[key] = {}
                for yr in YEARS:
                    if yr in ojc_yr_map:
                        ai, pi = ojc_yr_map[yr]
                        ann  = round(float(row[ai])) if row[ai] else 0
                        peak = round(float(row[pi])) if row[pi] else 0
                    else:
                        ann = peak = 0
                    cs[key][yr] = {'monthly': _uni2(ann), 'annual': ann, 'peak': peak}
            ojc_ref.close()
            logs.append("A1-4C / OM4-DP 보완 적용")
        except Exception as e:
            logs.append(f"OJC 참고파일 보완 건너뜀: {e}")

    # ── 메타데이터 초기화 ─────────────────────────────────────
    cable_meta_d = dict(cable_meta) if cable_meta else {}
    housing_meta = defaultdict(list)
    if housing_meta_in:
        for k, v in housing_meta_in.items():
            housing_meta[k] = list(v) if isinstance(v, list) else [v]

    # OJC 참고파일 품번 보완 (REF_ONLY_KEYS 전용)
    if ojc_ref_path:
        try:
            ojc_ref2 = openpyxl.load_workbook(ojc_ref_path, data_only=True)
            for row in ojc_ref2['케이블 사용내역'].iter_rows(min_row=4, values_only=True):
                pai, kind, bunho, pname = row[1], row[2], row[3], row[4]
                if not pai or not kind or not bunho: continue
                key = (str(pai).strip(), str(kind).strip())
                if key in REF_ONLY_KEYS and key not in cable_meta_d:
                    cable_meta_d[key] = {
                        '품번': bunho, '품명': pname,
                        '구매처': row[12], '리드타임': row[13],
                    }
            ojc_ref2.close()
        except Exception:
            pass
    cable_meta = cable_meta_d

    # ── Excel 생성 ───────────────────────────────────────────
    _prog(68, "케이블 시트 생성 중...")
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    def _cell_yr_data(ws, ri, yr_idx, yr, data, rf):
        """연도 데이터 2개 컬럼(연간·피크) 기입"""
        for j, typ in enumerate(['annual', 'peak']):
            col = C_YR0 + 2 * yr_idx + j
            v   = data.get(yr, {}).get(typ, 0) or None
            c   = ws.cell(ri, col, v)
            c.font = _font(size=9); c.border = BORDER
            c.number_format = '#,##0'; c.alignment = RIGHT
            if rf: c.fill = rf

    def _write_cable_sheet(ws):
        total_cols = C_NOTE
        ws.sheet_view.showGridLines = False

        # 행 1 — 제목
        ws.row_dimensions[1].height = 24
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        c = ws['A1']
        c.value = f'{yr_label} 케이블 생산자재 사용내역 (ROW 데이터 자동 생성)'
        c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR

        # 행 2 — 그룹 헤더
        ws.row_dimensions[2].height = 18
        bands = [('A2', 'E2', '기본정보', '374151')]
        for i, yr in enumerate(YEARS):
            cs_ = get_column_letter(C_YR0 + 2 * i)
            ce_ = get_column_letter(C_YR0 + 2 * i + 1)
            bands.append((f'{cs_}2', f'{ce_}2', f"20{yr}년", yr_colors[yr]))
        bands += [
            (get_column_letter(C_CURR)+'2', get_column_letter(C_CURR)+'2', '현재고', '7030A0'),
            (get_column_letter(C_ORD)+'2',  get_column_letter(C_ORD)+'2',  '기발주', '7030A0'),
            (get_column_letter(C_AVG)+'2',  get_column_letter(C_PAVG)+'2', f'{n_label} 분석', '375623'),
            (get_column_letter(C_VEN)+'2',  get_column_letter(C_VEN)+'2',  '구매처', '595959'),
            (get_column_letter(C_LT)+'2',   get_column_letter(C_LT)+'2',   '리드타임', '595959'),
            (get_column_letter(C_NOTE)+'2', get_column_letter(C_NOTE)+'2', '비고', '595959'),
        ]
        for s, e, lbl, bg in bands:
            if s != e: ws.merge_cells(f'{s}:{e}')
            c = ws[s]; c.value = lbl
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(bg); c.alignment = CTR; c.border = BORDER

        # 행 3 — 컬럼 헤더
        ws.row_dimensions[3].height = 40
        hdrs  = ['NO', '파이', '케이블종류', '품번', '품명']
        hfill = ['374151'] * 5
        wds   = [5, 8, 18, 16, 36]
        for i, yr in enumerate(YEARS):
            hdrs  += ['연간(m)', '피크(m)']
            hfill += [yr_colors[yr]] * 2
            wds   += [14, 14]
        hdrs  += ['현재고(m)', '기발주(m)', f'{n_label}평균(m)', f'{n_label}피크평균(m)', '구매처', '리드타임(일)', '비고']
        hfill += ['7030A0', '7030A0', '375623', '375623', '595959', '595959', '595959']
        wds   += [12, 10, 14, 14, 14, 10, 20]
        for ci, (h, bf, w) in enumerate(zip(hdrs, hfill, wds), 1):
            c = ws.cell(3, ci, h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(bf); c.alignment = CTR_W; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 데이터 행
        ri = 4; no = 0; unmatched = []
        for key in sorted(cs, key=lambda k: (PAI_ORDER.get(k[0], 9), k[1])):
            pai, ct = key; d = cs[key]
            if not any(d.get(yr, {}).get('annual', 0) > 0 for yr in YEARS): continue
            no += 1; meta = cable_meta.get(key, {}); matched = bool(meta.get('품번'))
            if not matched: unmatched.append(key)
            rf = _fill('F5F5F5') if ri % 2 == 0 else None
            ws.row_dimensions[ri].height = 17

            for ci, val in enumerate([no, pai, ct, meta.get('품번',''), meta.get('품명','')], 1):
                c = ws.cell(ri, ci, val); c.font = _font(size=9); c.border = BORDER
                c.alignment = CTR if ci == 1 else LEFT
                if rf: c.fill = rf

            for i, yr in enumerate(YEARS):
                _cell_yr_data(ws, ri, i, yr, d, rf)

            ws.cell(ri, C_CURR, None).font = _font(size=9)
            ws.cell(ri, C_CURR).border = BORDER; ws.cell(ri, C_CURR).number_format = '#,##0'
            ws.cell(ri, C_CURR).alignment = RIGHT
            if rf: ws.cell(ri, C_CURR).fill = rf

            ws.cell(ri, C_ORD, None).font = _font(size=9)
            ws.cell(ri, C_ORD).border = BORDER; ws.cell(ri, C_ORD).number_format = '#,##0'
            ws.cell(ri, C_ORD).alignment = RIGHT
            if rf: ws.cell(ri, C_ORD).fill = rf

            # N개년 평균 수식 — 모든 연간 컬럼 평균
            ann_cols = ','.join(f'{get_column_letter(C_YR0+2*i)}{ri}' for i in range(n))
            pk_cols  = ','.join(f'{get_column_letter(C_YR0+2*i+1)}{ri}' for i in range(n))
            for col, fml in [(C_AVG, f'=ROUND(AVERAGE({ann_cols}),0)'),
                             (C_PAVG, f'=ROUND(AVERAGE({pk_cols}),0)')]:
                c = ws.cell(ri, col, fml); c.font = _font(size=9); c.border = BORDER
                c.number_format = '#,##0'; c.alignment = RIGHT
                if rf: c.fill = rf

            ws.cell(ri, C_VEN, meta.get('구매처', '')).font = _font(size=9)
            ws.cell(ri, C_VEN).border = BORDER; ws.cell(ri, C_VEN).alignment = LEFT
            if rf: ws.cell(ri, C_VEN).fill = rf

            ws.cell(ri, C_LT, meta.get('리드타임', '')).font = _font(size=9)
            ws.cell(ri, C_LT).border = BORDER; ws.cell(ri, C_LT).alignment = CTR
            if rf: ws.cell(ri, C_LT).fill = rf

            from_ref = key in REF_ONLY_KEYS
            nv = '⚠ 품번 미입력' if not matched else ('※ OJC 참고 기준' if from_ref else '')
            c = ws.cell(ri, C_NOTE, nv)
            c.font = Font(name='Arial', size=8, color='C00000' if not matched else '595959')
            c.border = BORDER; c.alignment = LEFT
            if not matched: c.fill = _fill('FFD7D7')
            elif rf: c.fill = rf
            ri += 1

        # 합계 행
        last = ri - 1; tr = ri + 1; ws.row_dimensions[tr].height = 18
        c = ws.cell(tr, 2, '합  계')
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR; c.border = BORDER
        sum_cols = [C_YR0+2*i+j for i in range(n) for j in range(2)] + [C_AVG, C_PAVG]
        for ci in sum_cols:
            cl = get_column_letter(ci)
            c = ws.cell(tr, ci, f'=SUM({cl}4:{cl}{last})')
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(C_MAIN); c.number_format = '#,##0'; c.alignment = RIGHT; c.border = BORDER
        ws.freeze_panes = 'D4'
        return no, unmatched

    def _cell_h_yr_data(ws, ri2, yr_idx, yr, data, rf):
        for j, typ in enumerate(['annual', 'peak']):
            col = H_YR0 + 2 * yr_idx + j
            v   = data.get(yr, {}).get(typ, 0) or None
            c   = ws.cell(ri2, col, v)
            c.font = _font(size=9); c.border = BORDER
            c.number_format = '#,##0'; c.alignment = RIGHT
            if rf: c.fill = rf

    def _write_housing_sheet(ws):
        total_cols = H_NOTE
        ws.sheet_view.showGridLines = False

        ws.row_dimensions[1].height = 24
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        c = ws['A1']
        c.value = f'{yr_label} 하우징 생산자재 사용내역 (ROW 데이터 자동 생성)'
        c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        c.fill = _fill(C_MAIN); c.alignment = CTR

        ws.row_dimensions[2].height = 18
        bands = [('A2', 'F2', '기본정보', '374151')]
        for i, yr in enumerate(YEARS):
            hs_ = get_column_letter(H_YR0 + 2 * i)
            he_ = get_column_letter(H_YR0 + 2 * i + 1)
            bands.append((f'{hs_}2', f'{he_}2', f"20{yr}년", yr_colors[yr]))
        bands += [
            (get_column_letter(H_CURR)+'2', get_column_letter(H_CURR)+'2', '현재고', '7030A0'),
            (get_column_letter(H_ORD)+'2',  get_column_letter(H_ORD)+'2',  '기발주', '7030A0'),
            (get_column_letter(H_AVG)+'2',  get_column_letter(H_PAVG)+'2', f'{n_label} 분석', '375623'),
            (get_column_letter(H_VEN)+'2',  get_column_letter(H_VEN)+'2',  '구매처', '595959'),
            (get_column_letter(H_LT)+'2',   get_column_letter(H_LT)+'2',   '리드타임', '595959'),
            (get_column_letter(H_NOTE)+'2', get_column_letter(H_NOTE)+'2', '비고', '595959'),
        ]
        for s, e, lbl, bg in bands:
            if s != e: ws.merge_cells(f'{s}:{e}')
            c = ws[s]; c.value = lbl
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(bg); c.alignment = CTR; c.border = BORDER

        ws.row_dimensions[3].height = 40
        hdrs2  = ['NO', '파이', '하우징타입', '품번', '품명', '구매용도']
        hfill2 = ['374151'] * 6
        wds2   = [5, 8, 16, 16, 36, 14]
        for i, yr in enumerate(YEARS):
            hdrs2  += ['연간', '피크']
            hfill2 += [yr_colors[yr]] * 2
            wds2   += [12, 12]
        hdrs2  += ['현재고', '기발주', f'{n_label}평균', f'{n_label}피크평균', '구매처', '리드타임(일)', '비고']
        hfill2 += ['7030A0', '7030A0', '375623', '375623', '595959', '595959', '595959']
        wds2   += [10, 10, 12, 12, 14, 10, 20]
        for ci, (h, bf, w) in enumerate(zip(hdrs2, hfill2, wds2), 1):
            c = ws.cell(3, ci, h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = _fill(bf); c.alignment = CTR_W; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w

        ri2 = 4; no2 = 0
        for key in sorted(hs, key=lambda k: (PAI_ORDER.get(k[0], 9), k[1])):
            pai, htype = key; d = hs[key]
            if not any(d.get(yr, {}).get('annual', 0) > 0 for yr in YEARS): continue
            comps = housing_meta.get(key, [{}]); first_u = ri2
            for ci2, comp in enumerate(comps):
                no2 += 1; matched = bool(comp and comp.get('품번'))
                rf = _fill('F5F5F5') if ri2 % 2 == 0 else None
                ws.row_dimensions[ri2].height = 17
                for col, val in [(1, no2 if ci2==0 else ''), (2, pai if ci2==0 else ''), (3, htype if ci2==0 else '')]:
                    c = ws.cell(ri2, col, val); c.font = _font(size=9); c.border = BORDER; c.alignment = CTR
                    if rf: c.fill = rf
                for ci3, val in enumerate([comp.get('품번','') if comp else '', comp.get('품명','') if comp else ''], 4):
                    ws.cell(ri2, ci3, val).font = _font(size=9)
                    ws.cell(ri2, ci3).border = BORDER; ws.cell(ri2, ci3).alignment = LEFT
                    if rf: ws.cell(ri2, ci3).fill = rf
                pn = comp.get('품명','') if comp else ''; usage = ''
                for kw, label in [('SPRING','스프링'),('HOUSING','하우징 본체'),('DUST CAP','더스트캡'),
                                   ('STOPPER','스토퍼'),('FRAME','플러그 프레임'),('BOOT','부트'),('KIT','하우징 키트')]:
                    if kw in str(pn).upper(): usage = label; break
                ws.cell(ri2, 6, usage).font = _font(size=8, color='595959')
                ws.cell(ri2, 6).border = BORDER; ws.cell(ri2, 6).alignment = CTR
                if rf: ws.cell(ri2, 6).fill = rf

                if ci2 == 0:
                    for i, yr in enumerate(YEARS):
                        _cell_h_yr_data(ws, ri2, i, yr, d, rf)
                    ann_cols_h = ','.join(f'{get_column_letter(H_YR0+2*i)}{ri2}' for i in range(n))
                    pk_cols_h  = ','.join(f'{get_column_letter(H_YR0+2*i+1)}{ri2}' for i in range(n))
                    for col, fml in [(H_AVG,  f'=ROUND(AVERAGE({ann_cols_h}),0)'),
                                     (H_PAVG, f'=ROUND(AVERAGE({pk_cols_h}),0)')]:
                        c3 = ws.cell(ri2, col, fml); c3.font = _font(size=9); c3.border = BORDER
                        c3.number_format = '#,##0'; c3.alignment = RIGHT
                        if rf: c3.fill = rf
                else:
                    for col in range(H_YR0, H_PAVG + 1):
                        cl = get_column_letter(col)
                        c3 = ws.cell(ri2, col, f'={cl}{first_u}')
                        c3.font = _font(size=9); c3.border = BORDER
                        c3.number_format = '#,##0'; c3.alignment = RIGHT
                        if rf: c3.fill = rf

                for col, fld in [(H_CURR, '현재고'), (H_ORD, '기발주')]:
                    v = comp.get(fld) if comp and comp.get(fld) else None
                    ws.cell(ri2, col, v).font = _font(size=9); ws.cell(ri2, col).border = BORDER
                    ws.cell(ri2, col).number_format = '#,##0'; ws.cell(ri2, col).alignment = RIGHT
                    if rf: ws.cell(ri2, col).fill = rf

                ws.cell(ri2, H_VEN,  comp.get('구매처','') if comp else '').font = _font(size=9)
                ws.cell(ri2, H_VEN).border  = BORDER; ws.cell(ri2, H_VEN).alignment  = LEFT
                if rf: ws.cell(ri2, H_VEN).fill = rf
                ws.cell(ri2, H_LT, comp.get('리드타임','') if comp else '').font = _font(size=9)
                ws.cell(ri2, H_LT).border = BORDER; ws.cell(ri2, H_LT).alignment = CTR
                if rf: ws.cell(ri2, H_LT).fill = rf

                nv = '' if matched else '⚠ 품번 미입력'
                if pai == '0.9mm' and not nv: nv = '※ OJC 참고 기준'
                ws.cell(ri2, H_NOTE, nv).font = Font(name='Arial', size=8,
                    color='C00000' if '⚠' in nv else '595959')
                ws.cell(ri2, H_NOTE).border = BORDER; ws.cell(ri2, H_NOTE).alignment = LEFT
                if '⚠' in nv: ws.cell(ri2, H_NOTE).fill = _fill('FFD7D7')
                elif rf: ws.cell(ri2, H_NOTE).fill = rf
                ri2 += 1
        ws.freeze_panes = 'D4'
        return ri2 - 4

    ws_c = wb.create_sheet('케이블 사용내역')
    _prog(72, "케이블 사용내역 시트 작성 중...")
    n_cable, unmatched = _write_cable_sheet(ws_c)
    _prog(88, "하우징 사용내역 시트 작성 중...")
    ws_h = wb.create_sheet('하우징 사용내역')
    n_housing = _write_housing_sheet(ws_h)
    _prog(96, "파일 저장 중...")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    logs.append(f"Excel 생성 완료 — 케이블 {n_cable}행 / 하우징 {n_housing}행")
    if unmatched:
        logs.append(f"⚠ 품번 미입력 {len(unmatched)}개: {', '.join(str(k) for k in unmatched)}")
    _prog(100, "완료")

    return buf.read(), logs, list(cs.keys()), list(hs.keys()), YEARS, converted_bytes
