import openpyxl
wb = openpyxl.load_workbook('G3HVNQDBO0MAJPX.xlsx', read_only=True, data_only=True)
print('sheets:', wb.sheetnames)
ws = wb.active
print('max_row:', ws.max_row, 'max_col:', ws.max_column)
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
    print(f'Row{i+1}:', r)
wb.close()
