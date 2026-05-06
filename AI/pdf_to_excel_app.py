"""
PDF → Excel 변환기
버전: v1.0 (2026-04-27)
실행: python pdf_to_excel_app.py
빌드: pyinstaller --onefile --windowed pdf_to_excel_app.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import sys
import traceback

# ============================================================
# 변환 엔진
# ============================================================

def convert_pdf_to_excel(pdf_path: str, out_path: str, single_sheet: bool,
                          progress_cb=None, status_cb=None):
    """PDF를 Excel로 변환하는 핵심 로직"""

    import pdfplumber
    import fitz  # PyMuPDF
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    import io, tempfile, math

    def thin(color='AAAAAA'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def cell_style(cell, bold=False, size=9, bg=None, color='000000',
                   halign='left', valign='center', wrap=True, border=False, italic=False):
        cell.font = Font(name='Arial', bold=bold, size=size, color=color, italic=italic)
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if border:
            cell.border = thin()

    def merge_write(ws, r1, c1, r2, c2, val, **kw):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=val)
        cell_style(cell, **kw)
        return cell

    FILL_PAGE   = PatternFill('solid', start_color='1F3864')
    FILL_TBL_H  = PatternFill('solid', start_color='2E75B6')
    FILL_ALT    = PatternFill('solid', start_color='EBF3FB')
    FILL_WHITE  = PatternFill('solid', start_color='FFFFFF')
    FILL_TEXT   = PatternFill('solid', start_color='F9F9F9')

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    # PyMuPDF로 이미지 추출
    fitz_doc = fitz.open(pdf_path)
    plumber_pdf = pdfplumber.open(pdf_path)
    total_pages = len(plumber_pdf.pages)

    # 임시 이미지 폴더
    tmp_dir = tempfile.mkdtemp()

    # 단일 시트 모드: 하나의 ws에 모든 페이지 추가
    if single_sheet:
        ws = wb.create_sheet("변환결과")
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions['A'].width = 32
        current_row = [1]   # mutable int
    else:
        ws = None
        current_row = [1]

    def get_ws(page_idx):
        nonlocal ws
        if single_sheet:
            return ws
        sheet_name = f"Page {page_idx + 1}"
        w = wb.create_sheet(sheet_name)
        for col in range(1, 13):
            w.column_dimensions[get_column_letter(col)].width = 14
        w.column_dimensions['A'].width = 32
        return w

    for page_idx, page in enumerate(plumber_pdf.pages):
        if status_cb:
            status_cb(f"페이지 {page_idx + 1} / {total_pages} 처리 중...")
        if progress_cb:
            progress_cb(int(page_idx / total_pages * 85))

        cur_ws = get_ws(page_idx)
        if not single_sheet:
            current_row = [1]

        r = current_row[0]

        # ── 페이지 구분 헤더 ──
        if single_sheet or True:
            merge_write(cur_ws, r, 1, r, 12,
                        f"▌ Page {page_idx + 1}",
                        bold=True, size=10, color='FFFFFF', halign='left')
            cur_ws.cell(row=r, column=1).fill = FILL_PAGE
            for c in range(2, 13):
                cur_ws.cell(row=r, column=c).fill = FILL_PAGE
            cur_ws.row_dimensions[r].height = 16
            r += 1

        # ── 이미지 삽입 (PyMuPDF) ──
        fitz_page = fitz_doc[page_idx]
        img_infos = fitz_page.get_image_info()
        page_w = fitz_page.rect.width    # pt
        page_h = fitz_page.rect.height   # pt

        for img_idx, img_info in enumerate(img_infos):
            iw, ih = img_info['width'], img_info['height']
            # 실제 의미 있는 이미지만 (2px 구분선 제외)
            if ih <= 5 or iw <= 5:
                continue
            xref = None
            for xref_item in fitz_page.get_images(full=True):
                base = fitz_doc.extract_image(xref_item[0])
                if base['width'] == iw and base['height'] == ih:
                    xref = xref_item[0]
                    break
            if xref is None:
                continue

            base = fitz_doc.extract_image(xref)
            img_bytes = base['image']
            ext = base['ext']

            # PDF 내 이미지 위치 비율로 Excel 열 위치 추정
            bbox = img_info['bbox']   # (x0, y0, x1, y1) in pt
            x_ratio = (bbox[0] + bbox[2]) / 2 / page_w
            anchor_col = max(1, min(12, int(x_ratio * 12) + 1))

            # 이미지 크기 (Excel pt 단위, 최대 400px 너비)
            scale = min(400 / iw, 200 / ih, 1.0)
            xl_w = int(iw * scale)
            xl_h = int(ih * scale)
            needed_rows = max(1, math.ceil(xl_h / 15))

            img_path = os.path.join(tmp_dir, f"p{page_idx}_i{img_idx}.{ext}")
            with open(img_path, 'wb') as f:
                f.write(img_bytes)

            for rr in range(r, r + needed_rows):
                cur_ws.row_dimensions[rr].height = 15

            try:
                xl_img = XLImage(img_path)
                xl_img.width  = xl_w
                xl_img.height = xl_h
                col_letter = get_column_letter(anchor_col)
                xl_img.anchor = f"{col_letter}{r}"
                cur_ws.add_image(xl_img)
            except Exception:
                pass

            r += needed_rows

        # ── 텍스트 & 테이블 추출 ──
        tables = page.extract_tables()
        table_bboxes = [tbl.bbox for tbl in page.find_tables()] if tables else []

        # 테이블 먼저
        for t_idx, (tbl, bbox) in enumerate(zip(tables, table_bboxes)):
            if not tbl:
                continue
            # 테이블 제목 행 (첫 행이 헤더인지 확인)
            has_header = any(cell for cell in (tbl[0] or []) if cell)
            header_row = tbl[0] if has_header else None
            data_start  = 1 if has_header else 0

            num_cols = max(len(row) for row in tbl)
            # 컬럼 수에 맞게 병합 범위 결정 (최대 12컬럼)
            col_span = min(num_cols, 12)

            if header_row:
                for c_idx, cell_val in enumerate(header_row[:col_span], 1):
                    val = str(cell_val).replace('\n', ' ') if cell_val else ''
                    c = cur_ws.cell(row=r, column=c_idx, value=val)
                    cell_style(c, bold=True, size=9, bg='2E75B6', color='FFFFFF',
                               halign='center', border=True)
                cur_ws.row_dimensions[r].height = 18
                r += 1

            for row_idx, row in enumerate(tbl[data_start:]):
                fill = FILL_ALT if row_idx % 2 == 0 else FILL_WHITE
                for c_idx, cell_val in enumerate(row[:col_span], 1):
                    val = str(cell_val).replace('\n', ' ') if cell_val else ''
                    c = cur_ws.cell(row=r, column=c_idx, value=val)
                    cell_style(c, size=9, bg=fill.start_color.rgb,
                               halign='center' if c_idx > 2 else 'left',
                               border=True)
                cur_ws.row_dimensions[r].height = 14
                r += 1

            r += 1   # 테이블 후 빈 행

        # 텍스트 블록 (테이블 외 영역)
        text = page.extract_text()
        if text:
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            for line in lines:
                merge_write(cur_ws, r, 1, r, 12, line,
                            size=9, halign='left', bg='F9F9F9')
                cur_ws.row_dimensions[r].height = 13
                r += 1

        r += 1   # 페이지 후 공백
        current_row[0] = r

    plumber_pdf.close()
    fitz_doc.close()

    if progress_cb:
        progress_cb(92)
    if status_cb:
        status_cb("Excel 파일 저장 중...")

    wb.save(out_path)

    if progress_cb:
        progress_cb(100)
    if status_cb:
        status_cb("✅ 변환 완료!")


# ============================================================
# GUI
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF → Excel 변환기  v1.0")
        self.resizable(False, False)
        self.configure(bg='#F0F4FA')

        # DPI 보정
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w  = self.winfo_width()
        h  = self.winfo_height()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        PAD = dict(padx=18, pady=8)
        BG  = '#F0F4FA'
        ACCENT = '#1F3864'

        # ── 헤더 ──
        hdr = tk.Frame(self, bg=ACCENT, height=56)
        hdr.pack(fill='x')
        tk.Label(hdr, text="  📄→📊  PDF → Excel 변환기",
                 bg=ACCENT, fg='white',
                 font=('맑은 고딕', 14, 'bold')).pack(side='left', padx=16, pady=12)

        # ── 본문 ──
        body = tk.Frame(self, bg=BG, padx=20, pady=16)
        body.pack(fill='both', expand=True)

        # PDF 파일 선택
        tk.Label(body, text="PDF 파일", bg=BG,
                 font=('맑은 고딕', 10, 'bold'), fg=ACCENT).grid(row=0, column=0, sticky='w', pady=(0,2))

        file_row = tk.Frame(body, bg=BG)
        file_row.grid(row=1, column=0, sticky='ew', pady=(0,10))

        self.pdf_var = tk.StringVar()
        entry = tk.Entry(file_row, textvariable=self.pdf_var, width=46,
                         font=('맑은 고딕', 9), relief='solid', bd=1)
        entry.pack(side='left', ipady=4)

        tk.Button(file_row, text="  찾아보기  ",
                  command=self._browse_pdf,
                  bg=ACCENT, fg='white',
                  font=('맑은 고딕', 9, 'bold'),
                  relief='flat', cursor='hand2',
                  activebackground='#2E75B6',
                  activeforeground='white').pack(side='left', padx=(6,0), ipady=4)

        # 저장 위치
        tk.Label(body, text="저장 위치", bg=BG,
                 font=('맑은 고딕', 10, 'bold'), fg=ACCENT).grid(row=2, column=0, sticky='w', pady=(0,2))

        out_row = tk.Frame(body, bg=BG)
        out_row.grid(row=3, column=0, sticky='ew', pady=(0,10))

        self.out_var = tk.StringVar()
        tk.Entry(out_row, textvariable=self.out_var, width=46,
                 font=('맑은 고딕', 9), relief='solid', bd=1).pack(side='left', ipady=4)
        tk.Button(out_row, text="  찾아보기  ",
                  command=self._browse_out,
                  bg='#5B7BA6', fg='white',
                  font=('맑은 고딕', 9, 'bold'),
                  relief='flat', cursor='hand2',
                  activebackground='#2E75B6',
                  activeforeground='white').pack(side='left', padx=(6,0), ipady=4)

        # 옵션
        opt_frame = tk.LabelFrame(body, text="  옵션  ", bg=BG,
                                  font=('맑은 고딕', 9), fg=ACCENT,
                                  relief='solid', bd=1, padx=10, pady=6)
        opt_frame.grid(row=4, column=0, sticky='ew', pady=(0,12))

        self.sheet_var = tk.StringVar(value='single')
        tk.Radiobutton(opt_frame, text="단일 시트 (모든 페이지를 하나의 시트에)",
                       variable=self.sheet_var, value='single',
                       bg=BG, font=('맑은 고딕', 9)).pack(anchor='w')
        tk.Radiobutton(opt_frame, text="페이지별 시트 (각 페이지를 별도 시트에)",
                       variable=self.sheet_var, value='multi',
                       bg=BG, font=('맑은 고딕', 9)).pack(anchor='w')

        # 변환 버튼
        self.btn_convert = tk.Button(
            body, text="  ▶  변환 시작  ",
            command=self._start_convert,
            bg='#1F7A4B', fg='white',
            font=('맑은 고딕', 11, 'bold'),
            relief='flat', cursor='hand2',
            activebackground='#25A060',
            activeforeground='white')
        self.btn_convert.grid(row=5, column=0, sticky='ew', ipady=8, pady=(0,10))

        # 진행률
        self.progress = ttk.Progressbar(body, length=440, mode='determinate')
        self.progress.grid(row=6, column=0, sticky='ew', pady=(0,4))

        self.status_var = tk.StringVar(value="PDF 파일을 선택하고 변환을 시작하세요.")
        tk.Label(body, textvariable=self.status_var, bg=BG,
                 font=('맑은 고딕', 9), fg='#444444').grid(row=7, column=0, sticky='w')

        # 결과 열기 버튼
        self.btn_open = tk.Button(
            body, text="  📂  결과 파일 열기  ",
            command=self._open_result,
            bg='#2E75B6', fg='white',
            font=('맑은 고딕', 10, 'bold'),
            relief='flat', cursor='hand2',
            state='disabled',
            activebackground='#1F3864',
            activeforeground='white')
        self.btn_open.grid(row=8, column=0, sticky='ew', ipady=6, pady=(8,0))

        body.columnconfigure(0, weight=1)
        self._result_path = None

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF 파일", "*.pdf"), ("모든 파일", "*.*")])
        if path:
            self.pdf_var.set(path)
            # 저장 경로 자동 제안
            base = os.path.splitext(path)[0]
            self.out_var.set(base + "_변환.xlsx")
            self.btn_open.config(state='disabled')
            self.progress['value'] = 0
            self.status_var.set("변환 시작 버튼을 누르세요.")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="저장 위치 선택",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")])
        if path:
            self.out_var.set(path)

    def _start_convert(self):
        pdf_path = self.pdf_var.get().strip()
        out_path = self.out_var.get().strip()

        if not pdf_path:
            messagebox.showwarning("경고", "PDF 파일을 먼저 선택하세요.")
            return
        if not os.path.isfile(pdf_path):
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다:\n{pdf_path}")
            return
        if not out_path:
            messagebox.showwarning("경고", "저장 위치를 입력하세요.")
            return

        single = (self.sheet_var.get() == 'single')
        self.btn_convert.config(state='disabled')
        self.btn_open.config(state='disabled')
        self.progress['value'] = 0
        self._result_path = None

        def worker():
            try:
                convert_pdf_to_excel(
                    pdf_path, out_path, single,
                    progress_cb=lambda v: self.after(0, lambda: self.progress.configure(value=v)),
                    status_cb=lambda s: self.after(0, lambda: self.status_var.set(s))
                )
                self._result_path = out_path
                self.after(0, self._on_success)
            except Exception as e:
                err = traceback.format_exc()
                self.after(0, lambda: self._on_error(str(e), err))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self):
        self.btn_convert.config(state='normal')
        self.btn_open.config(state='normal')
        messagebox.showinfo("완료", f"변환이 완료되었습니다!\n\n{self._result_path}")

    def _on_error(self, msg, detail):
        self.btn_convert.config(state='normal')
        self.status_var.set(f"❌ 오류 발생: {msg}")
        messagebox.showerror("변환 오류",
                             f"변환 중 오류가 발생했습니다:\n\n{msg}\n\n"
                             f"자세한 내용:\n{detail[:600]}")

    def _open_result(self):
        if self._result_path and os.path.isfile(self._result_path):
            os.startfile(self._result_path)


# ============================================================
# 진입점
# ============================================================
if __name__ == '__main__':
    app = App()
    app.mainloop()
