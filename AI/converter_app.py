"""
AJW 구매조회 → 가공파일 변환기
버전: v4 (2026-04-24)
실행: python converter_app.py
빌드: pyinstaller --onefile --windowed converter_app.py
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, re, os, sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 변환 로직 (convert.py v4 통합본)
# ============================================================
CONN = r'(?:SC|LC|FC)'; FERR = r'(?:PC|APC)'; TYPE_PAT = f'({CONN}/{FERR})'

def extract_types(p):
    p = str(p).strip()
    if p.startswith('DROP'):
        m = re.search(r'\((.+?)\)', p)
        if m:
            t = re.findall(TYPE_PAT, m.group(1))
            if len(t) >= 2: return t[0], t[1]
            if len(t) == 1: return t[0], t[0]
        return '', ''
    if p.startswith('PIGTAIL'):
        t = re.findall(TYPE_PAT, p); return (t[0], '') if t else ('', '')
    if p.startswith('OJC-'):
        parts = p.split('-')
        conn = next((x for x in parts if re.match(r'^(SC|LC|FC)/(SC|LC|FC)$', x)), None)
        ferr = next((x for x in parts if re.match(r'^(PC|APC)/(PC|APC)$', x)), None)
        if conn and ferr:
            cA, cB = conn.split('/'); fA, fB = ferr.split('/')
            return f'{cA}/{fA}', f'{cB}/{fB}'
        return '', ''
    if p.startswith(('SOJC', 'DOJC', 'MOJC')):
        t = re.findall(TYPE_PAT, p)
        if len(t) >= 2: return t[0], t[1]
        if len(t) == 1: return t[0], t[0]
        return '', ''
    if p.startswith('Optical Cable Parts'):
        t = re.findall(TYPE_PAT, p)
        if len(t) >= 2: return t[0], t[1]
        if len(t) == 1: return t[0], t[0]
        return '', ''
    return '', ''

def derive_kind(p, g, core):
    if p.startswith('PIGTAIL'): return 'om1-pigtail' if re.search(r'-MM\b|-MM-|MM\(OM3\)', p) else 'pigtail'
    if p.startswith('DROP'): return 'drop'
    if p.startswith('Optical Cable Parts'): return 'a2'
    if p.startswith('MOJC'): return f'b3-{core}c'
    if 'OJC-C2' in p: return f'b3-{core}c' if core == 4 else f'a1-{core}c'
    if re.search(r'\bB3\b', g.upper()): return 'b3'
    if 'MM(OM3)' in p or re.search(r'\bOM3\b', g.upper()) or '-OM3' in p: return 'om3'
    if re.search(r'-MM\b|-MM-', p): return 'om1'
    for c, l in [('청', 'a1-청'), ('적', 'a1-적'), ('녹', 'a1-녹'), ('자', 'a1-자')]:
        if c in g: return l
    for c, l in [('청', 'a1-청'), ('적', 'a1-적'), ('녹', 'a1-녹'), ('자', 'a1-자')]:
        if c in p: return l
    return 'a1'

def derive_pai(p, g, kind):
    if p.startswith('PIGTAIL'):
        if '(0.9mm)' in p or '0.9mm' in p: return 0.9
        if '(2.0mm)' in p or re.search(r'-MM\b|-MM-', p) or 'MM(OM3)' in p: return '2.0mm'
        return 0.9
    if p.startswith('DROP'): return '3.0mm'
    if re.search(r'-MM\b|-MM-', p) or 'MM(OM3)' in p or 'OM3' in p: return '2.0mm'
    if p.startswith('MOJC') or 'OJC-C2' in p or 'OJC-A1' in p: return '2.0mm'
    if p.startswith('Optical Cable Parts'): return '2.0mm'
    if p.startswith(('SOJC', 'DOJC')):
        if '3.0mm' in g or '3.0MM' in g.upper() or '3MM' in g.upper(): return '3.0mm'
        return '2.0mm'
    if '3.0mm' in g or '3.0MM' in g.upper(): return '3.0mm'
    return '2.0mm'

def derive_core(p):
    m = re.search(r'MOJC-(?:SM|MM)-(\d+)C', p)
    if m: return int(m.group(1))
    m = re.search(r'PIGTAIL-[A-Z/()]+-[A-Z]+-(\d+)C\b', p)
    if m: return int(m.group(1))
    if p.startswith('PIGTAIL'):
        m = re.search(r'-(\d+)C\b', p); return int(m.group(1)) if m else 1
    m = re.search(r'OJC-C2-.*-(\d+)C', p)
    if m: return int(m.group(1))
    if p.startswith('Optical Cable Parts'):
        m = re.search(r'(\d+)Core', p); return int(m.group(1)) if m else 1
    if p.startswith('DROP') and '2C' in p: return 2
    if p.startswith('DROP'): return 1
    if p.startswith('DOJC'): return 2
    if p.startswith('SOJC'): return 1
    if p.endswith('-SP'): return 1
    if p.endswith('-DP'): return 2
    m = re.search(r'-(\d+)C$', p)
    if m: return int(m.group(1))
    return 1

def derive_length(p, g):
    if p.startswith('Optical Cable Parts'):
        m = re.search(r'-(\d+(?:\.\d+)?)m-', p, re.IGNORECASE)
        if m: return float(m.group(1))
    if 'OJC-A1' in p or 'OJC-C2' in p:
        m = re.search(r'-(?:SM|MM)(?:\(OM3\))?-(\d+(?:\.\d+)?)-(?:PC|APC)', p)
        if m: return float(m.group(1))
    m = re.search(r'(\d+(?:\.\d+)?)\s*[mM]\b', g)
    if m: return float(m.group(1))
    m = re.search(r'(\d+(?:\.\d+)?)\s*[mM]', g)
    if m: return float(m.group(1))
    m = re.search(r'\[(\d+(?:\.\d+)?)M', p, re.IGNORECASE)
    if m: return float(m.group(1))
    return None

MM_KINDS     = {'om1', 'om1-pigtail', 'om3'}
PRIMARY_20   = {'LC/PC': 23, 'LC/APC': 25, 'SC/PC': 27, 'SC/APC': 29, 'FC/PC': 31, 'FC/APC': 33}
SECONDARY_20 = {'LC/PC': 24, 'LC/APC': 26, 'SC/PC': 28, 'SC/APC': 30, 'FC/PC': 32, 'FC/APC': 34}
BEIGE_COL    = 35
PRIMARY_30   = {'LC/PC': 36, 'LC/APC': 38, 'SC/PC': 40, 'SC/APC': 42, 'FC/PC': 44, 'FC/APC': 46}
SECONDARY_30 = {'LC/PC': 37, 'LC/APC': 39, 'SC/PC': 41, 'SC/APC': 43, 'FC/PC': 45, 'FC/APC': 47}

def is_multicore(kind):
    return bool(re.match(r'^(b3|a1)-\d+c$', kind))

def calc_housing(row):
    t1 = row['타입1']; t2 = row['타입2']; pai = row['파이']
    kind = row['케이블종류']; core = int(row['코어수'])
    qty = sum(int(row[mm]) for mm in range(1, 13) if row[mm])
    if qty == 0 or not t1: return {}
    is_mm    = kind in MM_KINDS
    use_red  = (core >= 2) and not is_multicore(kind)
    is_20    = (pai == '2.0mm'); is_30 = (pai == '3.0mm'); cps = core
    res = {}
    def add(col, val):
        if col and val: res[col] = res.get(col, 0) + val
    def add_primary(t, amount):
        if t == 'LC/PC' and is_mm and is_20: add(BEIGE_COL, amount)
        elif is_20 and t in PRIMARY_20:       add(PRIMARY_20[t], amount)
        elif is_30 and t in PRIMARY_30:       add(PRIMARY_30[t], amount)
    if t1: add_primary(t1, cps * qty)
    if t2:
        if use_red:
            if is_20 and t2 in SECONDARY_20: add(SECONDARY_20[t2], cps * qty)
            elif is_30 and t2 in SECONDARY_30: add(SECONDARY_30[t2], cps * qty)
        else:
            add_primary(t2, cps * qty)
    return res

def build_merged(df_y):
    info  = df_y.groupby(['품목코드', '품목명', '규격명']).size().reset_index(name='_')[['품목코드', '품목명', '규격명']]
    pivot = df_y.pivot_table(index=['품목코드', '품목명', '규격명'], columns='월',
                              values='수량', aggfunc='sum', fill_value=0).reset_index()
    for mm in range(1, 13):
        if mm not in pivot.columns: pivot[mm] = 0
    merged = info.merge(pivot, on=['품목코드', '품목명', '규격명'], how='left').fillna(0)
    merged['코어수']    = merged['품목명'].apply(derive_core)
    merged['케이블종류'] = merged.apply(lambda r: derive_kind(r['품목명'], r['규격명'], r['코어수']), axis=1)
    merged['파이']      = merged.apply(lambda r: derive_pai(r['품목명'], r['규격명'], r['케이블종류']), axis=1)
    merged['케이블길이'] = merged.apply(lambda r: derive_length(r['품목명'], r['규격명']), axis=1)
    merged[['타입1', '타입2']] = merged['품목명'].apply(lambda x: pd.Series(extract_types(x)))
    return merged.sort_values('품목코드').reset_index(drop=True)

C1H = lambda YY: ['품목코드','품목명','규격명','케이블 종류','파이','코어수','케이블 길이','타입 1','타입2'] + \
                  [f'{YY}년{m:02d}월' for m in range(1,13)] + ['케이블 사용량','최고제작량','최고판매 케이블 소요량']
C2H = lambda YY: ['품목코드','품목명','규격명','케이블 종류','파이','코어수','케이블 길이','타입 1','타입2',
    *[f'{YY}년{m:02d}월' for m in range(1,13)],'합계',
    '2.0MM - LC/PC(청색 하우징 키트)','2.0MM - LC/PC(적색 하우징 키트)',
    '2.0MM - LC/APC(녹색 하우징 키트)','2.0MM - LC/APC(적색 하우징 키트)',
    '2.0MM - SC/PC(청) 하우징키트','2.0MM - SC/PC(적 부트) 하우징키트',
    '2.0MM - SC/APC(녹)하우징키트','2.0MM - SC/APC(적)하우징키트',
    '2.0MM - FC/PC(흑색 하우징키트)','2.0MM - FC/PC(적색 하우징키트)',
    '2.0MM - FC/APC(녹색 하우징 키트)','2.0MM - FC/APC(적색 하우징 키트)',
    '2.0MM - OJC LC/PC-MM 2.0 BEIGE',
    '3.0MM - LC/PC(청색 하우징 키트)','3.0MM - LC/PC(적색 하우징 키트)',
    '3.0MM - LC/APC(녹색 하우징 키트)','3.0MM - LC/APC(적색 하우징 키트)',
    '3.0MM - SC/PC(청) 하우징키트','3.0MM - SC/PC(적) 하우징키트',
    '3.0MM - SC/APC(녹)','3.0MM - SC/APC(적)',
    '3.0MM - FC/PC(흑색 하우징키트)','3.0MM - FC/PC(적색 하우징키트)',
    '3.0MM - FC/APC(녹색 하우징 키트)','3.0MM - FC/APC(적색 하우징 키트)',
    '검증(하우징 필요수량)','하우징 계산 수량']

def apply_style(ws, merged, nc):
    hf   = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    hfill= PatternFill('solid', start_color='305496')
    dfnt = Font(name='맑은 고딕', size=10)
    thin = Side(border_style='thin', color='BFBFBF')
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    cen  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la   = Alignment(horizontal='left', vertical='center')
    for c in range(1, nc+1):
        x = ws.cell(1, c); x.font=hf; x.fill=hfill; x.alignment=cen; x.border=bd
    for r in range(2, len(merged)+2):
        for c in range(1, nc+1):
            x = ws.cell(r, c); x.font=dfnt; x.border=bd
            x.alignment = la if c == 2 else cen
    ws.row_dimensions[1].height = 40; ws.freeze_panes = 'D2'
    if nc == 24:
        w = {1:16,2:40,3:22,4:22,5:8,6:7,7:10,8:10,9:10,22:14,23:12,24:20}
        for c in range(10, 22): w[c] = 10
    else:
        w = {1:16,2:40,3:22,4:22,5:8,6:7,7:10,8:10,9:10,22:10,48:14,49:14}
        for c in range(10, 22): w[c] = 10
        for c in range(23, 48): w[c] = 12
    for c, ww in w.items(): ws.column_dimensions[get_column_letter(c)].width = ww

def write_sheets(wb_out, merged, YY):
    ws_c = wb_out.create_sheet(f'{YY}년_케이블')
    ws_h = wb_out.create_sheet(f'{YY}년 하우징')
    for i, h in enumerate(C1H(YY), 1): ws_c.cell(1, i, h)
    for i, h in enumerate(C2H(YY), 1): ws_h.cell(1, i, h)
    for idx in range(len(merged)):
        r = idx + 2; row = merged.iloc[idx]
        for ws in [ws_c, ws_h]:
            ws.cell(r,1,row['품목코드']); ws.cell(r,2,row['품목명']); ws.cell(r,3,row['규격명'])
            ws.cell(r,4,row['케이블종류']); ws.cell(r,5,row['파이']); ws.cell(r,6,row['코어수'])
            ws.cell(r,7,row['케이블길이'] if pd.notna(row['케이블길이']) else None)
            ws.cell(r,8,row['타입1'] or None); ws.cell(r,9,row['타입2'] or None)
            for mm in range(1, 13):
                v = row[mm] if mm in row else 0; ws.cell(r, 9+mm, int(v) if v else None)
        ws_c.cell(r,22,f"=SUM(J{r}:U{r})*G{r}"); ws_c.cell(r,23,f"=MAX(J{r}:U{r})"); ws_c.cell(r,24,f"=W{r}*G{r}")
        ws_h.cell(r,22,f"=SUM(J{r}:U{r})")
        for col, val in calc_housing(row).items(): ws_h.cell(r, col, val)
        ws_h.cell(r,48,f"=IF(F{r}=1,V{r}*2,IF(F{r}=2,V{r}*F{r}*2,IF(F{r}>2,F{r}*2*V{r})))")
        ws_h.cell(r,49,f"=SUM(W{r}:AU{r})")
    apply_style(ws_c, merged, 24)
    apply_style(ws_h, merged, 49)

def parse_date(d):
    s = str(d).strip()
    m = re.match(r'^(\d{4})(\d{2})\d{2}', s)
    if m: return m.group(1), int(m.group(2))
    m = re.match(r'^(\d{2})/(\d{2})/\d{2}', s)
    if m: return f'20{m.group(1)}', int(m.group(2))
    return None, None

def detect_file_type(path):
    """구매현황 vs 통합 파일 자동 감지"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    if '구매현황' in sheets:
        return 'purchase'
    else:
        return 'integrated'

def run_conversion(input_path, output_path, log_fn):
    wb_in = load_workbook(input_path, data_only=True)
    file_type = detect_file_type(input_path)
    wb_out = Workbook(); wb_out.remove(wb_out.active)

    if file_type == 'purchase':
        # 구매현황 파일
        log_fn("📂 구매현황 파일 감지")
        ws_in = wb_in['구매현황']
        rows = []
        for r in range(3, ws_in.max_row + 1):
            buy_no = ws_in.cell(r,1).value; pc = ws_in.cell(r,5).value
            pr = ws_in.cell(r,6).value; qty = ws_in.cell(r,7).value
            if not buy_no or not pr or not qty: continue
            m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})', str(buy_no))
            if not m: continue
            year, month = m.group(1), int(m.group(2))
            s = re.sub(r'\s*외\s*\d+건\s*$', '', str(pr).strip())
            bm = re.match(r'^(.+?)\s*\[(.+)\]\s*$', s)
            pn = bm.group(1).strip() if bm else s
            gy = bm.group(2).strip() if bm else ''
            rows.append({'연도': year, '월': month, '품목코드': str(pc).strip() if pc else '',
                         '품목명': pn, '규격명': gy, '수량': int(qty)})
        df = pd.DataFrame(rows)
        for year in sorted(df['연도'].unique()):
            YY = year[-2:]
            dfY = df[df['연도'] == year].copy()
            merged = build_merged(dfY)
            write_sheets(wb_out, merged, YY)
            log_fn(f"  ✅ {year}년: {len(merged)}개 품목")

    else:
        # 통합 파일 (다년도)
        log_fn("📂 통합 파일 감지 (다년도)")
        sheet_name = wb_in.sheetnames[0]
        ws_in = wb_in[sheet_name]
        rows = []
        for r in range(2, ws_in.max_row + 1):
            date_raw = ws_in.cell(r,2).value; pc = ws_in.cell(r,4).value
            pn_raw = ws_in.cell(r,5).value; gy_raw = ws_in.cell(r,7).value
            qty = ws_in.cell(r,8).value
            if not date_raw or not pn_raw or not qty: continue
            year, month = parse_date(date_raw)
            if not year: continue
            pn = str(pn_raw).strip()
            gy = str(gy_raw).strip() if gy_raw and str(gy_raw).strip() not in ('', '\xa0') else ''
            try: q = int(float(str(qty)))
            except: continue
            if q <= 0: continue
            rows.append({'연도': year, '월': month, '품목코드': str(pc).strip() if pc else '',
                         '품목명': pn, '규격명': gy, '수량': q})
        df = pd.DataFrame(rows)
        for year in sorted(df['연도'].unique()):
            YY = year[-2:]
            dfY = df[df['연도'] == year].copy()
            merged = build_merged(dfY)
            write_sheets(wb_out, merged, YY)
            log_fn(f"  ✅ {year}년: {len(merged)}개 품목")

    wb_out.save(output_path)
    log_fn(f"\n✅ 저장 완료: {output_path}")
    log_fn(f"시트: {', '.join(wb_out.sheetnames)}")

# ============================================================
# GUI
# ============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("AJW 구매조회 → 가공파일 변환기")
        self.geometry("640x480")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self._build_ui()

    def _build_ui(self):
        # 헤더
        hdr = tk.Frame(self, bg="#1F4E79", height=60)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="AJW  구매조회 → 가공파일 변환기",
                 bg="#1F4E79", fg="white", font=("맑은 고딕", 14, "bold")).pack(pady=15)

        body = tk.Frame(self, bg="#F0F4F8", padx=24, pady=20)
        body.pack(fill=tk.BOTH, expand=True)

        # 입력 파일
        tk.Label(body, text="📂  입력 파일 (구매현황 또는 통합 xlsx)",
                 bg="#F0F4F8", font=("맑은 고딕", 10, "bold"), anchor="w").pack(fill=tk.X)
        row1 = tk.Frame(body, bg="#F0F4F8"); row1.pack(fill=tk.X, pady=(4, 12))
        self.input_var = tk.StringVar()
        tk.Entry(row1, textvariable=self.input_var, font=("맑은 고딕", 10),
                 width=52).pack(side=tk.LEFT)
        tk.Button(row1, text="찾아보기", command=self._browse_input,
                  bg="#2E75B6", fg="white", font=("맑은 고딕", 9),
                  relief=tk.FLAT, padx=10).pack(side=tk.LEFT, padx=(6, 0))

        # 출력 파일
        tk.Label(body, text="💾  출력 파일명",
                 bg="#F0F4F8", font=("맑은 고딕", 10, "bold"), anchor="w").pack(fill=tk.X)
        row2 = tk.Frame(body, bg="#F0F4F8"); row2.pack(fill=tk.X, pady=(4, 20))
        self.output_var = tk.StringVar()
        tk.Entry(row2, textvariable=self.output_var, font=("맑은 고딕", 10),
                 width=52).pack(side=tk.LEFT)
        tk.Button(row2, text="저장 위치", command=self._browse_output,
                  bg="#2E75B6", fg="white", font=("맑은 고딕", 9),
                  relief=tk.FLAT, padx=10).pack(side=tk.LEFT, padx=(6, 0))

        # 변환 버튼
        self.btn = tk.Button(body, text="▶  변환 시작", command=self._start,
                             bg="#1F4E79", fg="white", font=("맑은 고딕", 12, "bold"),
                             relief=tk.FLAT, pady=8, cursor="hand2")
        self.btn.pack(fill=tk.X, pady=(0, 12))

        # 진행 상황
        tk.Label(body, text="진행 상황", bg="#F0F4F8",
                 font=("맑은 고딕", 9, "bold"), anchor="w").pack(fill=tk.X)
        log_frame = tk.Frame(body, bg="#F0F4F8")
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log = tk.Text(log_frame, height=10, font=("Consolas", 9),
                           bg="#FFFFFF", relief=tk.SOLID, bd=1, wrap=tk.WORD)
        sb = tk.Scrollbar(log_frame, command=self.log.yview)
        self.log.configure(yscrollcommand=sb.set)
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="구매현황 또는 통합 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")])
        if path:
            self.input_var.set(path)
            # 출력 파일명 자동 제안
            base = os.path.splitext(path)[0]
            self.output_var.set(base + "_가공.xlsx")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="출력 파일 저장 위치",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")])
        if path:
            self.output_var.set(path)

    def _log(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def _start(self):
        input_path  = self.input_var.get().strip()
        output_path = self.output_var.get().strip()
        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("오류", "입력 파일을 선택해주세요."); return
        if not output_path:
            messagebox.showerror("오류", "출력 파일명을 입력해주세요."); return

        self.btn.config(state=tk.DISABLED, text="변환 중...")
        self.log.delete("1.0", tk.END)
        self._log(f"입력: {input_path}")
        self._log(f"출력: {output_path}")
        self._log("-" * 50)

        def job():
            try:
                run_conversion(input_path, output_path, self._log)
                self.btn.config(state=tk.NORMAL, text="▶  변환 시작")
                messagebox.showinfo("완료", f"변환이 완료됐습니다!\n\n{output_path}")
            except Exception as e:
                self._log(f"\n❌ 오류: {e}")
                self.btn.config(state=tk.NORMAL, text="▶  변환 시작")
                messagebox.showerror("오류", str(e))

        threading.Thread(target=job, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
